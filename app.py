from flask import Flask, request, redirect, url_for, session, render_template_string, jsonify, send_file, send_from_directory
import os
import math
import sqlite3
import datetime
import pytz
import uuid
from datetime import datetime, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import threading
import time

app = Flask(__name__)
app.secret_key = 'your-secret-key-here-change-it-to-something-secure-123456789'

# ============================================================
# TIME & UTILITY FUNCTIONS (ត្រូវដាក់ពីលើគេបង្អស់)
# ============================================================

def get_current_date():
    tz = pytz.timezone('Asia/Phnom_Penh')
    return datetime.now(tz).strftime('%Y-%m-%d')

def get_current_datetime():
    tz = pytz.timezone('Asia/Phnom_Penh')
    return datetime.now(tz)

def get_current_time():
    tz = pytz.timezone('Asia/Phnom_Penh')
    return datetime.now(tz).strftime('%Y-%m-%d %H:%M:%S')

def get_current_time_only():
    tz = pytz.timezone('Asia/Phnom_Penh')
    return datetime.now(tz).strftime('%H:%M:%S')

# ============================================================
# BACKGROUND AUTO-LOCK/UNLOCK THREAD
# ============================================================

def auto_lock_unlock_checker():
    """Background thread to automatically lock/unlock users based on deadlines"""
    while True:
        try:
            current_time = get_current_time_only()  # HH:MM:SS
            current_hhmm = current_time[:5]  # HH:MM

            # ===== Check System Lock (Auto-unlock) =====
            lock = get_system_lock_status()
            if lock.get('is_locked', 0) == 1:
                auto_unlock = lock.get('auto_unlock_time')
                if auto_unlock and current_hhmm >= auto_unlock:
                    print(f"🔄 Auto-unlocking system at {current_hhmm} (scheduled: {auto_unlock})")
                    update_system_lock(0, locked_by=None)
                    increment_data_version()
                    print("✅ System auto-unlocked successfully!")

            # ===== Auto-lock users based on attendance deadline =====
            conn = get_db_connection()
            active_settings = conn.execute('''
                SELECT user_id, check_in_deadline
                FROM attendance_settings
                WHERE is_active = 1
                AND check_in_deadline IS NOT NULL
                AND check_in_deadline != ''
            ''').fetchall()
            conn.close()

            for setting in active_settings:
                user_id = setting['user_id']
                deadline = setting['check_in_deadline']
                
                if deadline and current_hhmm >= deadline:
                    user_lock = get_user_lock_status(user_id)
                    if user_lock.get('is_locked', 0) != 1:
                        print(f"🔒 Auto-locking user {user_id} at {current_hhmm} (deadline: {deadline})")
                        system_lock = get_system_lock_status()
                        auto_unlock = system_lock.get('auto_unlock_time', '06:00')
                        update_user_lock(user_id, 1, auto_unlock_time=auto_unlock)
                        increment_data_version()
                        print(f"✅ User {user_id} auto-locked successfully!")

            # ===== Auto-unlock users based on auto_unlock_time =====
            conn = get_db_connection()
            locked_users = conn.execute('''
                SELECT user_id, auto_unlock_time
                FROM user_attendance_lock
                WHERE is_locked = 1
                AND auto_unlock_time IS NOT NULL
                AND auto_unlock_time != ''
            ''').fetchall()
            conn.close()

            for user in locked_users:
                auto_unlock = user['auto_unlock_time']
                if auto_unlock and current_hhmm >= auto_unlock:
                    user_id = user['user_id']
                    print(f"🔄 Auto-unlocking user {user_id} at {current_hhmm} (scheduled: {auto_unlock})")
                    update_user_lock(user_id, 0)
                    increment_data_version()
                    print(f"✅ User {user_id} auto-unlocked successfully!")

        except Exception as e:
            print(f"❌ Error in auto-lock/unlock checker: {e}")

        time.sleep(30)

def start_auto_lock_unlock_thread():
    thread = threading.Thread(target=auto_lock_unlock_checker, daemon=True)
    thread.start()
    print("✅ Auto-lock/unlock background thread started!")

start_auto_lock_unlock_thread()

# ===== UPLOAD CONFIGURATION =====
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024
app.config['ALLOWED_DISTANCE'] = 150

UPLOAD_FOLDER_LEAVES = os.path.join('static', 'uploads', 'leaves')
UPLOAD_FOLDER_MISSIONS = os.path.join('static', 'uploads', 'missions')
os.makedirs(UPLOAD_FOLDER_LEAVES, exist_ok=True)
os.makedirs(UPLOAD_FOLDER_MISSIONS, exist_ok=True)

DB_NAME = 'employees.db'

def get_db_connection():
    conn = sqlite3.connect(DB_NAME, timeout=20)
    conn.row_factory = sqlite3.Row
    return conn

def table_exists(table_name):
    conn = get_db_connection()
    result = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
        (table_name,)
    ).fetchone()
    conn.close()
    return result is not None

def column_exists(table_name, column_name):
    conn = get_db_connection()
    try:
        conn.execute(f"SELECT {column_name} FROM {table_name} LIMIT 1")
        exists = True
    except sqlite3.OperationalError:
        exists = False
    conn.close()
    return exists

def migrate_database():
    conn = get_db_connection()

    if not table_exists('attendance_settings'):
        conn.execute('''
            CREATE TABLE IF NOT EXISTS attendance_settings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                check_in_deadline TEXT,
                is_active INTEGER DEFAULT 0,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES users(id),
                UNIQUE(user_id)
            )
        ''')
        conn.commit()
        print("✅ Created table attendance_settings")

    if table_exists('attendance_settings'):
        if not column_exists('attendance_settings', 'check_in_deadline'):
            try:
                conn.execute("ALTER TABLE attendance_settings ADD COLUMN check_in_deadline TEXT")
                print("✅ Added column check_in_deadline")
            except sqlite3.OperationalError as e:
                print(f"⚠️ Could not add check_in_deadline: {e}")

        if not column_exists('attendance_settings', 'is_active'):
            try:
                conn.execute("ALTER TABLE attendance_settings ADD COLUMN is_active INTEGER DEFAULT 0")
                print("✅ Added column is_active")
            except sqlite3.OperationalError as e:
                print(f"⚠️ Could not add is_active: {e}")

    if not table_exists('user_attendance_lock'):
        conn.execute('''
            CREATE TABLE IF NOT EXISTS user_attendance_lock (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                is_locked INTEGER DEFAULT 0,
                auto_unlock_time TEXT,
                locked_by INTEGER,
                lock_start_time TEXT,
                lock_end_time TEXT,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES users(id),
                UNIQUE(user_id)
            )
        ''')
        conn.commit()
        print("✅ Created table user_attendance_lock")

    if not table_exists('system_attendance_lock'):
        conn.execute('''
            CREATE TABLE IF NOT EXISTS system_attendance_lock (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                is_locked INTEGER DEFAULT 0,
                lock_start_time TEXT,
                lock_end_time TEXT,
                auto_unlock_time TEXT,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                locked_by INTEGER,
                FOREIGN KEY (locked_by) REFERENCES users(id)
            )
        ''')
        conn.commit()
        print("✅ Created table system_attendance_lock")

        conn.execute('''
            INSERT INTO system_attendance_lock (is_locked, auto_unlock_time)
            VALUES (0, '06:00')
        ''')
        conn.commit()
        print("✅ Inserted default system lock record")

    if not table_exists('attendance'):
        conn.execute('''
            CREATE TABLE IF NOT EXISTS attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                date TEXT NOT NULL,
                check_in TEXT,
                check_out TEXT,
                total_hours REAL DEFAULT 0,
                shift INTEGER DEFAULT 1,
                check_in_lat REAL,
                check_in_lng REAL,
                check_out_lat REAL,
                check_out_lng REAL,
                check_in_distance REAL,
                check_out_distance REAL,
                FOREIGN KEY (user_id) REFERENCES users(id)
            )
        ''')
        conn.commit()
        print("✅ Created table attendance")
        conn.close()
        return

    columns_to_add = [
        ('check_in_lat', 'REAL'),
        ('check_in_lng', 'REAL'),
        ('check_out_lat', 'REAL'),
        ('check_out_lng', 'REAL'),
        ('check_in_distance', 'REAL'),
        ('check_out_distance', 'REAL'),
        ('shift', 'INTEGER DEFAULT 1')
    ]

    for col_name, col_type in columns_to_add:
        if not column_exists('attendance', col_name):
            try:
                conn.execute(f"ALTER TABLE attendance ADD COLUMN {col_name} {col_type}")
                print(f"✅ Added column {col_name}")
            except sqlite3.OperationalError as e:
                print(f"⚠️ Could not add {col_name}: {e}")

    if table_exists('leaves') and not column_exists('leaves', 'attachment'):
        try:
            conn.execute("ALTER TABLE leaves ADD COLUMN attachment TEXT")
            print("✅ Added column attachment to leaves table")
        except sqlite3.OperationalError as e:
            print(f"⚠️ Could not add attachment column: {e}")

    if table_exists('leaves') and not column_exists('leaves', 'admin_id'):
        try:
            conn.execute("ALTER TABLE leaves ADD COLUMN admin_id INTEGER")
            print("✅ Added column admin_id to leaves table")
        except sqlite3.OperationalError as e:
            print(f"⚠️ Could not add admin_id column to leaves: {e}")

    if table_exists('missions') and not column_exists('missions', 'attachment'):
        try:
            conn.execute("ALTER TABLE missions ADD COLUMN attachment TEXT")
            print("✅ Added column attachment to missions table")
        except sqlite3.OperationalError as e:
            print(f"⚠️ Could not add attachment column: {e}")

    if table_exists('missions') and not column_exists('missions', 'admin_id'):
        try:
            conn.execute("ALTER TABLE missions ADD COLUMN admin_id INTEGER")
            print("✅ Added column admin_id to missions table")
        except sqlite3.OperationalError as e:
            print(f"⚠️ Could not add admin_id column to missions: {e}")

    conn.commit()
    conn.close()

def init_db():
    conn = get_db_connection()

    conn.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            full_name TEXT NOT NULL,
            email TEXT,
            phone TEXT,
            role TEXT DEFAULT 'user',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            check_in TEXT,
            check_out TEXT,
            total_hours REAL DEFAULT 0,
            shift INTEGER DEFAULT 1,
            check_in_lat REAL,
            check_in_lng REAL,
            check_out_lat REAL,
            check_out_lng REAL,
            check_in_distance REAL,
            check_out_distance REAL,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS leaves (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            start_date TEXT NOT NULL,
            end_date TEXT NOT NULL,
            days REAL NOT NULL,
            reason TEXT,
            attachment TEXT,
            status TEXT DEFAULT 'pending',
            admin_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id),
            FOREIGN KEY (admin_id) REFERENCES users(id)
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS missions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            start_date TEXT NOT NULL,
            end_date TEXT NOT NULL,
            days REAL NOT NULL,
            destination TEXT,
            purpose TEXT,
            attachment TEXT,
            status TEXT DEFAULT 'pending',
            admin_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id),
            FOREIGN KEY (admin_id) REFERENCES users(id)
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS company_location (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            lat REAL NOT NULL,
            lng REAL NOT NULL,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS data_version (
            version INTEGER DEFAULT 1
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS attendance_settings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            check_in_deadline TEXT,
            is_active INTEGER DEFAULT 0,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id),
            UNIQUE(user_id)
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS system_attendance_lock (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            is_locked INTEGER DEFAULT 0,
            lock_start_time TEXT,
            lock_end_time TEXT,
            auto_unlock_time TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            locked_by INTEGER,
            FOREIGN KEY (locked_by) REFERENCES users(id)
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS user_attendance_lock (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            is_locked INTEGER DEFAULT 0,
            auto_unlock_time TEXT,
            locked_by INTEGER,
            lock_start_time TEXT,
            lock_end_time TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id),
            UNIQUE(user_id)
        )
    ''')

    admin = conn.execute("SELECT * FROM users WHERE username = 'admin'").fetchone()
    if not admin:
        conn.execute(
            "INSERT INTO users (username, password, full_name, role) VALUES (?, ?, ?, ?)",
            ('admin', '1234', 'អ្នកគ្រប់គ្រង', 'admin')
        )
        print("✅ Created admin user: admin / 1234")

    test_user = conn.execute("SELECT * FROM users WHERE username = 'user1'").fetchone()
    if not test_user:
        conn.execute(
            "INSERT INTO users (username, password, full_name, role) VALUES (?, ?, ?, ?)",
            ('user1', '1234', 'បុគ្គលិកសាកល្បង', 'user')
        )
        print("✅ Created test user: user1 / 1234")

    version = conn.execute("SELECT version FROM data_version").fetchone()
    if not version:
        conn.execute("INSERT INTO data_version (version) VALUES (1)")

    lock = conn.execute("SELECT * FROM system_attendance_lock LIMIT 1").fetchone()
    if not lock:
        conn.execute('''
            INSERT INTO system_attendance_lock (is_locked, auto_unlock_time)
            VALUES (0, '06:00')
        ''')
        print("✅ Inserted default system lock record")

    conn.commit()
    conn.close()

    migrate_database()
    print(f"✅ Database '{DB_NAME}' initialized successfully!")

# ============================================================
# SYSTEM ATTENDANCE LOCK FUNCTIONS
# ============================================================

def get_system_lock_status():
    try:
        conn = get_db_connection()
        result = conn.execute("SELECT * FROM system_attendance_lock ORDER BY id DESC LIMIT 1").fetchone()
        conn.close()
        return dict(result) if result else {'is_locked': 0, 'auto_unlock_time': '06:00'}
    except Exception as e:
        print(f"Error getting lock status: {e}")
        return {'is_locked': 0, 'auto_unlock_time': '06:00'}

def update_system_lock(is_locked, lock_start_time=None, lock_end_time=None, auto_unlock_time=None, locked_by=None):
    try:
        conn = get_db_connection()
        current = conn.execute("SELECT * FROM system_attendance_lock ORDER BY id DESC LIMIT 1").fetchone()

        if current:
            conn.execute('''
                UPDATE system_attendance_lock
                SET is_locked = ?,
                    lock_start_time = COALESCE(?, lock_start_time),
                    lock_end_time = COALESCE(?, lock_end_time),
                    auto_unlock_time = COALESCE(?, auto_unlock_time),
                    locked_by = COALESCE(?, locked_by),
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (is_locked, lock_start_time, lock_end_time, auto_unlock_time, locked_by, current['id']))
        else:
            conn.execute('''
                INSERT INTO system_attendance_lock
                (is_locked, lock_start_time, lock_end_time, auto_unlock_time, locked_by)
                VALUES (?, ?, ?, ?, ?)
            ''', (is_locked, lock_start_time, lock_end_time, auto_unlock_time, locked_by))

        conn.commit()
        conn.close()
        increment_data_version()
        return True
    except Exception as e:
        print(f"Error updating lock: {e}")
        return False

def check_system_lock_for_user(user_id):
    lock = get_system_lock_status()

    if lock.get('is_locked', 0) != 1:
        return True, None

    auto_unlock = lock.get('auto_unlock_time')
    if auto_unlock:
        current_time = get_current_time_only()
        current_hhmm = current_time[:5]
        if current_hhmm >= auto_unlock:
            print(f"🔄 Auto-unlocking system during check (current: {current_hhmm}, scheduled: {auto_unlock})")
            update_system_lock(0, locked_by=None)
            increment_data_version()
            return True, None

    return False, "⛔ ប្រព័ន្ធកំពុងបិទការចូលធ្វើការ! សូមរង់ចាំរហូតដល់ម៉ោងបើកដោយស្វ័យប្រវត្តិ ឬទាក់ទង Admin!"

def toggle_system_lock(lock_state, auto_unlock_time=None, locked_by=None):
    lock = get_system_lock_status()
    current_time = get_current_time()

    if lock_state == 1:
        return update_system_lock(
            is_locked=1,
            lock_start_time=current_time,
            auto_unlock_time=auto_unlock_time or lock.get('auto_unlock_time', '06:00'),
            locked_by=locked_by
        )
    else:
        return update_system_lock(
            is_locked=0,
            lock_end_time=current_time,
            locked_by=locked_by
        )

# ============================================================
# USER ATTENDANCE LOCK FUNCTIONS
# ============================================================

def get_user_lock_status(user_id):
    try:
        conn = get_db_connection()
        result = conn.execute(
            "SELECT * FROM user_attendance_lock WHERE user_id = ?",
            (user_id,)
        ).fetchone()
        conn.close()
        return dict(result) if result else {'is_locked': 0, 'auto_unlock_time': None}
    except Exception as e:
        print(f"Error getting user lock status: {e}")
        return {'is_locked': 0, 'auto_unlock_time': None}

def update_user_lock(user_id, is_locked, auto_unlock_time=None, locked_by=None):
    try:
        conn = get_db_connection()
        current_time = get_current_time()

        existing = conn.execute(
            "SELECT id FROM user_attendance_lock WHERE user_id = ?",
            (user_id,)
        ).fetchone()

        if existing:
            if is_locked == 1:
                conn.execute('''
                    UPDATE user_attendance_lock
                    SET is_locked = ?,
                        auto_unlock_time = ?,
                        locked_by = ?,
                        lock_start_time = ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE user_id = ?
                ''', (is_locked, auto_unlock_time, locked_by, current_time, user_id))
            else:
                conn.execute('''
                    UPDATE user_attendance_lock
                    SET is_locked = ?,
                        lock_end_time = ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE user_id = ?
                ''', (is_locked, current_time, user_id))
        else:
            conn.execute('''
                INSERT INTO user_attendance_lock
                (user_id, is_locked, auto_unlock_time, locked_by, lock_start_time)
                VALUES (?, ?, ?, ?, ?)
            ''', (user_id, is_locked, auto_unlock_time, locked_by, current_time))

        conn.commit()
        conn.close()
        increment_data_version()
        return True
    except Exception as e:
        print(f"Error updating user lock: {e}")
        return False

def check_user_lock(user_id):
    lock = get_user_lock_status(user_id)

    if lock.get('is_locked', 0) != 1:
        return True, None

    auto_unlock = lock.get('auto_unlock_time')
    if auto_unlock:
        current_time = get_current_time_only()
        current_hhmm = current_time[:5]
        if current_hhmm >= auto_unlock:
            print(f"🔄 Auto-unlocking user {user_id} during check (current: {current_hhmm}, scheduled: {auto_unlock})")
            update_user_lock(user_id, 0)
            increment_data_version()
            return True, None

    return False, "⛔ អ្នកត្រូវបានបិទការចូលធ្វើការដោយ Admin! សូមទាក់ទង Admin!"

def get_all_user_lock_status():
    try:
        conn = get_db_connection()
        results = conn.execute('''
            SELECT u.id, u.username, u.full_name,
                   l.is_locked, l.auto_unlock_time, l.lock_start_time, l.updated_at
            FROM users u
            LEFT JOIN user_attendance_lock l ON u.id = l.user_id
            WHERE u.role != 'admin'
            ORDER BY u.full_name
        ''').fetchall()
        conn.close()
        return [dict(row) for row in results]
    except Exception as e:
        print(f"Error getting all user locks: {e}")
        return []

def toggle_user_lock(user_id, lock_state, auto_unlock_time=None, locked_by=None):
    if lock_state == 1:
        return update_user_lock(user_id, 1, auto_unlock_time, locked_by)
    else:
        return update_user_lock(user_id, 0, locked_by=locked_by)

# ============================================================
# DATA VERSION FUNCTIONS
# ============================================================

def get_data_version():
    try:
        conn = get_db_connection()
        result = conn.execute("SELECT version FROM data_version").fetchone()
        conn.close()
        return result['version'] if result else 1
    except sqlite3.OperationalError:
        init_db()
        return 1

def increment_data_version():
    try:
        conn = get_db_connection()
        conn.execute("UPDATE data_version SET version = version + 1")
        conn.commit()
        conn.close()
    except sqlite3.OperationalError:
        init_db()
        conn = get_db_connection()
        conn.execute("UPDATE data_version SET version = version + 1")
        conn.commit()
        conn.close()

# ============================================================
# USER FUNCTIONS
# ============================================================

def get_user_by_id(user_id):
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    conn.close()
    return user

def get_user_by_username(username):
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
    conn.close()
    return user

def create_user(username, password, full_name, email=None, phone=None, role='user'):
    conn = get_db_connection()
    try:
        conn.execute(
            "INSERT INTO users (username, password, full_name, email, phone, role) VALUES (?, ?, ?, ?, ?, ?)",
            (username, password, full_name, email, phone, role)
        )
        conn.commit()
        conn.close()
        increment_data_version()
        return True
    except sqlite3.IntegrityError:
        conn.close()
        return False

def update_user(user_id, username, full_name, email, phone, role):
    conn = get_db_connection()
    try:
        conn.execute(
            "UPDATE users SET username = ?, full_name = ?, email = ?, phone = ?, role = ? WHERE id = ?",
            (username, full_name, email, phone, role, user_id)
        )
        conn.commit()
        conn.close()
        increment_data_version()
        return True
    except sqlite3.IntegrityError:
        conn.close()
        return False

def update_password(user_id, new_password):
    conn = get_db_connection()
    conn.execute("UPDATE users SET password = ? WHERE id = ?", (new_password, user_id))
    conn.commit()
    conn.close()
    increment_data_version()
    return True

def verify_password(user_id, password):
    conn = get_db_connection()
    user = conn.execute("SELECT password FROM users WHERE id = ?", (user_id,)).fetchone()
    conn.close()
    return user and user['password'] == password

def delete_user(user_id):
    conn = get_db_connection()
    try:
        conn.execute("DELETE FROM attendance WHERE user_id = ?", (user_id,))
        conn.execute("DELETE FROM leaves WHERE user_id = ?", (user_id,))
        conn.execute("DELETE FROM missions WHERE user_id = ?", (user_id,))
        conn.execute("DELETE FROM attendance_settings WHERE user_id = ?", (user_id,))
        conn.execute("DELETE FROM user_attendance_lock WHERE user_id = ?", (user_id,))
        conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
        conn.commit()
        conn.close()
        increment_data_version()
        return True
    except Exception:
        conn.close()
        return False

def get_all_users():
    conn = get_db_connection()
    users = conn.execute("SELECT * FROM users ORDER BY id DESC").fetchall()
    conn.close()
    return users

# ============================================================
# ATTENDANCE SETTINGS FUNCTIONS
# ============================================================

def get_attendance_setting(user_id):
    conn = get_db_connection()
    setting = conn.execute(
        "SELECT * FROM attendance_settings WHERE user_id = ?",
        (user_id,)
    ).fetchone()
    conn.close()
    return dict(setting) if setting else None

def get_all_attendance_settings():
    conn = get_db_connection()
    settings = conn.execute('''
        SELECT s.*, u.username, u.full_name
        FROM attendance_settings s
        JOIN users u ON s.user_id = u.id
        ORDER BY u.full_name
    ''').fetchall()
    conn.close()
    return [dict(row) for row in settings]

def save_attendance_setting(user_id, check_in_deadline, is_active):
    conn = get_db_connection()
    existing = conn.execute(
        "SELECT id FROM attendance_settings WHERE user_id = ?",
        (user_id,)
    ).fetchone()

    if existing:
        conn.execute(
            "UPDATE attendance_settings SET check_in_deadline = ?, is_active = ?, updated_at = CURRENT_TIMESTAMP WHERE user_id = ?",
            (check_in_deadline, is_active, user_id)
        )
    else:
        conn.execute(
            "INSERT INTO attendance_settings (user_id, check_in_deadline, is_active) VALUES (?, ?, ?)",
            (user_id, check_in_deadline, is_active)
        )
    conn.commit()
    conn.close()
    increment_data_version()
    return True

def check_attendance_deadline(user_id):
    setting = get_attendance_setting(user_id)
    if not setting:
        return True, None

    if setting.get('is_active') != 1:
        return True, None

    deadline = setting.get('check_in_deadline')
    if not deadline:
        return True, None

    current_time = get_current_time_only()
    current_hhmm = current_time[:5]

    if current_hhmm > deadline:
        return False, deadline
    return True, deadline

# ============================================================
# ATTENDANCE FUNCTIONS
# ============================================================

def get_company_location():
    conn = get_db_connection()
    location = conn.execute("SELECT lat, lng FROM company_location ORDER BY id DESC LIMIT 1").fetchone()
    conn.close()
    return dict(location) if location else None

def save_company_location(lat, lng):
    conn = get_db_connection()
    conn.execute("INSERT INTO company_location (lat, lng) VALUES (?, ?)", (lat, lng))
    conn.commit()
    conn.close()
    increment_data_version()
    return True

def check_in(user_id, lat, lng, distance, shift):
    date = get_current_date()
    check_in_time = get_current_time()
    conn = get_db_connection()

    existing = conn.execute(
        """SELECT id FROM attendance
           WHERE user_id = ?
           AND check_out IS NULL""",
        (user_id,)
    ).fetchone()

    if existing:
        conn.close()
        return False, "អ្នកបានចូលធ្វើការរួចហើយ! សូមចុច 'ចេញពីធ្វើការ' មុនពេលចូលម្តងទៀត!"

    can_checkin, deadline = check_attendance_deadline(user_id)
    if not can_checkin:
        return False, f"⛔ អ្នកលើសម៉ោងដែល Admin បានកំណត់ (ម៉ោងកំណត់: {deadline})! សូមទាក់ទងទៅអ្នកគ្រប់គ្រង!"

    conn.execute(
        """INSERT INTO attendance
           (user_id, date, check_in, shift, check_in_lat, check_in_lng, check_in_distance)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (user_id, date, check_in_time, shift, lat, lng, distance)
    )
    conn.commit()
    conn.close()
    increment_data_version()
    return True, "ចូលធ្វើការជោគជ័យ!"

def check_out(user_id, lat, lng, distance):
    check_out_time = get_current_datetime()
    conn = get_db_connection()

    record = conn.execute(
        """SELECT id, check_in, shift, date
           FROM attendance
           WHERE user_id = ?
           AND check_out IS NULL
           ORDER BY id DESC LIMIT 1""",
        (user_id,)
    ).fetchone()

    if not record:
        conn.close()
        return False, "មិនមានការចូលធ្វើការដែលមិនទាន់ចេញ!"

    check_in_time = datetime.strptime(record['check_in'], '%Y-%m-%d %H:%M:%S')
    tz = pytz.timezone('Asia/Phnom_Penh')
    check_in_time = tz.localize(check_in_time)

    if check_out_time <= check_in_time:
        check_out_time = check_out_time + timedelta(days=1)

    diff = check_out_time - check_in_time
    total_hours = diff.total_seconds() / 3600

    conn.execute(
        """UPDATE attendance
           SET check_out = ?, total_hours = ?, check_out_lat = ?, check_out_lng = ?, check_out_distance = ?
           WHERE id = ?""",
        (check_out_time.strftime('%Y-%m-%d %H:%M:%S'), total_hours, lat, lng, distance, record['id'])
    )
    conn.commit()
    conn.close()
    increment_data_version()

    shift_names = {1: 'វគ្គ 1 (ព្រឹក)', 2: 'វគ្គ 2 (រសៀល)', 3: 'វគ្គ 3 (យប់)'}
    shift = record['shift'] if 'shift' in record.keys() else 1

    hours = int(total_hours)
    minutes = int((total_hours - hours) * 60)
    return True, f"ចេញធ្វើការជោគជ័យ! ({shift_names.get(shift, '')}) ម៉ោងសរុប: {hours:02d}:{minutes:02d}"

def get_attendance_stats():
    today = get_current_date()
    conn = get_db_connection()

    total_users = conn.execute("SELECT COUNT(*) as count FROM users WHERE role != 'admin'").fetchone()['count']

    present_today = conn.execute(
        "SELECT COUNT(DISTINCT user_id) as count FROM attendance WHERE check_out IS NULL"
    ).fetchone()['count']

    leave_today = conn.execute(
        "SELECT COUNT(DISTINCT user_id) as count FROM leaves WHERE status = 'approved' AND ? BETWEEN start_date AND end_date",
        (today,)
    ).fetchone()['count']

    mission_today = conn.execute(
        "SELECT COUNT(DISTINCT user_id) as count FROM missions WHERE status = 'approved' AND ? BETWEEN start_date AND end_date",
        (today,)
    ).fetchone()['count']

    conn.close()

    return {
        'total_users': total_users,
        'present_today': present_today,
        'leave_today': leave_today,
        'mission_today': mission_today
    }

def get_checkin_status(user_id):
    conn = get_db_connection()
    record = conn.execute(
        """SELECT id, check_in, COALESCE(shift, 1) as shift
           FROM attendance
           WHERE user_id = ?
           AND check_out IS NULL
           ORDER BY id DESC LIMIT 1""",
        (user_id,)
    ).fetchone()
    conn.close()

    if record:
        return {
            'has_checkin': True,
            'check_in_time': record['check_in'],
            'shift': record['shift']
        }
    return {'has_checkin': False, 'check_in_time': None, 'shift': None}

# ============================================================
# DASHBOARD REPORT FUNCTIONS
# ============================================================

def get_work_history_report(start_date=None, end_date=None, limit=200):
    if not start_date:
        today = datetime.now()
        start_date = today.replace(day=1).strftime('%Y-%m-%d')
    if not end_date:
        end_date = datetime.now().strftime('%Y-%m-%d')

    conn = get_db_connection()

    attendance_records = conn.execute('''
        SELECT
            a.id,
            a.user_id,
            u.full_name,
            a.date,
            CASE
                WHEN a.shift = 1 THEN 'វគ្គ 1'
                WHEN a.shift = 2 THEN 'វគ្គ 2'
                WHEN a.shift = 3 THEN 'វគ្គ 3'
                ELSE 'វគ្គ 1'
            END as shift,
            a.check_in,
            a.check_out,
            a.total_hours,
            'attendance' as type,
            NULL as days,
            NULL as reason,
            NULL as destination,
            NULL as start_date,
            NULL as end_date,
            CASE
                WHEN a.check_in IS NOT NULL AND a.check_out IS NOT NULL THEN 'បានបិទ'
                WHEN a.check_in IS NOT NULL AND a.check_out IS NULL THEN 'កំពុងធ្វើការ'
                ELSE 'មិនទាន់ចូល'
            END as status,
            a.date as sort_date
        FROM attendance a
        JOIN users u ON a.user_id = u.id
        WHERE a.date BETWEEN ? AND ?
            AND a.check_in IS NOT NULL
    ''', (start_date, end_date)).fetchall()

    leave_records = conn.execute('''
        SELECT
            l.id,
            l.user_id,
            u.full_name,
            l.start_date as date,
            'ច្បាប់' as shift,
            NULL as check_in,
            NULL as check_out,
            NULL as total_hours,
            'leave' as type,
            l.days as days,
            l.reason as reason,
            NULL as destination,
            l.start_date,
            l.end_date,
            l.status,
            l.start_date as sort_date
        FROM leaves l
        JOIN users u ON l.user_id = u.id
        WHERE l.status = 'approved'
            AND l.start_date BETWEEN ? AND ?
    ''', (start_date, end_date)).fetchall()

    mission_records = conn.execute('''
        SELECT
            m.id,
            m.user_id,
            u.full_name,
            m.start_date as date,
            'បេសកម្ម' as shift,
            NULL as check_in,
            NULL as check_out,
            NULL as total_hours,
            'mission' as type,
            m.days as days,
            NULL as reason,
            m.destination as destination,
            m.start_date,
            m.end_date,
            m.status,
            m.start_date as sort_date
        FROM missions m
        JOIN users u ON m.user_id = u.id
        WHERE m.status = 'approved'
            AND m.start_date BETWEEN ? AND ?
    ''', (start_date, end_date)).fetchall()

    conn.close()

    result = []

    for row in attendance_records:
        item = dict(row)
        if item.get('total_hours'):
            hours = int(item['total_hours'])
            minutes = int((item['total_hours'] - hours) * 60)
            if item['total_hours'] < 0:
                hours = abs(hours)
                minutes = abs(minutes)
                item['total_hours_formatted'] = f"-{hours:02d}:{minutes:02d}"
            else:
                item['total_hours_formatted'] = f"{hours:02d}:{minutes:02d}"
            item['display_value'] = item['total_hours_formatted']
        else:
            item['total_hours_formatted'] = ''
            item['display_value'] = ''
        if item.get('check_in'):
            item['check_in_time'] = item['check_in'][11:16] if ' ' in item['check_in'] else item['check_in'][:5]
        else:
            item['check_in_time'] = ''
        if item.get('check_out'):
            item['check_out_time'] = item['check_out'][11:16] if ' ' in item['check_out'] else item['check_out'][:5]
        else:
            item['check_out_time'] = ''
        item['display_info'] = ''
        result.append(item)

    for row in leave_records:
        item = dict(row)
        item['shift'] = 'ច្បាប់'
        item['check_in_time'] = ''
        item['check_out_time'] = ''
        item['total_hours_formatted'] = ''
        if item.get('days'):
            item['display_value'] = f"{item['days']} ថ្ងៃ"
        else:
            item['display_value'] = ''
        item['display_info'] = f"មូលហេតុ: {item.get('reason', 'មិនបានបញ្ជាក់')}"
        item['status'] = 'បានអនុម័ត'
        result.append(item)

    for row in mission_records:
        item = dict(row)
        item['shift'] = 'បេសកម្ម'
        item['check_in_time'] = ''
        item['check_out_time'] = ''
        item['total_hours_formatted'] = ''
        if item.get('days'):
            item['display_value'] = f"{item['days']} ថ្ងៃ"
        else:
            item['display_value'] = ''
        item['display_info'] = f"ទីតាំង: {item.get('destination', 'មិនបានបញ្ជាក់')}"
        item['status'] = 'បានអនុម័ត'
        result.append(item)

    result.sort(key=lambda x: x['sort_date'] if x['sort_date'] else '', reverse=True)
    return result[:limit]

def get_monthly_summary_report(start_date, end_date):
    conn = get_db_connection()

    attendance_summary = conn.execute('''
        SELECT
            u.id,
            u.full_name,
            COUNT(DISTINCT a.date) as days_worked,
            COALESCE(SUM(a.total_hours), 0) as total_hours,
            COUNT(DISTINCT CASE WHEN a.shift = 3 THEN a.date END) as night_shifts,
            COALESCE(SUM(CASE WHEN a.shift = 3 THEN a.total_hours ELSE 0 END), 0) as night_hours
        FROM users u
        LEFT JOIN attendance a ON u.id = a.user_id
            AND a.date BETWEEN ? AND ?
            AND a.check_in IS NOT NULL
            AND a.check_out IS NOT NULL
        WHERE u.role != 'admin'
        GROUP BY u.id
        ORDER BY u.full_name
    ''', (start_date, end_date)).fetchall()

    leave_summary = conn.execute('''
        SELECT
            user_id,
            SUM(days) as total_leave_days
        FROM leaves
        WHERE status = 'approved'
            AND start_date BETWEEN ? AND ?
        GROUP BY user_id
    ''', (start_date, end_date)).fetchall()

    mission_summary = conn.execute('''
        SELECT
            user_id,
            SUM(days) as total_mission_days
        FROM missions
        WHERE status = 'approved'
            AND start_date BETWEEN ? AND ?
        GROUP BY user_id
    ''', (start_date, end_date)).fetchall()

    conn.close()

    leave_dict = {row['user_id']: row['total_leave_days'] for row in leave_summary}
    mission_dict = {row['user_id']: row['total_mission_days'] for row in mission_summary}

    result = []
    for emp in attendance_summary:
        item = dict(emp)
        item['total_leave_days'] = leave_dict.get(emp['id'], 0)
        item['total_mission_days'] = mission_dict.get(emp['id'], 0)

        total_hours = item['total_hours']
        if total_hours < 0:
            total_hours = abs(total_hours)
            sign = "-"
        else:
            sign = ""
        hours = int(total_hours)
        minutes = int((total_hours - hours) * 60)
        item['total_hours_formatted'] = f"{sign}{hours:02d}:{minutes:02d}"
        item['days_worked_display'] = f"{item['days_worked']} ថ្ងៃ"
        result.append(item)

    return result

# ============================================================
# ATTENDANCE FUNCTIONS (continued)
# ============================================================

def get_all_attendance(limit=100):
    conn = get_db_connection()

    attendances = conn.execute('''
        SELECT
            a.id,
            a.user_id,
            u.full_name,
            u.username,
            a.date,
            a.shift,
            a.check_in,
            a.check_out,
            a.total_hours,
            'attendance' as type,
            NULL as reason,
            NULL as destination,
            NULL as start_date,
            NULL as end_date,
            NULL as days,
            a.date as sort_date
        FROM attendance a
        JOIN users u ON a.user_id = u.id
        ORDER BY a.date DESC, a.id DESC
        LIMIT ?
    ''', (limit,)).fetchall()

    leaves = conn.execute('''
        SELECT
            l.id,
            l.user_id,
            u.full_name,
            u.username,
            l.start_date as date,
            NULL as shift,
            NULL as check_in,
            NULL as check_out,
            l.days as total_hours,
            'leave' as type,
            l.reason,
            NULL as destination,
            l.start_date,
            l.end_date,
            l.days,
            l.start_date as sort_date
        FROM leaves l
        JOIN users u ON l.user_id = u.id
        WHERE l.status = 'approved'
        ORDER BY l.start_date DESC, l.id DESC
        LIMIT ?
    ''', (limit,)).fetchall()

    missions = conn.execute('''
        SELECT
            m.id,
            m.user_id,
            u.full_name,
            u.username,
            m.start_date as date,
            NULL as shift,
            NULL as check_in,
            NULL as check_out,
            m.days as total_hours,
            'mission' as type,
            NULL as reason,
            m.destination,
            m.start_date,
            m.end_date,
            m.days,
            m.start_date as sort_date
        FROM missions m
        JOIN users u ON m.user_id = u.id
        WHERE m.status = 'approved'
        ORDER BY m.start_date DESC, m.id DESC
        LIMIT ?
    ''', (limit,)).fetchall()

    conn.close()

    all_data = []

    for row in attendances:
        all_data.append(dict(row))

    for row in leaves:
        all_data.append(dict(row))

    for row in missions:
        all_data.append(dict(row))

    all_data.sort(key=lambda x: x['sort_date'] if x['sort_date'] else '', reverse=True)

    return all_data[:limit]

def get_attendance_report(start_date, end_date):
    conn = get_db_connection()

    daily_report = conn.execute('''
        SELECT
            a.id,
            a.user_id,
            u.full_name,
            u.username,
            a.date,
            COALESCE(a.shift, 1) as shift,
            a.check_in,
            a.check_out,
            a.total_hours,
            'attendance' as type,
            NULL as reason,
            NULL as destination,
            NULL as start_date,
            NULL as end_date,
            NULL as days,
            CASE
                WHEN a.shift = 1 THEN 'វគ្គ 1'
                WHEN a.shift = 2 THEN 'វគ្គ 2'
                WHEN a.shift = 3 THEN 'វគ្គ 3'
                ELSE 'វគ្គ 1'
            END as shift_name,
            CASE
                WHEN a.check_in IS NOT NULL AND a.check_out IS NOT NULL THEN 'បានបិទ'
                WHEN a.check_in IS NOT NULL AND a.check_out IS NULL THEN 'កំពុងធ្វើការ'
                ELSE 'មិនទាន់ចូល'
            END as status
        FROM attendance a
        JOIN users u ON a.user_id = u.id
        WHERE a.date BETWEEN ? AND ?
        ORDER BY a.date DESC, a.user_id, a.shift
    ''', (start_date, end_date)).fetchall()

    leave_report = conn.execute('''
        SELECT
            l.id,
            l.user_id,
            u.full_name,
            u.username,
            l.start_date as date,
            NULL as shift,
            NULL as check_in,
            NULL as check_out,
            l.days as total_hours,
            'leave' as type,
            l.reason,
            NULL as destination,
            l.start_date,
            l.end_date,
            l.days,
            NULL as shift_name,
            l.status
        FROM leaves l
        JOIN users u ON l.user_id = u.id
        WHERE l.status = 'approved'
            AND l.start_date >= ? AND l.end_date <= ?
        ORDER BY l.start_date DESC
    ''', (start_date, end_date)).fetchall()

    mission_report = conn.execute('''
        SELECT
            m.id,
            m.user_id,
            u.full_name,
            u.username,
            m.start_date as date,
            NULL as shift,
            NULL as check_in,
            NULL as check_out,
            m.days as total_hours,
            'mission' as type,
            NULL as reason,
            m.destination,
            m.start_date,
            m.end_date,
            m.days,
            NULL as shift_name,
            m.status
        FROM missions m
        JOIN users u ON m.user_id = u.id
        WHERE m.status = 'approved'
            AND m.start_date >= ? AND m.end_date <= ?
        ORDER BY m.start_date DESC
    ''', (start_date, end_date)).fetchall()

    all_reports = []

    for row in daily_report:
        all_reports.append(dict(row))

    for row in leave_report:
        all_reports.append(dict(row))

    for row in mission_report:
        all_reports.append(dict(row))

    all_reports.sort(key=lambda x: x['date'] if x['date'] else '', reverse=True)

    monthly_summary = conn.execute('''
        SELECT
            u.id,
            u.username,
            u.full_name,
            u.role,
            COUNT(DISTINCT a.date) as days_worked,
            COALESCE(SUM(a.total_hours), 0) as total_hours,
            COUNT(DISTINCT CASE WHEN a.shift = 3 THEN a.date END) as night_shifts,
            COALESCE(SUM(CASE WHEN a.shift = 3 THEN a.total_hours ELSE 0 END), 0) as night_hours
        FROM users u
        LEFT JOIN attendance a ON u.id = a.user_id
            AND a.date BETWEEN ? AND ?
            AND a.check_in IS NOT NULL
            AND a.check_out IS NOT NULL
        WHERE u.role != 'admin'
        GROUP BY u.id
        ORDER BY total_hours DESC
    ''', (start_date, end_date)).fetchall()

    leave_days = conn.execute('''
        SELECT
            user_id,
            SUM(days) as total_leave_days
        FROM leaves
        WHERE status = 'approved'
            AND start_date >= ? AND end_date <= ?
        GROUP BY user_id
    ''', (start_date, end_date)).fetchall()

    mission_days = conn.execute('''
        SELECT
            user_id,
            SUM(days) as total_mission_days
        FROM missions
        WHERE status = 'approved'
            AND start_date >= ? AND end_date <= ?
        GROUP BY user_id
    ''', (start_date, end_date)).fetchall()

    conn.close()

    summary = []
    for emp in monthly_summary:
        emp_dict = dict(emp)
        emp_dict['total_leave_days'] = 0
        emp_dict['total_mission_days'] = 0

        for leave in leave_days:
            if leave['user_id'] == emp['id']:
                emp_dict['total_leave_days'] = leave['total_leave_days']
                break

        for mission in mission_days:
            if mission['user_id'] == emp['id']:
                emp_dict['total_mission_days'] = mission['total_mission_days']
                break

        summary.append(emp_dict)

    return {
        'daily': all_reports,
        'summary': summary
    }

def get_monthly_summary(year, month):
    start_date = f"{year}-{month:02d}-01"
    if month == 12:
        end_date = f"{year}-12-31"
    else:
        next_month = month + 1
        end_dt = datetime(year, next_month, 1) - timedelta(days=1)
        end_date = end_dt.strftime('%Y-%m-%d')
    return get_attendance_report(start_date, end_date)

# ============================================================
# ATTENDANCE MANAGEMENT FUNCTIONS
# ============================================================

def get_attendance_by_id(attendance_id):
    try:
        conn = get_db_connection()
        data = conn.execute('''
            SELECT a.*, u.full_name, u.username
            FROM attendance a
            JOIN users u ON u.id = a.user_id
            WHERE a.id = ?
        ''', (attendance_id,)).fetchone()
        conn.close()
        return data
    except Exception as e:
        print(f"Error in get_attendance_by_id: {e}")
        return None

def update_attendance(attendance_id, check_in=None, check_out=None, date=None, shift=None):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        updates = []
        params = []

        if check_in is not None:
            updates.append("check_in = ?")
            params.append(check_in)
        if check_out is not None:
            updates.append("check_out = ?")
            params.append(check_out)
        if date is not None:
            updates.append("date = ?")
            params.append(date)
        if shift is not None:
            updates.append("shift = ?")
            params.append(shift)

        if check_in is not None and check_out is not None:
            try:
                from datetime import datetime
                check_in_dt = datetime.strptime(check_in, '%Y-%m-%d %H:%M:%S')
                check_out_dt = datetime.strptime(check_out, '%Y-%m-%d %H:%M:%S')
                total_hours = (check_out_dt - check_in_dt).total_seconds() / 3600
                updates.append("total_hours = ?")
                params.append(total_hours)
            except:
                pass

        if not updates:
            conn.close()
            return False

        params.append(attendance_id)
        sql = f"UPDATE attendance SET {', '.join(updates)} WHERE id = ?"
        cursor.execute(sql, params)
        conn.commit()
        conn.close()
        increment_data_version()
        return True
    except Exception as e:
        print(f"Error in update_attendance: {e}")
        return False

def delete_attendance(attendance_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM attendance WHERE id = ?", (attendance_id,))
        conn.commit()
        conn.close()
        increment_data_version()
        return True
    except Exception as e:
        print(f"Error in delete_attendance: {e}")
        return False

# ============================================================
# LEAVE FUNCTIONS
# ============================================================

def create_leave(user_id, start_date, end_date, days, reason='', attachment=None):
    conn = get_db_connection()
    conn.execute(
        "INSERT INTO leaves (user_id, start_date, end_date, days, reason, attachment, status) VALUES (?, ?, ?, ?, ?, ?, 'pending')",
        (user_id, start_date, end_date, days, reason, attachment)
    )
    conn.commit()
    conn.close()
    increment_data_version()
    return True

def get_pending_leaves():
    conn = get_db_connection()
    leaves = conn.execute('''
        SELECT l.*, u.username, u.full_name
        FROM leaves l
        JOIN users u ON l.user_id = u.id
        WHERE l.status = 'pending'
        ORDER BY l.created_at DESC
    ''').fetchall()
    conn.close()
    return [dict(row) for row in leaves]

def approve_leave(leave_id, admin_id):
    conn = get_db_connection()
    conn.execute(
        "UPDATE leaves SET status = 'approved', admin_id = ? WHERE id = ?",
        (admin_id, leave_id)
    )
    conn.commit()
    conn.close()
    increment_data_version()
    return True

def reject_leave(leave_id):
    conn = get_db_connection()
    conn.execute(
        "UPDATE leaves SET status = 'rejected' WHERE id = ?",
        (leave_id,)
    )
    conn.commit()
    conn.close()
    increment_data_version()
    return True

# ============================================================
# MISSION FUNCTIONS
# ============================================================

def create_mission(user_id, start_date, end_date, days, destination='', purpose='', attachment=None):
    conn = get_db_connection()
    conn.execute(
        "INSERT INTO missions (user_id, start_date, end_date, days, destination, purpose, attachment, status) VALUES (?, ?, ?, ?, ?, ?, ?, 'pending')",
        (user_id, start_date, end_date, days, destination, purpose, attachment)
    )
    conn.commit()
    conn.close()
    increment_data_version()
    return True

def get_pending_missions():
    conn = get_db_connection()
    missions = conn.execute('''
        SELECT m.*, u.username, u.full_name
        FROM missions m
        JOIN users u ON m.user_id = u.id
        WHERE m.status = 'pending'
        ORDER BY m.created_at DESC
    ''').fetchall()
    conn.close()
    return [dict(row) for row in missions]

def approve_mission(mission_id, admin_id):
    conn = get_db_connection()
    conn.execute(
        "UPDATE missions SET status = 'approved', admin_id = ? WHERE id = ?",
        (admin_id, mission_id)
    )
    conn.commit()
    conn.close()
    increment_data_version()
    return True

def reject_mission(mission_id):
    conn = get_db_connection()
    conn.execute(
        "UPDATE missions SET status = 'rejected' WHERE id = ?",
        (mission_id,)
    )
    conn.commit()
    conn.close()
    increment_data_version()
    return True

# ============================================================
# CLEAN DATA FUNCTIONS
# ============================================================

def clean_all_data():
    conn = get_db_connection()
    conn.execute("DELETE FROM attendance")
    conn.execute("DELETE FROM leaves")
    conn.execute("DELETE FROM missions")
    conn.commit()
    conn.close()
    increment_data_version()
    return True

def clean_attendance_only():
    conn = get_db_connection()
    conn.execute("DELETE FROM attendance")
    conn.commit()
    conn.close()
    increment_data_version()
    return True

def clean_leaves_only():
    conn = get_db_connection()
    conn.execute("DELETE FROM leaves")
    conn.commit()
    conn.close()
    increment_data_version()
    return True

def clean_missions_only():
    conn = get_db_connection()
    conn.execute("DELETE FROM missions")
    conn.commit()
    conn.close()
    increment_data_version()
    return True

# ============================================================
# HELPER FUNCTIONS
# ============================================================

def haversine_distance(lat1, lon1, lat2, lon2):
    R = 6371000
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    a = math.sin(delta_phi/2)**2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    return R * c

def save_uploaded_file(file, user_id, folder_type):
    if file and file.filename:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_id = str(uuid.uuid4())[:8]
        file_ext = os.path.splitext(file.filename)[1]
        new_filename = f"{folder_type}_{user_id}_{timestamp}_{unique_id}{file_ext}"

        if folder_type == 'leave':
            upload_folder = UPLOAD_FOLDER_LEAVES
        else:
            upload_folder = UPLOAD_FOLDER_MISSIONS

        file_path = os.path.join(upload_folder, new_filename)
        file.save(file_path)
        return f"/static/uploads/{folder_type}s/{new_filename}"
    return None

# ============================================================
# ROUTES
# ============================================================

@app.route('/')
def home():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    return redirect(url_for('dashboard'))

@app.route('/dashboard')
def dashboard():
    if not session.get('logged_in'):
        return redirect(url_for('login'))

    stats = get_attendance_stats()

    today = datetime.now()
    start_date = today.replace(day=1).strftime('%Y-%m-%d')
    end_date = today.strftime('%Y-%m-%d')
    work_history = get_work_history_report(start_date, end_date, 200)
    monthly_summary = get_monthly_summary_report(start_date, end_date)
    company_loc = get_company_location()

    pending_leaves = get_pending_leaves()
    pending_missions = get_pending_missions()
    pending_count = len(pending_leaves) + len(pending_missions)

    data_version = get_data_version()
    current_month = today.strftime('%m')
    current_year = today.strftime('%Y')
    allowed_distance = app.config['ALLOWED_DISTANCE']

    lock_status = get_system_lock_status()
    user_lock = get_user_lock_status(session.get('user_id'))

    return render_template_string(DASHBOARD_HTML,
                                   session=session,
                                   stats=stats,
                                   work_history=work_history,
                                   monthly_summary=monthly_summary,
                                   company_lat=company_loc['lat'] if company_loc else '',
                                   company_lng=company_loc['lng'] if company_loc else '',
                                   pending_count=pending_count,
                                   data_version=data_version,
                                   current_month=current_month,
                                   current_year=current_year,
                                   allowed_distance=allowed_distance,
                                   lock_status=lock_status,
                                   user_lock=user_lock)

@app.route('/get_data_version')
def get_data_version_route():
    if not session.get('logged_in'):
        return jsonify({'version': 0})
    version = get_data_version()
    return jsonify({'version': version})

@app.route('/get_checkin_status')
def get_checkin_status_route():
    if not session.get('logged_in'):
        return jsonify({'has_checkin': False, 'check_in_time': None, 'shift': None})

    user_id = session.get('user_id')
    if not user_id:
        return jsonify({'has_checkin': False, 'check_in_time': None, 'shift': None})

    status = get_checkin_status(user_id)
    return jsonify(status)

@app.route('/get_system_lock_status')
def get_system_lock_status_route():
    if not session.get('logged_in'):
        return jsonify({'is_locked': 0, 'auto_unlock_time': '06:00'})
    lock = get_system_lock_status()
    return jsonify(lock)

@app.route('/toggle_system_lock', methods=['POST'])
def toggle_system_lock_route():
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'អ្នកមិនមែនជា Admin!'})

    data = request.get_json()
    lock_state = data.get('lock_state', 0)
    auto_unlock_time = data.get('auto_unlock_time', '06:00')
    user_id = session.get('user_id')

    result = toggle_system_lock(lock_state, auto_unlock_time, user_id)
    if result:
        status = "បិទ" if lock_state == 1 else "បើក"
        return jsonify({'success': True, 'message': f'✅ បាន{status}ប្រព័ន្ធចូលធ្វើការជោគជ័យ!'})
    return jsonify({'success': False, 'message': '❌ មិនអាចប្តូរស្ថានភាពបាន!'})

# ============================================================
# USER LOCK ROUTES
# ============================================================

@app.route('/get_user_lock/<int:user_id>')
def get_user_lock_route(user_id):
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    lock = get_user_lock_status(user_id)
    return jsonify(lock)

@app.route('/get_all_user_locks')
def get_all_user_locks_route():
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify([]), 403
    locks = get_all_user_lock_status()
    return jsonify(locks)

@app.route('/toggle_user_lock', methods=['POST'])
def toggle_user_lock_route():
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'អ្នកមិនមែនជា Admin!'})

    data = request.get_json()
    user_id = data.get('user_id')
    lock_state = data.get('lock_state', 0)
    auto_unlock_time = data.get('auto_unlock_time', '')
    admin_id = session.get('user_id')

    if not user_id:
        return jsonify({'success': False, 'message': 'មិនមាន user_id!'})

    if auto_unlock_time:
        try:
            datetime.strptime(auto_unlock_time, '%H:%M')
        except ValueError:
            return jsonify({'success': False, 'message': 'ទ្រង់ទ្រាយម៉ោងមិនត្រឹមត្រូវ! សូមប្រើ HH:MM'})

    if user_id == admin_id:
        return jsonify({'success': False, 'message': '❌ អ្នកមិនអាចបិទគណនីរបស់ខ្លួនឯងបានទេ!'})

    result = toggle_user_lock(user_id, lock_state, auto_unlock_time if auto_unlock_time else None, admin_id)
    if result:
        user = get_user_by_id(user_id)
        status = "បិទ" if lock_state == 1 else "បើក"
        return jsonify({'success': True, 'message': f'✅ បាន{status}ការចូលធ្វើការរបស់ {user["full_name"]} ជោគជ័យ!'})
    return jsonify({'success': False, 'message': '❌ មិនអាចប្តូរស្ថានភាពបាន!'})

@app.route('/check_in', methods=['POST'])
def check_in_route():
    try:
        if not session.get('logged_in'):
            return jsonify({'success': False, 'message': 'មិនទាន់ Login'})

        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'success': False, 'message': 'មិនមាន user_id ក្នុង Session'})

        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'មិនមានទិន្នន័យផ្ញើមក'})

        user_lat = data.get('user_lat')
        user_lng = data.get('user_lng')
        company_lat = data.get('company_lat')
        company_lng = data.get('company_lng')
        shift = data.get('shift', 1)

        if user_lat is None or user_lng is None:
            return jsonify({'success': False, 'message': 'មិនមានទីតាំងរបស់អ្នក!'})

        if company_lat is None or company_lng is None:
            return jsonify({'success': False, 'message': 'សូមឲ្យ Admin កំណត់ទីតាំងក្រុមហ៊ុនជាមុនសិន!'})

        allowed, lock_message = check_system_lock_for_user(user_id)
        if not allowed:
            return jsonify({'success': False, 'message': lock_message})

        allowed, user_lock_message = check_user_lock(user_id)
        if not allowed:
            return jsonify({'success': False, 'message': user_lock_message})

        can_checkin, deadline = check_attendance_deadline(user_id)
        if not can_checkin:
            system_lock = get_system_lock_status()
            auto_unlock = system_lock.get('auto_unlock_time', '06:00')
            update_user_lock(user_id, 1, auto_unlock_time=auto_unlock)
            increment_data_version()
            return jsonify({'success': False, 'message': f'⛔ អ្នកលើសម៉ោងដែល Admin បានកំណត់ (ម៉ោងកំណត់: {deadline})! ប្រព័ន្ធបានបិទការចូលធ្វើការរបស់អ្នកដោយស្វ័យប្រវត្តិ!'})

        distance = haversine_distance(user_lat, user_lng, company_lat, company_lng)
        allowed_distance = app.config['ALLOWED_DISTANCE']

        if distance > allowed_distance:
            return jsonify({
                'success': False,
                'message': f'អ្នកនៅឆ្ងាយពីទីតាំងក្រុមហ៊ុន {round(distance, 2)} ម៉ែត្រ (អនុញ្ញាតត្រឹម {allowed_distance} ម៉ែត្រ)! មិនអាចចូលធ្វើការបាន!'
            })

        status = get_checkin_status(user_id)
        if status['has_checkin']:
            return jsonify({'success': False, 'message': 'អ្នកបានចូលធ្វើការរួចហើយ! សូមចុច "ចេញពីធ្វើការ" មុនពេលចូលម្តងទៀត!'})

        result, message = check_in(user_id, user_lat, user_lng, distance, shift)
        if result:
            shift_names = {1: 'វគ្គ 1 (ព្រឹក)', 2: 'វគ្គ 2 (រសៀល)', 3: 'វគ្គ 3 (យប់)'}
            return jsonify({
                'success': True,
                'message': f'✅ {message} ({shift_names.get(shift, "វគ្គ " + str(shift))}) (ចម្ងាយ {round(distance, 2)} ម៉ែត្រ)'
            })
        else:
            return jsonify({'success': False, 'message': '❌ ' + message})

    except Exception as e:
        print(f"Error in check_in: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'កំហុស: {str(e)}'})

@app.route('/check_out', methods=['POST'])
def check_out_route():
    try:
        if not session.get('logged_in'):
            return jsonify({'success': False, 'message': 'មិនទាន់ Login'})

        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'success': False, 'message': 'មិនមាន user_id ក្នុង Session'})

        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'មិនមានទិន្នន័យផ្ញើមក'})

        user_lat = data.get('user_lat')
        user_lng = data.get('user_lng')
        company_lat = data.get('company_lat')
        company_lng = data.get('company_lng')

        if user_lat is None or user_lng is None:
            return jsonify({'success': False, 'message': 'មិនមានទីតាំងរបស់អ្នក!'})

        if company_lat is None or company_lng is None:
            return jsonify({'success': False, 'message': 'សូមឲ្យ Admin កំណត់ទីតាំងក្រុមហ៊ុនជាមុនសិន!'})

        allowed, lock_message = check_system_lock_for_user(user_id)
        if not allowed:
            return jsonify({'success': False, 'message': lock_message})

        allowed, user_lock_message = check_user_lock(user_id)
        if not allowed:
            return jsonify({'success': False, 'message': user_lock_message})

        distance = haversine_distance(user_lat, user_lng, company_lat, company_lng)
        allowed_distance = app.config['ALLOWED_DISTANCE']

        if distance > allowed_distance:
            return jsonify({
                'success': False,
                'message': f'អ្នកនៅឆ្ងាយពីទីតាំងក្រុមហ៊ុន {round(distance, 2)} ម៉ែត្រ (អនុញ្ញាតត្រឹម {allowed_distance} ម៉ែត្រ)! មិនអាចចេញធ្វើការបាន!'
            })

        status = get_checkin_status(user_id)
        if not status['has_checkin']:
            return jsonify({'success': False, 'message': 'អ្នកមិនទាន់ចូលធ្វើការនៅថ្ងៃនេះទេ!'})

        result, message = check_out(user_id, user_lat, user_lng, distance)
        if result:
            return jsonify({'success': True, 'message': '✅ ' + message})
        else:
            return jsonify({'success': False, 'message': '❌ ' + message})

    except Exception as e:
        print(f"Error in check_out: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'កំហុស: {str(e)}'})

@app.route('/get_company_location')
def get_company_location_route():
    location = get_company_location()
    if location:
        return jsonify({'lat': location['lat'], 'lng': location['lng']})
    return jsonify({'lat': None, 'lng': None})

@app.route('/save_location', methods=['POST'])
def save_location():
    if not session.get('logged_in'):
        return jsonify({'success': False, 'message': 'មិនទាន់ Login'})
    data = request.get_json()
    lat = data.get('lat')
    lng = data.get('lng')
    if lat is None or lng is None:
        return jsonify({'success': False, 'message': 'ទិន្នន័យមិនត្រឹមត្រូវ'})
    result = save_company_location(lat, lng)
    if result:
        return jsonify({'success': True, 'message': 'រក្សាទុកជោគជ័យ'})
    return jsonify({'success': False, 'message': 'មិនអាចរក្សាទុកបាន'})

@app.route('/register', methods=['GET', 'POST'])
def register():
    if not session.get('logged_in') or session.get('role') != 'admin':
        return redirect(url_for('dashboard'))
    message = ''
    message_type = ''
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        full_name = request.form.get('full_name')
        email = request.form.get('email')
        phone = request.form.get('phone')
        role = request.form.get('role')
        if not username or not password or not full_name:
            message = 'សូមបំពេញព័ត៌មានឲ្យបានពេញលេញ!'
            message_type = 'error'
        elif password != confirm_password:
            message = 'ពាក្យសម្ងាត់មិនត្រូវគ្នា!'
            message_type = 'error'
        elif len(password) < 4:
            message = 'ពាក្យសម្ងាត់ត្រូវមានយ៉ាងតិច 4 តួ!'
            message_type = 'error'
        else:
            if create_user(username, password, full_name, email, phone, role):
                message = f'✅ ចុះឈ្មោះអ្នកប្រើ "{username}" ជោគជ័យ!'
                message_type = 'success'
            else:
                message = f'❌ ឈ្មោះអ្នកប្រើ "{username}" មានរួចហើយ!'
                message_type = 'error'
    return render_template_string(REGISTER_HTML, session=session, message=message, message_type=message_type)

@app.route('/manage_users')
def manage_users():
    if not session.get('logged_in') or session.get('role') != 'admin':
        return redirect(url_for('dashboard'))
    users = get_all_users()
    settings = get_all_attendance_settings()
    settings_dict = {s['user_id']: s for s in settings}
    user_locks = get_all_user_lock_status()
    user_locks_dict = {u['id']: u for u in user_locks}
    return render_template_string(USER_MANAGEMENT_HTML,
                                   session=session,
                                   users=users,
                                   settings=settings_dict,
                                   user_locks=user_locks_dict,
                                   message='',
                                   message_type='')

@app.route('/get_user/<int:user_id>')
def get_user(user_id):
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    user = get_user_by_id(user_id)
    if not user:
        return jsonify({'error': 'User not found'}), 404
    return jsonify({'id': user['id'], 'username': user['username'], 'full_name': user['full_name'], 'email': user['email'] or '', 'phone': user['phone'] or '', 'role': user['role']})

@app.route('/get_attendance_setting/<int:user_id>')
def get_attendance_setting_route(user_id):
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    setting = get_attendance_setting(user_id)
    if setting:
        return jsonify(setting)
    return jsonify({'user_id': user_id, 'check_in_deadline': '', 'is_active': 0})

@app.route('/save_attendance_setting', methods=['POST'])
def save_attendance_setting_route():
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'អ្នកមិនមែនជា Admin!'})

    data = request.get_json()
    user_id = data.get('user_id')
    check_in_deadline = data.get('check_in_deadline', '').strip()
    is_active = data.get('is_active', 0)

    if not user_id:
        return jsonify({'success': False, 'message': 'មិនមាន user_id!'})

    if check_in_deadline:
        try:
            datetime.strptime(check_in_deadline, '%H:%M')
        except ValueError:
            return jsonify({'success': False, 'message': 'ទ្រង់ទ្រាយម៉ោងមិនត្រឹមត្រូវ! សូមប្រើ HH:MM (ឧទាហរណ៍: 08:00)'})

    result = save_attendance_setting(user_id, check_in_deadline, int(is_active))
    if result:
        return jsonify({'success': True, 'message': '✅ រក្សាទុកការកំណត់ជោគជ័យ!'})
    return jsonify({'success': False, 'message': '❌ មិនអាចរក្សាទុកបាន!'})

@app.route('/update_user/<int:user_id>', methods=['POST'])
def update_user_route(user_id):
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'អ្នកមិនមែនជា Admin!'})
    data = request.get_json()
    result = update_user(user_id, data.get('username'), data.get('full_name'), data.get('email'), data.get('phone'), data.get('role'))
    if result:
        return jsonify({'success': True, 'message': '✅ កែប្រែអ្នកប្រើជោគជ័យ!'})
    return jsonify({'success': False, 'message': '❌ មិនអាចកែប្រែបាន!'})

@app.route('/add_user', methods=['POST'])
def add_user_route():
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'អ្នកមិនមែនជា Admin!'})
    data = request.get_json()
    if create_user(data.get('username'), data.get('password'), data.get('full_name'), data.get('email'), data.get('phone'), data.get('role')):
        return jsonify({'success': True, 'message': f'✅ បានបន្ថែមអ្នកប្រើ "{data.get("username")}" ជោគជ័យ!'})
    return jsonify({'success': False, 'message': '❌ ឈ្មោះអ្នកប្រើមានរួចហើយ!'})

@app.route('/delete_user/<int:user_id>', methods=['POST'])
def delete_user_route(user_id):
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'អ្នកមិនមែនជា Admin!'})
    if user_id == session.get('user_id'):
        return jsonify({'success': False, 'message': '❌ អ្នកមិនអាចលុបគណនីរបស់ខ្លួនឯងបានទេ!'})
    if delete_user(user_id):
        return jsonify({'success': True, 'message': '✅ លុបអ្នកប្រើជោគជ័យ!'})
    return jsonify({'success': False, 'message': '❌ មិនអាចលុបបាន!'})

@app.route('/change_password', methods=['GET', 'POST'])
def change_password():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    message = ''
    message_type = ''
    user_id = session.get('user_id')
    if request.method == 'POST':
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        if not current_password or not new_password or not confirm_password:
            message = 'សូមបំពេញព័ត៌មានឲ្យបានពេញលេញ!'
            message_type = 'error'
        elif new_password != confirm_password:
            message = 'ពាក្យសម្ងាត់ថ្មីមិនត្រូវគ្នា!'
            message_type = 'error'
        elif len(new_password) < 4:
            message = 'ពាក្យសម្ងាត់ថ្មីត្រូវមានយ៉ាងតិច 4 តួ!'
            message_type = 'error'
        elif not verify_password(user_id, current_password):
            message = 'ពាក្យសម្ងាត់បច្ចុប្បន្នមិនត្រឹមត្រូវ!'
            message_type = 'error'
        else:
            if update_password(user_id, new_password):
                message = '✅ បានប្តូរពាក្យសម្ងាត់ជោគជ័យ!'
                message_type = 'success'
            else:
                message = '❌ មិនអាចប្តូរពាក្យសម្ងាត់បាន!'
                message_type = 'error'
    return render_template_string(CHANGE_PASSWORD_HTML, session=session, message=message, message_type=message_type)

@app.route('/admin_reset_password/<int:user_id>', methods=['POST'])
def admin_reset_password(user_id):
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'អ្នកមិនមែនជា Admin!'})
    data = request.get_json()
    new_password = data.get('new_password')
    if not new_password or len(new_password) < 4:
        return jsonify({'success': False, 'message': 'ពាក្យសម្ងាត់ត្រូវមានយ៉ាងតិច 4 តួ!'})
    if update_password(user_id, new_password):
        return jsonify({'success': True, 'message': '✅ បានកំណត់ពាក្យសម្ងាត់ថ្មីជោគជ័យ!'})
    return jsonify({'success': False, 'message': '❌ មិនអាចកំណត់ពាក្យសម្ងាត់បាន!'})

@app.route('/clean_data', methods=['POST'])
def clean_data():
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'អ្នកមិនមែនជា Admin!'})
    data = request.get_json()
    clean_type = data.get('type', 'all')
    if clean_type == 'all':
        result = clean_all_data()
        message = '✅ បានសម្អាតទិន្នន័យទាំងអស់ជោគជ័យ!'
    elif clean_type == 'attendance':
        result = clean_attendance_only()
        message = '✅ បានសម្អាតទិន្នន័យវត្តមានជោគជ័យ!'
    elif clean_type == 'leaves':
        result = clean_leaves_only()
        message = '✅ បានសម្អាតទិន្នន័យសុំច្បាប់ជោគជ័យ!'
    elif clean_type == 'missions':
        result = clean_missions_only()
        message = '✅ បានសម្អាតទិន្នន័យបេសកម្មជោគជ័យ!'
    else:
        return jsonify({'success': False, 'message': 'ប្រភេទមិនត្រឹមត្រូវ!'})
    if result:
        return jsonify({'success': True, 'message': message})
    return jsonify({'success': False, 'message': '❌ មិនអាចសម្អាតទិន្នន័យបាន!'})

@app.route('/request_leave', methods=['POST'])
def request_leave():
    if not session.get('logged_in'):
        return jsonify({'success': False, 'message': 'មិនទាន់ Login'})

    user_id = session.get('user_id')

    attachment = None
    if 'attachment' in request.files:
        file = request.files['attachment']
        attachment = save_uploaded_file(file, user_id, 'leave')

    if request.is_json:
        data = request.get_json()
        result = create_leave(
            user_id,
            data.get('start_date'),
            data.get('end_date'),
            data.get('days'),
            data.get('reason', ''),
            attachment
        )
        days = data.get('days')
    else:
        result = create_leave(
            user_id,
            request.form.get('start_date'),
            request.form.get('end_date'),
            request.form.get('days'),
            request.form.get('reason', ''),
            attachment
        )
        days = request.form.get('days')

    if result:
        return jsonify({'success': True, 'message': f'✅ បានសុំច្បាប់ {days} ថ្ងៃរួចហើយ! រង់ចាំការអនុម័តពី Admin'})
    return jsonify({'success': False, 'message': '❌ មិនអាចសុំច្បាប់បាន!'})

@app.route('/request_mission', methods=['POST'])
def request_mission():
    if not session.get('logged_in'):
        return jsonify({'success': False, 'message': 'មិនទាន់ Login'})

    user_id = session.get('user_id')

    attachment = None
    if 'attachment' in request.files:
        file = request.files['attachment']
        attachment = save_uploaded_file(file, user_id, 'mission')

    if request.is_json:
        data = request.get_json()
        result = create_mission(
            user_id,
            data.get('start_date'),
            data.get('end_date'),
            data.get('days'),
            data.get('destination', ''),
            data.get('purpose', ''),
            attachment
        )
        days = data.get('days')
    else:
        result = create_mission(
            user_id,
            request.form.get('start_date'),
            request.form.get('end_date'),
            request.form.get('days'),
            request.form.get('destination', ''),
            request.form.get('purpose', ''),
            attachment
        )
        days = request.form.get('days')

    if result:
        return jsonify({'success': True, 'message': f'✅ បានសុំបេសកម្ម {days} ថ្ងៃរួចហើយ! រង់ចាំការអនុម័តពី Admin'})
    return jsonify({'success': False, 'message': '❌ មិនអាចសុំបេសកម្មបាន!'})

@app.route('/get_pending_requests')
def get_pending_requests():
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'leaves': [], 'missions': []})
    leaves = get_pending_leaves()
    missions = get_pending_missions()
    return jsonify({'leaves': leaves, 'missions': missions})

@app.route('/check_new_requests')
def check_new_requests():
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'new': False, 'count': 0, 'message': ''})
    pending_leaves = get_pending_leaves()
    pending_missions = get_pending_missions()
    total_pending = len(pending_leaves) + len(pending_missions)
    last_count = session.get('last_pending_count', 0)
    is_new = total_pending > last_count
    session['last_pending_count'] = total_pending
    message = ''
    if is_new and total_pending > 0:
        new_count = total_pending - last_count
        message = f'📬 មានសំណើថ្មី {new_count} ករណី! សូមចុច "📬 សំណើ" ដើម្បីអនុម័ត'
    return jsonify({'new': is_new, 'count': total_pending, 'message': message})

@app.route('/approve_request', methods=['POST'])
def approve_request():
    try:
        if not session.get('logged_in'):
            return jsonify({'success': False, 'message': 'សូម Login ជាមុន!'})

        if session.get('role') != 'admin':
            return jsonify({'success': False, 'message': 'អ្នកមិនមែនជា Admin!'})

        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'មិនមានទិន្នន័យផ្ញើមក!'})

        req_type = data.get('type')
        req_id = data.get('id')
        admin_id = session.get('user_id')

        print(f"🔄 Approving: type={req_type}, id={req_id}, admin={admin_id}")

        if req_type == 'leave':
            result = approve_leave(req_id, admin_id)
            print(f"📝 Approve leave result: {result}")
        elif req_type == 'mission':
            result = approve_mission(req_id, admin_id)
            print(f"📝 Approve mission result: {result}")
        else:
            return jsonify({'success': False, 'message': 'ប្រភេទមិនត្រឹមត្រូវ'})

        if result:
            return jsonify({'success': True, 'message': '✅ បានអនុម័តជោគជ័យ!'})
        return jsonify({'success': False, 'message': '❌ មិនអាចអនុម័តបាន!'})
    except Exception as e:
        print(f"❌ Error in approve_request: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'កំហុស: {str(e)}'})

@app.route('/reject_request', methods=['POST'])
def reject_request():
    try:
        if not session.get('logged_in'):
            return jsonify({'success': False, 'message': 'សូម Login ជាមុន!'})

        if session.get('role') != 'admin':
            return jsonify({'success': False, 'message': 'អ្នកមិនមែនជា Admin!'})

        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'មិនមានទិន្នន័យផ្ញើមក!'})

        req_type = data.get('type')
        req_id = data.get('id')

        print(f"🔄 Rejecting: type={req_type}, id={req_id}")

        if req_type == 'leave':
            result = reject_leave(req_id)
            print(f"📝 Reject leave result: {result}")
        elif req_type == 'mission':
            result = reject_mission(req_id)
            print(f"📝 Reject mission result: {result}")
        else:
            return jsonify({'success': False, 'message': 'ប្រភេទមិនត្រឹមត្រូវ'})

        if result:
            return jsonify({'success': True, 'message': '✅ បានបដិសេធជោគជ័យ!'})
        return jsonify({'success': False, 'message': '❌ មិនអាចបដិសេធបាន!'})
    except Exception as e:
        print(f"❌ Error in reject_request: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'កំហុស: {str(e)}'})

@app.route('/report')
def report():
    if not session.get('logged_in'):
        return redirect(url_for('login'))

    start_date = request.args.get('start')
    end_date = request.args.get('end')
    month = request.args.get('month')
    year = request.args.get('year')
    current_year = datetime.now().year

    if month and year:
        month = int(month)
        year = int(year)
        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year}-12-31"
        else:
            next_month = month + 1
            end_dt = datetime(year, next_month, 1) - timedelta(days=1)
            end_date = end_dt.strftime('%Y-%m-%d')
    elif start_date and end_date:
        pass
    else:
        today = datetime.now()
        start_date = today.replace(day=1).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')

    result = get_attendance_report(start_date, end_date)
    daily_data = result.get('daily', [])
    summary_data = result.get('summary', [])

    return render_template_string(REPORT_HTML,
                                   session=session,
                                   daily=daily_data,
                                   summary=summary_data,
                                   start_date=start_date,
                                   end_date=end_date,
                                   current_year=current_year,
                                   user_role=session.get('role', 'user'))

@app.route('/export_excel')
def export_excel():
    if not session.get('logged_in'):
        return redirect(url_for('login'))

    start_date = request.args.get('start')
    end_date = request.args.get('end')
    month = request.args.get('month')
    year = request.args.get('year')

    if month and year:
        month = int(month)
        year = int(year)
        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year}-12-31"
        else:
            next_month = month + 1
            end_dt = datetime(year, next_month, 1) - timedelta(days=1)
            end_date = end_dt.strftime('%Y-%m-%d')
    elif start_date and end_date:
        pass
    else:
        today = datetime.now()
        start_date = today.replace(day=1).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')

    result = get_attendance_report(start_date, end_date)
    daily_data = result.get('daily', [])
    summary_data = result.get('summary', [])

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "ប្រចាំថ្ងៃ"

    ws1.merge_cells('A1:H1')
    ws1['A1'] = f"របាយការណ៍ប្រចាំថ្ងៃ\nចាប់ពី {start_date} ដល់ {end_date}"
    ws1['A1'].font = Font(size=16, bold=True)
    ws1['A1'].alignment = Alignment(horizontal='center')

    headers = ['ល.រ', 'ឈ្មោះបុគ្គលិក', 'កាលបរិច្ឆេទ', 'ប្រភេទ', 'ចាប់ផ្តើម', 'បញ្ចប់', 'ចំនួន', 'ស្ថានភាព/ព័ត៌មាន']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1a73e8', end_color='1a73e8', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    for row_idx, item in enumerate(daily_data, 4):
        ws1.cell(row=row_idx, column=1, value=row_idx-3)
        ws1.cell(row=row_idx, column=2, value=item.get('full_name', ''))
        ws1.cell(row=row_idx, column=3, value=item.get('date', ''))

        if item.get('type') == 'attendance':
            ws1.cell(row=row_idx, column=4, value='វត្តមាន')
        elif item.get('type') == 'leave':
            ws1.cell(row=row_idx, column=4, value='ច្បាប់')
        elif item.get('type') == 'mission':
            ws1.cell(row=row_idx, column=4, value='បេសកម្ម')
        else:
            ws1.cell(row=row_idx, column=4, value='-')

        if item.get('type') == 'attendance':
            ws1.cell(row=row_idx, column=5, value=item.get('check_in', '')[:16] if item.get('check_in') else '')
            ws1.cell(row=row_idx, column=6, value=item.get('check_out', '')[:16] if item.get('check_out') else '')
            if item.get('total_hours'):
                hours = int(item['total_hours'])
                minutes = int((item['total_hours'] - hours) * 60)
                ws1.cell(row=row_idx, column=7, value=f"{hours:02d}:{minutes:02d}")
            else:
                ws1.cell(row=row_idx, column=7, value='-')
            if item.get('check_in') and item.get('check_out'):
                ws1.cell(row=row_idx, column=8, value='បានបិទ')
            elif item.get('check_in') and not item.get('check_out'):
                ws1.cell(row=row_idx, column=8, value='កំពុងធ្វើការ')
            else:
                ws1.cell(row=row_idx, column=8, value='មិនទាន់ចូល')
        else:
            ws1.cell(row=row_idx, column=5, value=item.get('start_date', ''))
            ws1.cell(row=row_idx, column=6, value=item.get('end_date', ''))
            ws1.cell(row=row_idx, column=7, value=f"{item.get('days', 0)} ថ្ងៃ")
            if item.get('type') == 'leave':
                ws1.cell(row=row_idx, column=8, value=f"ច្បាប់: {item.get('reason', '')}")
            else:
                ws1.cell(row=row_idx, column=8, value=f"បេសកម្ម: {item.get('destination', '')}")

    for col in range(1, 9):
        ws1.column_dimensions[chr(64 + col)].width = 18

    ws2 = wb.create_sheet("សង្ខេបប្រចាំខែ")

    ws2.merge_cells('A1:H1')
    ws2['A1'] = f"របាយការណ៍សង្ខេបប្រចាំខែ\nចាប់ពី {start_date} ដល់ {end_date}"
    ws2['A1'].font = Font(size=16, bold=True)
    ws2['A1'].alignment = Alignment(horizontal='center')

    headers2 = ['ល.រ', 'ឈ្មោះបុគ្គលិក', 'ថ្ងៃធ្វើការ', 'ម៉ោងធ្វើការសរុប', 'វគ្គយប់', 'ម៉ោងយប់', 'ចំនួនថ្ងៃសុំច្បាប់', 'ចំនួនថ្ងៃបេសកម្ម']
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1a73e8', end_color='1a73e8', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    for row_idx, item in enumerate(summary_data, 4):
        ws2.cell(row=row_idx, column=1, value=row_idx-3)
        ws2.cell(row=row_idx, column=2, value=item.get('full_name', ''))
        ws2.cell(row=row_idx, column=3, value=item.get('days_worked', 0))

        total_hours = item.get('total_hours', 0)
        hours = int(total_hours)
        minutes = int((total_hours - hours) * 60)
        ws2.cell(row=row_idx, column=4, value=f"{hours:02d}:{minutes:02d}")

        ws2.cell(row=row_idx, column=5, value=item.get('night_shifts', 0))

        night_hours = item.get('night_hours', 0)
        nhours = int(night_hours)
        nminutes = int((night_hours - nhours) * 60)
        ws2.cell(row=row_idx, column=6, value=f"{nhours:02d}:{nminutes:02d}")

        ws2.cell(row=row_idx, column=7, value=item.get('total_leave_days', 0))
        ws2.cell(row=row_idx, column=8, value=item.get('total_mission_days', 0))

    for col in range(1, 9):
        ws2.column_dimensions[chr(64 + col)].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"របាយការណ៍_{start_date}_ដល់_{end_date}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/check_auto_unlock')
def check_auto_unlock():
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'})

    try:
        current_time = get_current_time_only()[:5]
        unlocked_items = []

        lock = get_system_lock_status()
        if lock.get('is_locked', 0) == 1:
            auto_unlock = lock.get('auto_unlock_time')
            if auto_unlock and current_time >= auto_unlock:
                update_system_lock(0, locked_by=None)
                unlocked_items.append('system')
                print(f"✅ System auto-unlocked via API at {current_time}")

        conn = get_db_connection()
        locked_users = conn.execute('''
            SELECT user_id, auto_unlock_time
            FROM user_attendance_lock
            WHERE is_locked = 1
            AND auto_unlock_time IS NOT NULL
            AND auto_unlock_time != ''
        ''').fetchall()
        conn.close()

        for user in locked_users:
            auto_unlock = user['auto_unlock_time']
            if auto_unlock and current_time >= auto_unlock:
                update_user_lock(user['user_id'], 0)
                unlocked_items.append(f"user_{user['user_id']}")
                print(f"✅ User {user['user_id']} auto-unlocked via API at {current_time}")

        if unlocked_items:
            increment_data_version()
            return jsonify({
                'success': True,
                'message': f'Auto-unlocked: {", ".join(unlocked_items)}',
                'unlocked': unlocked_items
            })
        else:
            return jsonify({
                'success': True,
                'message': 'No auto-unlock needed at this time',
                'unlocked': []
            })

    except Exception as e:
        print(f"❌ Error in check_auto_unlock: {e}")
        return jsonify({'success': False, 'message': str(e)})

# ============================================================
# PWA & STATIC ROUTES
# ============================================================

@app.route('/static/manifest.json')
def manifest():
    return {
        "name": "ប្រព័ន្ធគ្រប់គ្រងបុគ្គលិក",
        "short_name": "HR System",
        "start_url": "/dashboard",
        "display": "standalone",
        "background_color": "#1a73e8",
        "theme_color": "#1a73e8",
        "orientation": "portrait",
        "icons": [
            {
                "src": "/static/icon-192.png",
                "sizes": "192x192",
                "type": "image/png",
                "purpose": "any maskable"
            },
            {
                "src": "/static/icon-512.png",
                "sizes": "512x512",
                "type": "image/png",
                "purpose": "any maskable"
            }
        ]
    }

@app.route('/static/sw.js')
def service_worker():
    return '''
const CACHE_NAME = 'hr-system-v1';
const ASSETS = [
    '/',
    '/dashboard',
    '/static/manifest.json'
];

self.addEventListener('install', function(e) {
    e.waitUntil(
        caches.open(CACHE_NAME).then(function(cache) {
            return cache.addAll(ASSETS);
        }).then(function() {
            return self.skipWaiting();
        })
    );
});

self.addEventListener('activate', function(e) {
    e.waitUntil(
        caches.keys().then(function(cacheNames) {
            return Promise.all(
                cacheNames.filter(function(name) {
                    return name !== CACHE_NAME;
                }).map(function(name) {
                    return caches.delete(name);
                })
            );
        }).then(function() {
            return self.clients.claim();
        })
    );
});

self.addEventListener('fetch', function(e) {
    e.respondWith(
        caches.match(e.request).then(function(response) {
            return response || fetch(e.request).then(function(fetchResponse) {
                return caches.open(CACHE_NAME).then(function(cache) {
                    cache.put(e.request, fetchResponse.clone());
                    return fetchResponse;
                });
            });
        }).catch(function() {
            return new Response('Offline', {
                status: 503,
                statusText: 'Service Unavailable'
            });
        })
    );
});
    ''', 200, {'Content-Type': 'application/javascript'}

@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory('static', filename)

@app.route('/edit_attendance/<int:attendance_id>', methods=['GET', 'POST'])
def edit_attendance(attendance_id):
    if not session.get('logged_in') or session.get('role') != 'admin':
        return redirect(url_for('dashboard'))

    attendance = get_attendance_by_id(attendance_id)
    if not attendance:
        return "មិនមានទិន្នន័យ", 404

    if request.method == 'POST':
        check_in_date = request.form.get('check_in_date')
        check_in_time = request.form.get('check_in_time')
        check_out_date = request.form.get('check_out_date')
        check_out_time = request.form.get('check_out_time')
        shift = request.form.get('shift')

        check_in_full = None
        check_out_full = None

        if check_in_date and check_in_time:
            check_in_full = f"{check_in_date} {check_in_time}:00"
        if check_out_date and check_out_time:
            check_out_full = f"{check_out_date} {check_out_time}:00"

        result = update_attendance(attendance_id, check_in_full, check_out_full, check_in_date, shift)
        if result:
            return redirect(url_for('dashboard'))
        else:
            return "មិនអាចកែប្រែបាន!", 500

    edit_form_html = '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>កែប្រែទិន្នន័យវត្តមាន</title>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <style>
            * { margin: 0; padding: 0; box-sizing: border-box; }
            body { font-family: 'Khmer OS', 'Arial', sans-serif; background: #f0f2f5; padding: 20px; }
            .container { max-width: 600px; margin: 0 auto; background: white; padding: 30px; border-radius: 12px; box-shadow: 0 2px 10px rgba(0,0,0,0.08); }
            h2 { color: #1a73e8; margin-bottom: 20px; }
            .info { background: #f8f9fa; padding: 12px; border-radius: 8px; margin-bottom: 20px; }
            .info p { margin: 5px 0; }
            label { display: block; margin: 12px 0 5px; font-weight: 600; color: #555; }
            input, select { width: 100%; padding: 10px 14px; border: 2px solid #e8ecf1; border-radius: 8px; font-size: 14px; font-family: 'Khmer OS', 'Arial', sans-serif; }
            input:focus, select:focus { outline: none; border-color: #1a73e8; }
            .row { display: grid; grid-template-columns: 1fr 1fr; gap: 10px; }
            .btn-group { display: flex; gap: 10px; margin-top: 20px; flex-wrap: wrap; }
            .btn-save { flex: 1; background: #1a73e8; color: white; padding: 12px 20px; border: none; border-radius: 8px; cursor: pointer; font-size: 16px; font-family: 'Khmer OS', 'Arial', sans-serif; }
            .btn-save:hover { background: #1557b0; }
            .btn-cancel { flex: 1; background: #dc3545; color: white; padding: 12px 20px; border: none; border-radius: 8px; cursor: pointer; font-size: 16px; font-family: 'Khmer OS', 'Arial', sans-serif; text-decoration: none; text-align: center; }
            .btn-cancel:hover { background: #b02a37; }
            .hint { font-size: 12px; color: #888; margin-top: 3px; }
            @media (max-width: 600px) {
                .container { padding: 20px; }
                .row { grid-template-columns: 1fr; }
                .btn-group { flex-direction: column; }
            }
        </style>
    </head>
    <body>
        <div class="container">
            <h2>✏️ កែប្រែទិន្នន័យវត្តមាន</h2>
            <div class="info">
                <p><strong>ឈ្មោះបុគ្គលិក:</strong> {{ attendance.full_name }}</p>
                <p><strong>ឈ្មោះអ្នកប្រើ:</strong> {{ attendance.username }}</p>
                <p><strong>វគ្គបច្ចុប្បន្ន:</strong>
                    {% if attendance.shift == 1 %}វគ្គ 1 (ព្រឹក 07:00-11:00)
                    {% elif attendance.shift == 2 %}វគ្គ 2 (រសៀល 13:00-17:00)
                    {% elif attendance.shift == 3 %}វគ្គ 3 (យប់ 18:00-08:00)
                    {% else %}វគ្គ 1{% endif %}
                </p>
            </div>
            <form method="POST">
                <label>📅 កាលបរិច្ឆេទចូល</label>
                <input type="date" name="check_in_date" value="{{ attendance.date }}" required>

                <label>⏰ ម៉ោងចូល</label>
                <input type="time" name="check_in_time" value="{{ attendance.check_in.split(' ')[1][:5] if attendance.check_in else '' }}">

                <label>📅 កាលបរិច្ឆេទចេញ</label>
                <input type="date" name="check_out_date" value="{{ attendance.check_out.split(' ')[0] if attendance.check_out else attendance.date }}">

                <label>⏰ ម៉ោងចេញ</label>
                <input type="time" name="check_out_time" value="{{ attendance.check_out.split(' ')[1][:5] if attendance.check_out else '' }}">
                <div class="hint">💡 សម្រាប់វគ្គយប់ (វគ្គ 3) ប្រសិនបើចេញព្រឹកថ្ងៃបន្ទាប់ សូមកំណត់កាលបរិច្ឆេទចេញជាថ្ងៃបន្ទាប់</div>

                <label>វគ្គ</label>
                <select name="shift">
                    <option value="1" {% if attendance.shift == 1 %}selected{% endif %}>វគ្គ 1 (ព្រឹក 07:00-11:00)</option>
                    <option value="2" {% if attendance.shift == 2 %}selected{% endif %}>វគ្គ 2 (រសៀល 13:00-17:00)</option>
                    <option value="3" {% if attendance.shift == 3 %}selected{% endif %}>វគ្គ 3 (យប់ 18:00-08:00)</option>
                </select>

                <div class="btn-group">
                    <button type="submit" class="btn-save">💾 រក្សាទុក</button>
                    <a href="/dashboard" class="btn-cancel">❌ បោះបង់</a>
                </div>
            </form>
        </div>
    </body>
    </html>
    '''
    return render_template_string(edit_form_html, attendance=attendance)

@app.route('/delete_attendance/<int:attendance_id>', methods=['POST'])
def delete_attendance_route(attendance_id):
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'អ្នកមិនមែនជា Admin!'})

    result = delete_attendance(attendance_id)
    if result:
        return jsonify({'success': True, 'message': '✅ លុបទិន្នន័យជោគជ័យ!'})
    else:
        return jsonify({'success': False, 'message': '❌ មិនអាចលុបទិន្នន័យបាន!'})

@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory(os.path.join('static', 'uploads'), filename)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = get_user_by_username(username)
        if user and user['password'] == password:
            session['logged_in'] = True
            session['username'] = username
            session['user_id'] = user['id']
            session['role'] = user['role']
            session['full_name'] = user['full_name']
            return redirect(url_for('dashboard'))
        else:
            return '''
            <!DOCTYPE html>
            <html>
            <head><title>កំហុស</title><meta charset="UTF-8">
            <style>
                body { font-family:'Khmer OS',Arial; text-align:center; padding:50px; background:#f0f2f5; }
                .box { background:white; padding:40px; border-radius:16px; max-width:400px; margin:0 auto; box-shadow:0 4px 20px rgba(0,0,0,0.08); }
                h3 { color:#dc3545; }
                .back-link { display:inline-block; margin-top:20px; color:#1a73e8; text-decoration:none; padding:10px 25px; border:2px solid #1a73e8; border-radius:10px; }
                .back-link:hover { background:#1a73e8; color:white; }
            </style>
            </head>
            <body>
                <div class="box">
                    <h3>❌ ឈ្មោះអ្នកប្រើ ឬ ពាក្យសម្ងាត់មិនត្រឹមត្រូវ!</h3>
                    <a href="/login" class="back-link">← ត្រលប់មកវិញ</a>
                </div>
            </body>
            </html>
            '''
    return '''
<!DOCTYPE html>
<html>
<head>
    <title>ចូលប្រើប្រព័ន្ធ</title>
    <meta charset="UTF-8">
    <link rel="manifest" href="/static/manifest.json">
    <meta name="theme-color" content="#1a73e8">
    <meta name="apple-mobile-web-app-capable" content="yes">
    <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
    <meta name="apple-mobile-web-app-title" content="HR System">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        html, body {
            height: 100%;
            width: 100%;
            font-family: 'Khmer OS', Arial, sans-serif;
            background: #ffffff;
            margin: 0;
            padding: 0;
            overflow: hidden;
        }
        .login-container {
            display: flex;
            justify-content: center;
            align-items: center;
            height: 100vh;
            width: 100vw;
            padding: 15px;
            background: #ffffff;
        }
        .login-box {
            width: 100%;
            height: 100%;
            max-width: 480px;
            max-height: 600px;
            background: #ffffff;
            padding: 40px 30px;
            border-radius: 24px;
            box-shadow: 0 10px 40px rgba(0, 0, 0, 0.06);
            border: 1px solid #f0f0f0;
            display: flex;
            flex-direction: column;
            justify-content: center;
            animation: fadeInUp 0.5s ease;
        }
        @keyframes fadeInUp {
            from { opacity: 0; transform: translateY(20px); }
            to { opacity: 1; transform: translateY(0); }
        }
        .logo-icon {
            text-align: center;
            font-size: 65px;
            margin-bottom: 8px;
        }
        .login-box h2 {
            text-align: center;
            color: #1a1a2e;
            font-size: 30px;
            font-weight: 700;
            margin-bottom: 4px;
        }
        .login-box .sub-title {
            text-align: center;
            color: #888;
            font-size: 15px;
            margin-bottom: 28px;
        }
        .login-box .form-group {
            margin-bottom: 16px;
        }
        .login-box input {
            width: 100%;
            padding: 16px 18px;
            border: 2px solid #e8ecf1;
            border-radius: 14px;
            font-size: 17px;
            box-sizing: border-box;
            font-family: 'Khmer OS', Arial, sans-serif;
            transition: all 0.3s;
            background: #f8f9fa;
        }
        .login-box input:focus {
            outline: none;
            border-color: #1a73e8;
            background: white;
            box-shadow: 0 0 0 4px rgba(26, 115, 232, 0.08);
        }
        .login-box button {
            width: 100%;
            padding: 16px;
            background: #1a73e8;
            color: white;
            border: none;
            border-radius: 14px;
            font-size: 18px;
            cursor: pointer;
            font-family: 'Khmer OS', Arial, sans-serif;
            font-weight: 600;
            transition: all 0.3s;
            margin-top: 4px;
        }
        .login-box button:hover {
            transform: translateY(-2px);
            box-shadow: 0 8px 25px rgba(26, 115, 232, 0.35);
        }
        .login-box button:active { transform: scale(0.98); }
        .login-box .hint {
            text-align: center;
            margin-top: 16px;
            color: #999;
            font-size: 13px;
        }
        .login-box .hint b { color: #1a73e8; }
        .login-box .footer-text {
            text-align: center;
            margin-top: 20px;
            padding-top: 16px;
            border-top: 1px solid #eee;
            color: #ccc;
            font-size: 12px;
        }
    </style>
</head>
<body>
    <div class="login-container">
        <div class="login-box">
            <div class="logo-icon">🏢</div>
            <h2 style="font-family: 'Khmer', 'Khmer OS Muol Light', 'Khmer OS', Arial, sans-serif; font-weight: 300; color: #1a73e8;">ប្រព័ន្ធគ្រប់គ្រងបុគ្គលិក</h2>
            <div class="sub-title">សូមបញ្ចូលឈ្មោះ និងពាក្យសម្ងាត់</div>
            <form method="POST" style="flex:1;display:flex;flex-direction:column;justify-content:center;">
                <div class="form-group">
                    <input type="text" name="username" placeholder="ឈ្មោះអ្នកប្រើ" required>
                </div>
                <div class="form-group">
                    <input type="password" name="password" placeholder="ពាក្យសម្ងាត់" required>
                </div>
                <button type="submit">🔐 Login</button>
            </form>
            <div class="hint"><b>ពត៌មានបន្ថែមៈ ទំនាក់ទំនងលោក YEN SONY</b></div>
            <div class="hint"><b>ទូរស័ព្ទ៖ +855 92 740 067</b></div>
            <div class="footer-text">© 2026 ប្រព័ន្ធគ្រប់គ្រងបុគ្គលិក</div>
        </div>
    </div>
</body>
</html>
'''

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

DASHBOARD_HTML = r'''
<!DOCTYPE html>
<html>
<head>
    <title>ប្រព័ន្ធគ្រប់គ្រងបុគ្គលិក</title>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
    <link rel="manifest" href="/static/manifest.json">
    <meta name="theme-color" content="#1a73e8">
    <meta name="apple-mobile-web-app-capable" content="yes">
    <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
    <meta name="apple-mobile-web-app-title" content="HR System">
    <link href="https://fonts.googleapis.com/css2?family=Khmer&display=swap" rel="stylesheet">
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif; background: #f0f2f5; padding: 15px; min-height: 100vh; }
        .container { max-width: 1200px; margin: 0 auto; }
        .header {
            background: linear-gradient(135deg, #1a73e8, #0d47a1);
            color: white;
            padding: 12px 20px;
            border-radius: 12px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            box-shadow: 0 4px 15px rgba(26, 115, 232, 0.3);
            margin-bottom: 15px;
            flex-wrap: wrap;
            gap: 10px;
        }
        .header-left h2 { font-size: 18px; font-weight: 600; white-space: nowrap; }
        .header-right {
            display: flex;
            align-items: center;
            gap: 8px;
            flex-wrap: wrap;
        }
        .header-right .nav-btn {
            color: white;
            text-decoration: none;
            padding: 6px 16px;
            border-radius: 20px;
            font-size: 13px;
            font-weight: 500;
            transition: all 0.3s;
            background: rgba(255,255,255,0.1);
            border: 1px solid rgba(255,255,255,0.15);
            white-space: nowrap;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
            cursor: pointer;
        }
        .header-right .nav-btn:hover { background: rgba(255,255,255,0.25); transform: translateY(-2px); }
        .header-right .nav-btn.location-btn {
            background: rgba(255, 193, 7, 0.3);
            border-color: #ffc107;
        }
        .header-right .nav-btn.location-btn:hover {
            background: rgba(255, 193, 7, 0.5);
        }
        .header-right .user-name-btn {
            color: white;
            padding: 6px 16px;
            border-radius: 20px;
            font-size: 13px;
            font-weight: 500;
            background: rgba(255,255,255,0.2);
            border: 1px solid rgba(255,255,255,0.25);
            white-space: nowrap;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
            cursor: default;
            display: flex;
            align-items: center;
            gap: 6px;
        }
        .header-right .logout-link {
            color: white;
            text-decoration: none;
            background: rgba(255,255,255,0.2);
            padding: 6px 16px;
            border-radius: 20px;
            font-size: 13px;
            transition: all 0.3s;
            font-weight: 500;
            border: 1px solid rgba(255,255,255,0.15);
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
        }
        .header-right .logout-link:hover { background: rgba(255,255,255,0.35); transform: scale(1.05); }
        .stats {
            display: grid;
            grid-template-columns: repeat(4, 1fr);
            gap: 15px;
            margin-bottom: 20px;
        }
        .stat-card {
            background: white;
            padding: 18px 15px;
            border-radius: 12px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.06);
            text-align: center;
            transition: all 0.3s;
        }
        .stat-card:hover { transform: translateY(-2px); box-shadow: 0 6px 20px rgba(0,0,0,0.1); }
        .stat-card .stat-icon { font-size: 28px; display: block; margin-bottom: 4px; }
        .stat-card .number { font-size: 30px; font-weight: 700; color: #1a73e8; }
        .stat-card .number.green { color: #34a853; }
        .stat-card .number.orange { color: #fbbc04; }
        .stat-card .number.purple { color: #7c3aed; }
        .stat-card .label { font-size: 13px; color: #888; margin-top: 2px; }
        .action-buttons {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 12px;
            margin-bottom: 20px;
        }
        .action-buttons .btn {
            padding: 14px 10px;
            border: none;
            border-radius: 12px;
            cursor: pointer;
            font-size: 16px;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
            transition: all 0.3s;
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            gap: 4px;
            font-weight: 500;
            min-height: 65px;
            color: white;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        }
        .action-buttons .btn .icon { font-size: 28px; }
        .action-buttons .btn .sub-text { font-size: 12px; opacity: 0.85; font-weight: 400; }
        .action-buttons .btn:hover { transform: translateY(-3px); box-shadow: 0 6px 25px rgba(0,0,0,0.15); }
        .action-buttons .btn:active { transform: scale(0.95); }
        .action-buttons .btn.disabled {
            opacity: 0.5;
            cursor: not-allowed;
            transform: none !important;
        }
        .btn-checkin { background: linear-gradient(135deg, #34a853, #1e7e34); }
        .btn-checkout { background: linear-gradient(135deg, #dc3545, #b02a37); }
        .btn-leave { background: linear-gradient(135deg, #fbbc04, #e5a800); color: #333; }
        .btn-mission { background: linear-gradient(135deg, #1a73e8, #0d47a1); }
        .table-container {
            background: white;
            border-radius: 12px;
            padding: 20px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.06);
            overflow-x: auto;
        }
        .table-container .table-title {
            font-size: 18px;
            font-weight: 600;
            color: #333;
            margin-bottom: 15px;
            display: flex;
            align-items: center;
            gap: 10px;
        }
        .table-container .table-title .badge-count {
            background: #1a73e8;
            color: white;
            padding: 2px 12px;
            border-radius: 20px;
            font-size: 14px;
        }
        table { width: 100%; border-collapse: collapse; }
        table thead th {
            background: #f8f9fa;
            color: #555;
            padding: 12px 15px;
            text-align: left;
            font-size: 13px;
            font-weight: 600;
            border-bottom: 2px solid #e8ecf1;
        }
        table tbody td {
            padding: 12px 15px;
            border-bottom: 1px solid #f0f2f5;
            color: #333;
            font-size: 14px;
        }
        table tbody tr:hover { background: #f8f9fa; }
        .status-badge {
            display: inline-block;
            padding: 4px 14px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: 600;
        }
        .status-badge.present { background: #e8f5e9; color: #2e7d32; }
        .status-badge.outside { background: #ffebee; color: #c62828; }
        .status-badge.working { background: #fff3cd; color: #856404; }
        .status-badge.leave { background: #fff3cd; color: #856404; }
        .status-badge.mission { background: #d1ecf1; color: #0c5460; }
        .status-badge.approved { background: #d4edda; color: #155724; }
        .shift-badge {
            display: inline-block;
            padding: 2px 10px;
            border-radius: 12px;
            font-size: 11px;
            font-weight: 600;
        }
        .shift-1 { background: #e3f2fd; color: #0d47a1; }
        .shift-2 { background: #fff3e0; color: #e65100; }
        .shift-3 { background: #f3e5f5; color: #4a148c; }
        .shift-leave { background: #fff3cd; color: #856404; }
        .shift-mission { background: #d1ecf1; color: #0c5460; }
        .footer { text-align: center; padding: 20px 0 5px; color: #aaa; font-size: 12px; }
        .modal {
            display: none;
            position: fixed;
            top: 0; left: 0; width: 100%; height: 100%;
            background: rgba(0,0,0,0.5);
            justify-content: center;
            align-items: center;
            z-index: 999;
        }
        .modal.show { display: flex; }
        .modal-content {
            background: white;
            padding: 30px;
            border-radius: 16px;
            max-width: 500px;
            width: 90%;
            box-shadow: 0 10px 40px rgba(0,0,0,0.3);
            max-height: 90vh;
            overflow-y: auto;
        }
        .modal-content h3 { color: #333; margin-bottom: 10px; }
        .modal-content p { color: #666; font-size: 14px; margin-bottom: 15px; }
        .modal-content input, .modal-content textarea, .modal-content select {
            width: 100%;
            padding: 10px 14px;
            border: 2px solid #e8ecf1;
            border-radius: 10px;
            font-size: 14px;
            margin-bottom: 10px;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
        }
        .modal-content input:focus, .modal-content textarea:focus, .modal-content select:focus {
            outline: none;
            border-color: #1a73e8;
        }
        .modal-content .btn-group {
            display: flex;
            gap: 10px;
            flex-wrap: wrap;
        }
        .modal-content .btn-group button {
            flex: 1;
            padding: 12px;
            border: none;
            border-radius: 10px;
            cursor: pointer;
            font-size: 15px;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
            font-weight: 500;
            min-width: 80px;
        }
        .btn-save { background: #1a73e8; color: white; }
        .btn-save:hover { background: #1557b0; }
        .btn-success { background: #34a853; color: white; }
        .btn-success:hover { background: #1e7e34; }
        .btn-danger { background: #dc3545; color: white; }
        .btn-danger:hover { background: #b02a37; }
        .btn-cancel { background: #e8ecf1; color: #333; }
        .btn-cancel:hover { background: #d5d8dd; }
        .btn-get-location {
            background: #34a853;
            color: white;
            padding: 10px;
            border: none;
            border-radius: 10px;
            cursor: pointer;
            font-size: 14px;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
            width: 100%;
            margin-bottom: 10px;
        }
        .btn-get-location:hover { background: #1e7e34; }
        .location-info {
            background: #f8f9fa;
            padding: 10px;
            border-radius: 8px;
            font-size: 13px;
            color: #555;
            margin-bottom: 10px;
        }
        .request-item {
            background: #f8f9fa;
            padding: 12px 15px;
            border-radius: 10px;
            margin-bottom: 10px;
            border-left: 4px solid #ffc107;
        }
        .request-item .request-user { font-weight: 600; color: #333; }
        .request-item .request-detail { font-size: 13px; color: #666; margin-top: 3px; }
        .request-item .request-actions { margin-top: 8px; display: flex; gap: 8px; }
        .request-item .request-actions button {
            padding: 5px 15px;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            font-size: 13px;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
        }
        .request-item .request-actions .approve-btn {
            background: #34a853;
            color: white;
        }
        .request-item .request-actions .approve-btn:hover {
            background: #1e7e34;
        }
        .request-item .request-actions .reject-btn {
            background: #dc3545;
            color: white;
        }
        .request-item .request-actions .reject-btn:hover {
            background: #b02a37;
        }
        .status-message {
            padding: 10px 15px;
            border-radius: 8px;
            margin-bottom: 10px;
            font-size: 14px;
            display: none;
        }
        .status-message.info {
            background: #d1ecf1;
            color: #0c5460;
            border: 1px solid #bee5eb;
            display: block;
        }
        .status-message.success {
            background: #d4edda;
            color: #155724;
            border: 1px solid #c3e6cb;
            display: block;
        }
        .status-message.warning {
            background: #fff3cd;
            color: #856404;
            border: 1px solid #ffc107;
            display: block;
        }
        .status-message.danger {
            background: #f8d7da;
            color: #721c24;
            border: 1px solid #f5c6cb;
            display: block;
        }
        .status-message.locked {
            background: #f8d7da;
            color: #721c24;
            border: 2px solid #dc3545;
            display: block;
            font-weight: 600;
        }
        .install-btn {
            background: #34a853;
            color: white;
            border: none;
            border-radius: 20px;
            padding: 6px 16px;
            cursor: pointer;
            font-size: 13px;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
            font-weight: 500;
            display: none;
            animation: pulse-green 2s infinite;
        }
        @keyframes pulse-green {
            0% { box-shadow: 0 0 0 0 rgba(52, 168, 83, 0.4); }
            70% { box-shadow: 0 0 0 10px rgba(52, 168, 83, 0); }
            100% { box-shadow: 0 0 0 0 rgba(52, 168, 83, 0); }
        }
        .install-btn:hover { background: #1e7e34; transform: scale(1.05); }
        .summary-grid {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 15px;
            margin-top: 15px;
        }
        .summary-card {
            background: white;
            border-radius: 12px;
            padding: 15px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.06);
        }
        .summary-card .card-title {
            font-size: 15px;
            font-weight: 600;
            color: #555;
            margin-bottom: 10px;
            padding-bottom: 8px;
            border-bottom: 2px solid #f0f2f5;
        }
        .summary-item {
            display: flex;
            justify-content: space-between;
            padding: 6px 0;
            border-bottom: 1px solid #f5f6f8;
            font-size: 14px;
        }
        .summary-item .name { color: #333; }
        .summary-item .value { font-weight: 600; color: #1a73e8; }
        .summary-item .value.leave { color: #fbbc04; }
        .summary-item .value.mission { color: #7c3aed; }
        .summary-item .value.hours { color: #34a853; }
        .deadline-badge {
            display: inline-block;
            padding: 2px 10px;
            border-radius: 12px;
            font-size: 11px;
            font-weight: 600;
            background: #ff9800;
            color: white;
        }
        .deadline-badge.active {
            background: #4caf50;
        }
        .deadline-badge.inactive {
            background: #9e9e9e;
        }
        .user-lock-badge {
            display: inline-block;
            padding: 2px 10px;
            border-radius: 12px;
            font-size: 11px;
            font-weight: 600;
            background: #dc3545;
            color: white;
        }
        .user-lock-badge.unlocked {
            background: #4caf50;
        }
        .lock-banner {
            background: linear-gradient(135deg, #dc3545, #b02a37);
            color: white;
            padding: 12px 20px;
            border-radius: 12px;
            display: none;
            align-items: center;
            justify-content: space-between;
            margin-bottom: 15px;
            flex-wrap: wrap;
            gap: 10px;
        }
        .lock-banner.show { display: flex; }
        .lock-banner .lock-icon { font-size: 24px; margin-right: 10px; }
        .lock-banner .lock-text {
            font-size: 16px;
            font-weight: 500;
            flex: 1;
        }
        .lock-banner .lock-time {
            font-size: 13px;
            opacity: 0.8;
        }
        .lock-banner .lock-btn {
            background: rgba(255,255,255,0.2);
            border: 1px solid rgba(255,255,255,0.3);
            color: white;
            padding: 6px 16px;
            border-radius: 20px;
            cursor: pointer;
            font-size: 13px;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
            transition: all 0.3s;
        }
        .lock-banner .lock-btn:hover {
            background: rgba(255,255,255,0.35);
        }
        .lock-banner.unlocked {
            background: linear-gradient(135deg, #34a853, #1e7e34);
        }
        .user-lock-banner {
            background: linear-gradient(135deg, #ff9800, #e65100);
            color: white;
            padding: 10px 16px;
            border-radius: 12px;
            display: none;
            align-items: center;
            justify-content: space-between;
            margin-bottom: 15px;
            flex-wrap: wrap;
            gap: 8px;
        }
        .user-lock-banner.show { display: flex; }
        .user-lock-banner .lock-icon { font-size: 20px; margin-right: 8px; }
        .user-lock-banner .lock-text {
            font-size: 14px;
            font-weight: 500;
            flex: 1;
        }
        .user-lock-banner .lock-time {
            font-size: 12px;
            opacity: 0.8;
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <div class="header-left">
                <h2>🏢 ប្រព័ន្ធគ្រប់គ្រងបុគ្គលិក</h2>
            </div>
            <div class="header-right">
                <button id="installBtn" class="install-btn">📲 ដំឡើងកម្មវិធី</button>
                <a href="/report" class="nav-btn">📊 របាយការណ៍</a>
                {% if session.role == 'admin' %}
                <a href="/register" class="nav-btn">📝 ចុះឈ្មោះ</a>
                <a href="/manage_users" class="nav-btn">👤 គ្រប់គ្រងអ្នកប្រើ</a>
                <button class="nav-btn location-btn" onclick="openLocationModal()">📍 កំណត់ទីតាំង</button>
                <button class="nav-btn" onclick="openCleanModal()" style="background:rgba(220,53,69,0.2);border-color:#dc3545;">🧹 សម្អាតទិន្នន័យ</button>
                <button class="nav-btn" onclick="openRequestsModal()" style="background:rgba(255,193,7,0.2);border-color:#ffc107;">
                    📬 សំណើ
                    {% if pending_count > 0 %}
                    <span style="background:red;color:white;border-radius:50%;padding:0 6px;font-size:11px;margin-left:3px;">{{ pending_count }}</span>
                    {% endif %}
                </button>
                <button class="nav-btn" onclick="openSystemLockModal()" style="background:rgba(33,150,243,0.2);border-color:#2196f3;">
                    🔒 បិទ/បើកប្រព័ន្ធ
                </button>
                {% endif %}
                <a href="/change_password" class="nav-btn">🔑 ប្តូរពាក្យសម្ងាត់</a>
                <span class="user-name-btn">👤 {{ session.username }}</span>
                <a href="/logout" class="logout-link">🚪 ចាកចេញ</a>
            </div>
        </div>

        <!-- User Lock Banner -->
        <div id="userLockBanner" class="user-lock-banner {% if user_lock.is_locked == 1 %}show{% endif %}">
            <div style="display:flex;align-items:center;gap:8px;">
                <span class="lock-icon">🔒</span>
                <span class="lock-text" id="userLockText">
                    {% if user_lock.is_locked == 1 %}
                    អ្នកត្រូវបានបិទការចូលធ្វើការដោយ Admin
                    {% if user_lock.auto_unlock_time %}
                    (បើកដោយស្វ័យប្រវត្តិនៅម៉ោង {{ user_lock.auto_unlock_time }})
                    {% endif %}
                    {% endif %}
                </span>
            </div>
        </div>

        <!-- System Lock Banner -->
        <div id="lockBanner" class="lock-banner show">
            <div style="display:flex;align-items:center;gap:10px;">
                <span class="lock-icon">🔒</span>
                <span class="lock-text" id="lockText">ប្រព័ន្ធកំពុងបិទការចូលធ្វើការ</span>
                <span class="lock-time" id="lockTime">បើកដោយស្វ័យប្រវត្តិនៅម៉ោង 06:00</span>
            </div>
            {% if session.role == 'admin' %}
            <button class="lock-btn" id="lockToggleBtn" onclick="openSystemLockModal()">🔒 បិទ/បើកប្រព័ន្ធ</button>
            {% endif %}
        </div>

        <div id="statusMessage" class="status-message"></div>

        <div class="stats">
            <div class="stat-card">
                <span class="stat-icon">👥</span>
                <div class="number">{{ stats.total_users }}</div>
                <div class="label">បុគ្គលិកសរុប</div>
            </div>
            <div class="stat-card">
                <span class="stat-icon">✅</span>
                <div class="number green">{{ stats.present_today }}</div>
                <div class="label">កំពុងធ្វើការ</div>
            </div>
            <div class="stat-card">
                <span class="stat-icon">📋</span>
                <div class="number orange">{{ stats.leave_today }}</div>
                <div class="label">សុំច្បាប់ (Approved)</div>
            </div>
            <div class="stat-card">
                <span class="stat-icon">🚗</span>
                <div class="number purple">{{ stats.mission_today }}</div>
                <div class="label">ចុះបេសកម្ម (Approved)</div>
            </div>
        </div>

        <div class="action-buttons" id="actionButtons">
            <button class="btn btn-checkin" onclick="openCheckInModal()" id="checkinBtn">
                <span class="icon">✅</span>
                ចូលធ្វើការ
                <span class="sub-text">ជ្រើសរើសវគ្គ</span>
            </button>
            <button class="btn btn-checkout" onclick="handleCheckOut()" id="checkoutBtn">
                <span class="icon">🚪</span>
                ចេញពីធ្វើការ
                <span class="sub-text">ចុចដើម្បីចេញ</span>
            </button>
            <button class="btn btn-leave" onclick="handleLeave()">
                <span class="icon">📋</span>
                សុំច្បាប់
                <span class="sub-text">ស្នើសុំច្បាប់</span>
            </button>
            <button class="btn btn-mission" onclick="handleMission()">
                <span class="icon">🚗</span>
                បេសកម្ម
                <span class="sub-text">ស្នើសុំបេសកម្ម</span>
            </button>
        </div>

        <div class="table-container">
            <div class="table-title">
                📋 ប្រវត្តិការងារបុគ្គលិក
                <span class="badge-count">{{ work_history|length }}</span>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>ល.រ</th>
                        <th>ឈ្មោះ</th>
                        <th>កាលបរិច្ឆេទ</th>
                        <th>វគ្គ</th>
                        <th>ម៉ោងចូល</th>
                        <th>ម៉ោងចេញ</th>
                        <th>ម៉ោង/ថ្ងៃ</th>
                        <th>ស្ថានភាព</th>
                        <th>ព័ត៌មាន</th>
                        {% if session.role == 'admin' %}
                        <th>សកម្មភាព</th>
                        {% endif %}
                    </tr>
                </thead>
                <tbody>
    {% if work_history %}
        {% for item in work_history %}
        <tr>
            <td>{{ loop.index }}</td>
            <td>{{ item.full_name }}</td>
            <td>{{ item.date }}</td>
            <td>
                {% if item.type == 'attendance' %}
                    {% if item.shift == 'វគ្គ 1' %}
                    <span class="shift-badge shift-1">វគ្គ 1</span>
                    {% elif item.shift == 'វគ្គ 2' %}
                    <span class="shift-badge shift-2">វគ្គ 2</span>
                    {% elif item.shift == 'វគ្គ 3' %}
                    <span class="shift-badge shift-3">វគ្គ 3</span>
                    {% else %}
                    <span class="shift-badge shift-1">{{ item.shift }}</span>
                    {% endif %}
                {% elif item.type == 'leave' %}
                    <span class="shift-badge shift-leave">📋 ច្បាប់</span>
                {% elif item.type == 'mission' %}
                    <span class="shift-badge shift-mission">🚗 បេសកម្ម</span>
                {% endif %}
            </td>
            <td>{{ item.check_in_time or '' }}</td>
            <td>{{ item.check_out_time or '' }}</td>
            <td>{{ item.display_value or '' }}</td>
            <td>
                {% if item.type == 'attendance' %}
                    {% if item.status == 'បានបិទ' %}
                    <span class="status-badge present">បានបិទ</span>
                    {% elif item.status == 'កំពុងធ្វើការ' %}
                    <span class="status-badge working">កំពុងធ្វើការ</span>
                    {% else %}
                    <span class="status-badge outside">មិនទាន់ចូល</span>
                    {% endif %}
                {% else %}
                    <span class="status-badge approved">បានអនុម័ត</span>
                {% endif %}
            </td>
            <td>{{ item.display_info or '' }}</td>
            {% if session.role == 'admin' %}
            <td>
                {% if item.type == 'attendance' %}
                <a href="/edit_attendance/{{ item.id }}"
                   style="background:#1a73e8;color:white;padding:4px 12px;border-radius:6px;text-decoration:none;font-size:12px;display:inline-block;margin:2px;">
                   ✏️ កែ
                </a>
                <button onclick="deleteAttendance({{ item.id }})"
                        style="background:#dc3545;color:white;padding:4px 12px;border:none;border-radius:6px;cursor:pointer;font-size:12px;margin:2px;">
                   🗑️ លុប
                </button>
                {% else %}
                <span style="font-size:11px;color:#999;">មិនអាចកែ</span>
                {% endif %}
            </td>
            {% endif %}
        </tr>
        {% endfor %}
    {% else %}
        <tr>
            <td colspan="{% if session.role == 'admin' %}10{% else %}9{% endif %}"
                style="text-align:center;padding:30px;color:#aaa;">
                📭 មិនមានទិន្នន័យសម្រាប់ខែនេះ
            </td>
        </tr>
    {% endif %}
    </tbody>
            </table>
        </div>

        <div class="summary-grid">
            <div class="summary-card">
                <div class="card-title">📊 សង្ខេបបុគ្គលិក</div>
                <div class="summary-item" style="font-weight:600;color:#555;border-bottom:2px solid #e8ecf1;">
                    <span>ឈ្មោះ</span>
                    <span>ថ្ងៃធ្វើការ</span>
                </div>
                {% for emp in monthly_summary %}
                <div class="summary-item">
                    <span class="name">{{ emp.full_name }}</span>
                    <span class="value">{{ emp.days_worked_display }}</span>
                </div>
                {% endfor %}
                {% if not monthly_summary %}
                <div style="text-align:center;padding:15px;color:#aaa;">មិនមានទិន្នន័យ</div>
                {% endif %}
            </div>
            <div class="summary-card">
                <div class="card-title">📋 សង្ខេបវត្តមាន និងច្បាប់</div>
                <div class="summary-item" style="font-weight:600;color:#555;border-bottom:2px solid #e8ecf1;">
                    <span>ឈ្មោះ</span>
                    <span style="display:flex;gap:15px;">
                        <span>ម៉ោង</span>
                        <span>ច្បាប់</span>
                        <span>បេសកម្ម</span>
                    </span>
                </div>
                {% for emp in monthly_summary %}
                <div class="summary-item">
                    <span class="name">{{ emp.full_name }}</span>
                    <span style="display:flex;gap:50px;font-weight:600;">
                        <span class="value hours">{{ emp.total_hours_formatted }}</span>
                        <span class="value leave">{{ emp.total_leave_days }}</span>
                        <span class="value mission">{{ emp.total_mission_days }}</span>
                    </span>
                </div>
                {% endfor %}
                {% if not monthly_summary %}
                <div style="text-align:center;padding:15px;color:#aaa;">មិនមានទិន្នន័យ</div>
                {% endif %}
            </div>
        </div>

        <div class="footer">
            © 2026 ប្រព័ន្ធគ្រប់គ្រងបុគ្គលិក | រក្សាសិទ្ធិគ្រប់យ៉ាង
        </div>
    </div>

    <!-- Modals & Scripts -->
    <div id="checkInModal" class="modal">
        <div class="modal-content">
            <h3>✅ ចូលធ្វើការ</h3>
            <p>សូមជ្រើសរើសវគ្គដែលអ្នកចង់ចូល</p>
            <div id="checkInLocationInfo" class="location-info">
                ⏳ កំពុងចាប់យកទីតាំង...
            </div>
            <select id="shiftSelect">
                <option value="1">វគ្គ 1 (ព្រឹក 07:00-11:00)</option>
                <option value="2">វគ្គ 2 (រសៀល 13:00-17:00)</option>
                <option value="3">វគ្គ 3 (យប់ 18:00-08:00)</option>
            </select>
            <div class="btn-group">
                <button class="btn-success" onclick="submitCheckIn()">✅ ចូលធ្វើការ</button>
                <button class="btn-cancel" onclick="closeModal('checkInModal')">បោះបង់</button>
            </div>
        </div>
    </div>

    <div id="locationModal" class="modal">
        <div class="modal-content">
            <h3>📍 កំណត់ទីតាំងក្រុមហ៊ុន</h3>
            <p>សូមកំណត់ទីតាំងក្រុមហ៊ុនសម្រាប់ពិនិត្យមើលវត្តមានបុគ្គលិក (ចម្ងាយអនុញ្ញាត {{ allowed_distance }} ម៉ែត្រ)</p>
            <div id="locationInfo" class="location-info">
                {% if company_lat and company_lng %}
                📍 ទីតាំងបច្ចុប្បន្ន: {{ company_lat }}, {{ company_lng }}
                {% else %}
                ⚠️ មិនទាន់មានកំណត់ទីតាំង
                {% endif %}
            </div>
            <button class="btn-get-location" onclick="getCurrentLocation()">📡 ចាប់យកទីតាំងបច្ចុប្បន្ន</button>
            <input type="text" id="companyLat" placeholder="រយៈទទឹង (Latitude)" value="{{ company_lat or '' }}">
            <input type="text" id="companyLng" placeholder="រយៈបណ្តោយ (Longitude)" value="{{ company_lng or '' }}">
            <div class="btn-group">
                <button class="btn-save" onclick="saveCompanyLocation()">💾 រក្សាទុក</button>
                <button class="btn-cancel" onclick="closeLocationModal()">បោះបង់</button>
            </div>
        </div>
    </div>

    <div id="systemLockModal" class="modal">
        <div class="modal-content">
            <h3>🔒 បិទ/បើកប្រព័ន្ធចូលធ្វើការ</h3>
            <p style="color:#dc3545;font-weight:600;">⚠️ ពេលបិទប្រព័ន្ធ បុគ្គលិកនឹងមិនអាចចុចចូល/ចេញធ្វើការបានទេ!</p>
            <div id="lockStatusInfo" class="location-info">
                <span id="lockStatusText">ប្រព័ន្ធកំពុងបើក</span>
            </div>
            <div style="margin-bottom:15px;">
                <label style="font-weight:600;color:#555;font-size:14px;">ម៉ោងបើកដោយស្វ័យប្រវត្តិ (HH:MM)</label>
                <input type="time" id="autoUnlockTime" value="06:00" step="60">
                <div style="font-size:12px;color:#888;margin-top:4px;">
                    🔓 ពេលដល់ម៉ោងនេះ ប្រព័ន្ធនឹងបើកដោយស្វ័យប្រវត្តិ
                </div>
            </div>
            <div class="btn-group">
                <button id="lockToggleBtnModal" class="btn-danger" onclick="toggleSystemLockFromModal(1)">🔒 បិទប្រព័ន្ធឥឡូវនេះ</button>
                <button class="btn-success" onclick="toggleSystemLockFromModal(0)" id="unlockBtnModal" style="display:none;">🔓 បើកប្រព័ន្ធឥឡូវនេះ</button>
                <button class="btn-cancel" onclick="closeModal('systemLockModal')">បោះបង់</button>
            </div>
            <div id="unlockScheduledInfo" style="display:none;margin-top:12px;padding:10px;background:#fff3cd;border-radius:8px;color:#856404;font-size:14px;">
                ⏰ ប្រព័ន្ធនឹងបើកដោយស្វ័យប្រវត្តិនៅម៉ោង <span id="scheduledUnlockTime">06:00</span>
            </div>
        </div>
    </div>

    <div id="requestsModal" class="modal">
        <div class="modal-content">
            <h3>📬 សំណើរង់ចាំការអនុម័ត</h3>
            <div id="requestsList">
                <p style="text-align:center;color:#888;padding:20px;">កំពុងផ្ទុក...</p>
            </div>
            <div class="btn-group">
                <button class="btn-cancel" onclick="closeRequestsModal()" style="width:100%;">បិទ</button>
            </div>
        </div>
    </div>

    <div id="cleanModal" class="modal">
        <div class="modal-content" style="max-width:500px;">
            <h3>🧹 សម្អាតទិន្នន័យ</h3>
            <p style="color:#dc3545;font-weight:600;margin-bottom:15px;">⚠️ ការសម្អាតទិន្នន័យមិនអាចស្តារឡើងវិញបានទេ!</p>
            <p style="margin-bottom:15px;color:#666;">សូមជ្រើសរើសប្រភេទទិន្នន័យដែលចង់សម្អាត៖</p>
            <div style="display:flex;flex-direction:column;gap:10px;">
                <button onclick="cleanData('all')" style="padding:12px;background:#dc3545;color:white;border:none;border-radius:8px;cursor:pointer;font-size:15px;font-family:'Khmer OS','Khmer OS Muol','Arial',sans-serif;">🗑️ សម្អាតទិន្នន័យទាំងអស់</button>
                <button onclick="cleanData('attendance')" style="padding:12px;background:#fbbc04;color:#333;border:none;border-radius:8px;cursor:pointer;font-size:15px;font-family:'Khmer OS','Khmer OS Muol','Arial',sans-serif;">🗑️ សម្អាតតែវត្តមាន</button>
                <button onclick="cleanData('leaves')" style="padding:12px;background:#fbbc04;color:#333;border:none;border-radius:8px;cursor:pointer;font-size:15px;font-family:'Khmer OS','Khmer OS Muol','Arial',sans-serif;">🗑️ សម្អាតតែសុំច្បាប់</button>
                <button onclick="cleanData('missions')" style="padding:12px;background:#fbbc04;color:#333;border:none;border-radius:8px;cursor:pointer;font-size:15px;font-family:'Khmer OS','Khmer OS Muol','Arial',sans-serif;">🗑️ សម្អាតតែបេសកម្ម</button>
                <button onclick="closeModal('cleanModal')" style="padding:12px;background:#e8ecf1;color:#333;border:none;border-radius:8px;cursor:pointer;font-size:15px;font-family:'Khmer OS','Khmer OS Muol','Arial',sans-serif;">❌ បោះបង់</button>
            </div>
        </div>
    </div>

    <script>
    var currentDataVersion = {{ data_version }};
    var dataCheckInterval = null;
    var userLocation = null;
    var notificationInterval = null;
    var isFirstCheck = true;
    var deferredPrompt = null;
    var systemLocked = false;
    var autoUnlockTime = '06:00';
    var userLocked = {{ user_lock.is_locked|default(0) }};

    function updateLockUI(lockStatus) {
        var banner = document.getElementById('lockBanner');
        var lockText = document.getElementById('lockText');
        var lockTime = document.getElementById('lockTime');
        systemLocked = lockStatus.is_locked == 1;
        autoUnlockTime = lockStatus.auto_unlock_time || '06:00';

        if (systemLocked) {
            banner.className = 'lock-banner show';
            lockText.textContent = '🔒 ប្រព័ន្ធកំពុងបិទការចូលធ្វើការ';
            lockTime.textContent = '⏰ បើកដោយស្វ័យប្រវត្តិនៅម៉ោង ' + autoUnlockTime;
        } else {
            banner.className = 'lock-banner show unlocked';
            lockText.textContent = '🔓 ប្រព័ន្ធកំពុងបើក';
            lockTime.textContent = '✅ អនុញ្ញាតអោយចុចចូល/ចេញធ្វើការ';
        }
    }

    function getSystemLockStatus() {
        fetch('/get_system_lock_status')
        .then(function(res) { return res.json(); })
        .then(function(data) { updateLockUI(data); })
        .catch(function(err) { console.log('Lock status error:', err); });
    }

    function openModal(id) { document.getElementById(id).classList.add('show'); }
    function closeModal(id) { document.getElementById(id).classList.remove('show'); }
    function closeLocationModal() { closeModal('locationModal'); }
    function closeRequestsModal() { closeModal('requestsModal'); }
    function openLocationModal() { openModal('locationModal'); }
    function openRequestsModal() { openModal('requestsModal'); }
    function openCleanModal() { openModal('cleanModal'); }

    getSystemLockStatus();
    setInterval(getSystemLockStatus, 60000);
    </script>
</body>
</html>
'''

REGISTER_HTML = '''<!DOCTYPE html>
<html>
<head><title>ចុះឈ្មោះ</title><meta charset="UTF-8"></head>
<body><h2>ចុះឈ្មោះ</h2></body>
</html>'''

CHANGE_PASSWORD_HTML = '''<!DOCTYPE html>
<html>
<head><title>ប្តូរពាក្យសម្ងាត់</title><meta charset="UTF-8"></head>
<body><h2>ប្តូរពាក្យសម្ងាត់</h2></body>
</html>'''

USER_MANAGEMENT_HTML = '''<!DOCTYPE html>
<html>
<head><title>គ្រប់គ្រងអ្នកប្រើ</title><meta charset="UTF-8"></head>
<body><h2>គ្រប់គ្រងអ្នកប្រើ</h2></body>
</html>'''

REPORT_HTML = '''<!DOCTYPE html>
<html>
<head><title>របាយការណ៍</title><meta charset="UTF-8"></head>
<body><h2>របាយការណ៍</h2></body>
</html>'''

# ============================================================
# MAIN
# ============================================================

if __name__ == '__main__':
    init_db()
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port)
