from flask import Flask, request, redirect, url_for, session, render_template_string, jsonify, send_file, send_from_directory
import os
import math
import sqlite3
import datetime
import pytz
import uuid
from datetime import datetime, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

app = Flask(__name__)
app.secret_key = 'your-secret-key-here-change-it-to-something-secure-123456789'

import threading
import time

# ============================================================
# BACKGROUND AUTO-LOCK/UNLOCK THREAD
# ============================================================

def auto_lock_unlock_checker():
    """Background thread to automatically lock/unlock users based on deadlines"""
    while True:
        try:
            current_time = get_current_time_only()  # HH:MM:SS
            current_hhmm = current_time[:5]  # HH:MM

            # ===== Check System Lock (Auto-unlock) =====
            lock = get_system_lock_status()
            if lock.get('is_locked', 0) == 1:
                auto_unlock = lock.get('auto_unlock_time')
                if auto_unlock and current_hhmm >= auto_unlock:
                    print(f"🔄 Auto-unlocking system at {current_hhmm} (scheduled: {auto_unlock})")
                    update_system_lock(0, locked_by=None)
                    increment_data_version()
                    print("✅ System auto-unlocked successfully!")

            # ===== Auto-lock users based on attendance deadline =====
            conn = get_db_connection()
            active_settings = conn.execute('''
                SELECT user_id, check_in_deadline
                FROM attendance_settings
                WHERE is_active = 1
                AND check_in_deadline IS NOT NULL
                AND check_in_deadline != ''
            ''').fetchall()
            conn.close()

            for setting in active_settings:
                user_id = setting['user_id']
                deadline = setting['check_in_deadline']
                
                if deadline and current_hhmm >= deadline:
                    user_lock = get_user_lock_status(user_id)
                    if user_lock.get('is_locked', 0) != 1:
                        print(f"🔒 Auto-locking user {user_id} at {current_hhmm} (deadline: {deadline})")
                        system_lock = get_system_lock_status()
                        auto_unlock = system_lock.get('auto_unlock_time', '06:00')
                        update_user_lock(user_id, 1, auto_unlock_time=auto_unlock)
                        increment_data_version()
                        print(f"✅ User {user_id} auto-locked successfully!")

            # ===== Auto-unlock users based on auto_unlock_time =====
            conn = get_db_connection()
            locked_users = conn.execute('''
                SELECT user_id, auto_unlock_time
                FROM user_attendance_lock
                WHERE is_locked = 1
                AND auto_unlock_time IS NOT NULL
                AND auto_unlock_time != ''
            ''').fetchall()
            conn.close()

            for user in locked_users:
                auto_unlock = user['auto_unlock_time']
                if auto_unlock and current_hhmm >= auto_unlock:
                    user_id = user['user_id']
                    print(f"🔄 Auto-unlocking user {user_id} at {current_hhmm} (scheduled: {auto_unlock})")
                    update_user_lock(user_id, 0)
                    increment_data_version()
                    print(f"✅ User {user_id} auto-unlocked successfully!")

        except Exception as e:
            print(f"❌ Error in auto-lock/unlock checker: {e}")

        time.sleep(30)

def start_auto_lock_unlock_thread():
    thread = threading.Thread(target=auto_lock_unlock_checker, daemon=True)
    thread.start()
    print("✅ Auto-lock/unlock background thread started!")

start_auto_lock_unlock_thread()

# ===== UPLOAD CONFIGURATION =====
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024
app.config['ALLOWED_DISTANCE'] = 150

UPLOAD_FOLDER_LEAVES = os.path.join('static', 'uploads', 'leaves')
UPLOAD_FOLDER_MISSIONS = os.path.join('static', 'uploads', 'missions')
os.makedirs(UPLOAD_FOLDER_LEAVES, exist_ok=True)
os.makedirs(UPLOAD_FOLDER_MISSIONS, exist_ok=True)

DB_NAME = 'employees.db'

def get_db_connection():
    conn = sqlite3.connect(DB_NAME, timeout=20)
    conn.row_factory = sqlite3.Row
    return conn

def get_current_date():
    tz = pytz.timezone('Asia/Phnom_Penh')
    return datetime.now(tz).strftime('%Y-%m-%d')

def get_current_datetime():
    tz = pytz.timezone('Asia/Phnom_Penh')
    return datetime.now(tz)

def get_current_time():
    tz = pytz.timezone('Asia/Phnom_Penh')
    return datetime.now(tz).strftime('%Y-%m-%d %H:%M:%S')

def get_current_time_only():
    tz = pytz.timezone('Asia/Phnom_Penh')
    return datetime.now(tz).strftime('%H:%M:%S')

def table_exists(table_name):
    conn = get_db_connection()
    result = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
        (table_name,)
    ).fetchone()
    conn.close()
    return result is not None

def column_exists(table_name, column_name):
    conn = get_db_connection()
    try:
        conn.execute(f"SELECT {column_name} FROM {table_name} LIMIT 1")
        exists = True
    except sqlite3.OperationalError:
        exists = False
    conn.close()
    return exists

def migrate_database():
    conn = get_db_connection()

    if not table_exists('attendance_settings'):
        conn.execute('''
            CREATE TABLE IF NOT EXISTS attendance_settings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                check_in_deadline TEXT,
                is_active INTEGER DEFAULT 0,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES users(id),
                UNIQUE(user_id)
            )
        ''')
        conn.commit()
        print("✅ Created table attendance_settings")

    if table_exists('attendance_settings'):
        if not column_exists('attendance_settings', 'check_in_deadline'):
            try:
                conn.execute("ALTER TABLE attendance_settings ADD COLUMN check_in_deadline TEXT")
                print("✅ Added column check_in_deadline")
            except sqlite3.OperationalError as e:
                print(f"⚠️ Could not add check_in_deadline: {e}")

        if not column_exists('attendance_settings', 'is_active'):
            try:
                conn.execute("ALTER TABLE attendance_settings ADD COLUMN is_active INTEGER DEFAULT 0")
                print("✅ Added column is_active")
            except sqlite3.OperationalError as e:
                print(f"⚠️ Could not add is_active: {e}")

    if not table_exists('user_attendance_lock'):
        conn.execute('''
            CREATE TABLE IF NOT EXISTS user_attendance_lock (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                is_locked INTEGER DEFAULT 0,
                auto_unlock_time TEXT,
                locked_by INTEGER,
                lock_start_time TEXT,
                lock_end_time TEXT,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES users(id),
                UNIQUE(user_id)
            )
        ''')
        conn.commit()
        print("✅ Created table user_attendance_lock")

    if not table_exists('system_attendance_lock'):
        conn.execute('''
            CREATE TABLE IF NOT EXISTS system_attendance_lock (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                is_locked INTEGER DEFAULT 0,
                lock_start_time TEXT,
                lock_end_time TEXT,
                auto_unlock_time TEXT,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                locked_by INTEGER,
                FOREIGN KEY (locked_by) REFERENCES users(id)
            )
        ''')
        conn.commit()
        print("✅ Created table system_attendance_lock")

        conn.execute('''
            INSERT INTO system_attendance_lock (is_locked, auto_unlock_time)
            VALUES (0, '06:00')
        ''')
        conn.commit()
        print("✅ Inserted default system lock record")

    if not table_exists('attendance'):
        conn.execute('''
            CREATE TABLE IF NOT EXISTS attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                date TEXT NOT NULL,
                check_in TEXT,
                check_out TEXT,
                total_hours REAL DEFAULT 0,
                shift INTEGER DEFAULT 1,
                check_in_lat REAL,
                check_in_lng REAL,
                check_out_lat REAL,
                check_out_lng REAL,
                check_in_distance REAL,
                check_out_distance REAL,
                FOREIGN KEY (user_id) REFERENCES users(id)
            )
        ''')
        conn.commit()
        print("✅ Created table attendance")
        conn.close()
        return

    columns_to_add = [
        ('check_in_lat', 'REAL'),
        ('check_in_lng', 'REAL'),
        ('check_out_lat', 'REAL'),
        ('check_out_lng', 'REAL'),
        ('check_in_distance', 'REAL'),
        ('check_out_distance', 'REAL'),
        ('shift', 'INTEGER DEFAULT 1')
    ]

    for col_name, col_type in columns_to_add:
        if not column_exists('attendance', col_name):
            try:
                conn.execute(f"ALTER TABLE attendance ADD COLUMN {col_name} {col_type}")
                print(f"✅ Added column {col_name}")
            except sqlite3.OperationalError as e:
                print(f"⚠️ Could not add {col_name}: {e}")

    if table_exists('leaves') and not column_exists('leaves', 'attachment'):
        try:
            conn.execute("ALTER TABLE leaves ADD COLUMN attachment TEXT")
            print("✅ Added column attachment to leaves table")
        except sqlite3.OperationalError as e:
            print(f"⚠️ Could not add attachment column: {e}")

    if table_exists('leaves') and not column_exists('leaves', 'admin_id'):
        try:
            conn.execute("ALTER TABLE leaves ADD COLUMN admin_id INTEGER")
            print("✅ Added column admin_id to leaves table")
        except sqlite3.OperationalError as e:
            print(f"⚠️ Could not add admin_id column to leaves: {e}")

    if table_exists('missions') and not column_exists('missions', 'attachment'):
        try:
            conn.execute("ALTER TABLE missions ADD COLUMN attachment TEXT")
            print("✅ Added column attachment to missions table")
        except sqlite3.OperationalError as e:
            print(f"⚠️ Could not add attachment column: {e}")

    if table_exists('missions') and not column_exists('missions', 'admin_id'):
        try:
            conn.execute("ALTER TABLE missions ADD COLUMN admin_id INTEGER")
            print("✅ Added column admin_id to missions table")
        except sqlite3.OperationalError as e:
            print(f"⚠️ Could not add admin_id column to missions: {e}")

    conn.commit()
    conn.close()

def init_db():
    conn = get_db_connection()

    conn.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            full_name TEXT NOT NULL,
            email TEXT,
            phone TEXT,
            role TEXT DEFAULT 'user',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            check_in TEXT,
            check_out TEXT,
            total_hours REAL DEFAULT 0,
            shift INTEGER DEFAULT 1,
            check_in_lat REAL,
            check_in_lng REAL,
            check_out_lat REAL,
            check_out_lng REAL,
            check_in_distance REAL,
            check_out_distance REAL,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS leaves (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            start_date TEXT NOT NULL,
            end_date TEXT NOT NULL,
            days REAL NOT NULL,
            reason TEXT,
            attachment TEXT,
            status TEXT DEFAULT 'pending',
            admin_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id),
            FOREIGN KEY (admin_id) REFERENCES users(id)
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS missions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            start_date TEXT NOT NULL,
            end_date TEXT NOT NULL,
            days REAL NOT NULL,
            destination TEXT,
            purpose TEXT,
            attachment TEXT,
            status TEXT DEFAULT 'pending',
            admin_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id),
            FOREIGN KEY (admin_id) REFERENCES users(id)
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS company_location (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            lat REAL NOT NULL,
            lng REAL NOT NULL,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS data_version (
            version INTEGER DEFAULT 1
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS attendance_settings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            check_in_deadline TEXT,
            is_active INTEGER DEFAULT 0,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id),
            UNIQUE(user_id)
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS system_attendance_lock (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            is_locked INTEGER DEFAULT 0,
            lock_start_time TEXT,
            lock_end_time TEXT,
            auto_unlock_time TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            locked_by INTEGER,
            FOREIGN KEY (locked_by) REFERENCES users(id)
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS user_attendance_lock (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            is_locked INTEGER DEFAULT 0,
            auto_unlock_time TEXT,
            locked_by INTEGER,
            lock_start_time TEXT,
            lock_end_time TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id),
            UNIQUE(user_id)
        )
    ''')

    admin = conn.execute("SELECT * FROM users WHERE username = 'admin'").fetchone()
    if not admin:
        conn.execute(
            "INSERT INTO users (username, password, full_name, role) VALUES (?, ?, ?, ?)",
            ('admin', '1234', 'អ្នកគ្រប់គ្រង', 'admin')
        )
        print("✅ Created admin user: admin / 1234")

    test_user = conn.execute("SELECT * FROM users WHERE username = 'user1'").fetchone()
    if not test_user:
        conn.execute(
            "INSERT INTO users (username, password, full_name, role) VALUES (?, ?, ?, ?)",
            ('user1', '1234', 'បុគ្គលិកសាកល្បង', 'user')
        )
        print("✅ Created test user: user1 / 1234")

    version = conn.execute("SELECT version FROM data_version").fetchone()
    if not version:
        conn.execute("INSERT INTO data_version (version) VALUES (1)")

    lock = conn.execute("SELECT * FROM system_attendance_lock LIMIT 1").fetchone()
    if not lock:
        conn.execute('''
            INSERT INTO system_attendance_lock (is_locked, auto_unlock_time)
            VALUES (0, '06:00')
        ''')
        print("✅ Inserted default system lock record")

    conn.commit()
    conn.close()

    migrate_database()
    print(f"✅ Database '{DB_NAME}' initialized successfully!")

# ============================================================
# SYSTEM ATTENDANCE LOCK FUNCTIONS
# ============================================================

def get_system_lock_status():
    try:
        conn = get_db_connection()
        result = conn.execute("SELECT * FROM system_attendance_lock ORDER BY id DESC LIMIT 1").fetchone()
        conn.close()
        return dict(result) if result else {'is_locked': 0, 'auto_unlock_time': '06:00'}
    except Exception as e:
        print(f"Error getting lock status: {e}")
        return {'is_locked': 0, 'auto_unlock_time': '06:00'}

def update_system_lock(is_locked, lock_start_time=None, lock_end_time=None, auto_unlock_time=None, locked_by=None):
    try:
        conn = get_db_connection()
        current = conn.execute("SELECT * FROM system_attendance_lock ORDER BY id DESC LIMIT 1").fetchone()

        if current:
            conn.execute('''
                UPDATE system_attendance_lock
                SET is_locked = ?,
                    lock_start_time = COALESCE(?, lock_start_time),
                    lock_end_time = COALESCE(?, lock_end_time),
                    auto_unlock_time = COALESCE(?, auto_unlock_time),
                    locked_by = COALESCE(?, locked_by),
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (is_locked, lock_start_time, lock_end_time, auto_unlock_time, locked_by, current['id']))
        else:
            conn.execute('''
                INSERT INTO system_attendance_lock
                (is_locked, lock_start_time, lock_end_time, auto_unlock_time, locked_by)
                VALUES (?, ?, ?, ?, ?)
            ''', (is_locked, lock_start_time, lock_end_time, auto_unlock_time, locked_by))

        conn.commit()
        conn.close()
        increment_data_version()
        return True
    except Exception as e:
        print(f"Error updating lock: {e}")
        return False

def check_system_lock_for_user(user_id):
    lock = get_system_lock_status()

    if lock.get('is_locked', 0) != 1:
        return True, None

    auto_unlock = lock.get('auto_unlock_time')
    if auto_unlock:
        current_time = get_current_time_only()
        current_hhmm = current_time[:5]
        if current_hhmm >= auto_unlock:
            print(f"🔄 Auto-unlocking system during check (current: {current_hhmm}, scheduled: {auto_unlock})")
            update_system_lock(0, locked_by=None)
            increment_data_version()
            return True, None

    return False, "⛔ ប្រព័ន្ធកំពុងបិទការចូលធ្វើការ! សូមរង់ចាំរហូតដល់ម៉ោងបើកដោយស្វ័យប្រវត្តិ ឬទាក់ទង Admin!"

def toggle_system_lock(lock_state, auto_unlock_time=None, locked_by=None):
    lock = get_system_lock_status()
    current_time = get_current_time()

    if lock_state == 1:
        return update_system_lock(
            is_locked=1,
            lock_start_time=current_time,
            auto_unlock_time=auto_unlock_time or lock.get('auto_unlock_time', '06:00'),
            locked_by=locked_by
        )
    else:
        return update_system_lock(
            is_locked=0,
            lock_end_time=current_time,
            locked_by=locked_by
        )

# ============================================================
# USER ATTENDANCE LOCK FUNCTIONS
# ============================================================

def get_user_lock_status(user_id):
    try:
        conn = get_db_connection()
        result = conn.execute(
            "SELECT * FROM user_attendance_lock WHERE user_id = ?",
            (user_id,)
        ).fetchone()
        conn.close()
        return dict(result) if result else {'is_locked': 0, 'auto_unlock_time': None}
    except Exception as e:
        print(f"Error getting user lock status: {e}")
        return {'is_locked': 0, 'auto_unlock_time': None}

def update_user_lock(user_id, is_locked, auto_unlock_time=None, locked_by=None):
    try:
        conn = get_db_connection()
        current_time = get_current_time()

        existing = conn.execute(
            "SELECT id FROM user_attendance_lock WHERE user_id = ?",
            (user_id,)
        ).fetchone()

        if existing:
            if is_locked == 1:
                conn.execute('''
                    UPDATE user_attendance_lock
                    SET is_locked = ?,
                        auto_unlock_time = ?,
                        locked_by = ?,
                        lock_start_time = ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE user_id = ?
                ''', (is_locked, auto_unlock_time, locked_by, current_time, user_id))
            else:
                conn.execute('''
                    UPDATE user_attendance_lock
                    SET is_locked = ?,
                        lock_end_time = ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE user_id = ?
                ''', (is_locked, current_time, user_id))
        else:
            conn.execute('''
                INSERT INTO user_attendance_lock
                (user_id, is_locked, auto_unlock_time, locked_by, lock_start_time)
                VALUES (?, ?, ?, ?, ?)
            ''', (user_id, is_locked, auto_unlock_time, locked_by, current_time))

        conn.commit()
        conn.close()
        increment_data_version()
        return True
    except Exception as e:
        print(f"Error updating user lock: {e}")
        return False

def check_user_lock(user_id):
    lock = get_user_lock_status(user_id)

    if lock.get('is_locked', 0) != 1:
        return True, None

    auto_unlock = lock.get('auto_unlock_time')
    if auto_unlock:
        current_time = get_current_time_only()
        current_hhmm = current_time[:5]
        if current_hhmm >= auto_unlock:
            print(f"🔄 Auto-unlocking user {user_id} during check (current: {current_hhmm}, scheduled: {auto_unlock})")
            update_user_lock(user_id, 0)
            increment_data_version()
            return True, None

    return False, "⛔ អ្នកត្រូវបានបិទការចូលធ្វើការដោយ Admin! សូមទាក់ទង Admin!"

def get_all_user_lock_status():
    try:
        conn = get_db_connection()
        results = conn.execute('''
            SELECT u.id, u.username, u.full_name,
                   l.is_locked, l.auto_unlock_time, l.lock_start_time, l.updated_at
            FROM users u
            LEFT JOIN user_attendance_lock l ON u.id = l.user_id
            WHERE u.role != 'admin'
            ORDER BY u.full_name
        ''').fetchall()
        conn.close()
        return [dict(row) for row in results]
    except Exception as e:
        print(f"Error getting all user locks: {e}")
        return []

def toggle_user_lock(user_id, lock_state, auto_unlock_time=None, locked_by=None):
    if lock_state == 1:
        return update_user_lock(user_id, 1, auto_unlock_time, locked_by)
    else:
        return update_user_lock(user_id, 0, locked_by=locked_by)

# ============================================================
# DATA VERSION FUNCTIONS
# ============================================================

def get_data_version():
    try:
        conn = get_db_connection()
        result = conn.execute("SELECT version FROM data_version").fetchone()
        conn.close()
        return result['version'] if result else 1
    except sqlite3.OperationalError:
        init_db()
        return 1

def increment_data_version():
    try:
        conn = get_db_connection()
        conn.execute("UPDATE data_version SET version = version + 1")
        conn.commit()
        conn.close()
    except sqlite3.OperationalError:
        init_db()
        conn = get_db_connection()
        conn.execute("UPDATE data_version SET version = version + 1")
        conn.commit()
        conn.close()

# ============================================================
# USER FUNCTIONS
# ============================================================

def get_user_by_id(user_id):
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    conn.close()
    return user

def get_user_by_username(username):
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
    conn.close()
    return user

def create_user(username, password, full_name, email=None, phone=None, role='user'):
    conn = get_db_connection()
    try:
        conn.execute(
            "INSERT INTO users (username, password, full_name, email, phone, role) VALUES (?, ?, ?, ?, ?, ?)",
            (username, password, full_name, email, phone, role)
        )
        conn.commit()
        conn.close()
        increment_data_version()
        return True
    except sqlite3.IntegrityError:
        conn.close()
        return False

def update_user(user_id, username, full_name, email, phone, role):
    conn = get_db_connection()
    try:
        conn.execute(
            "UPDATE users SET username = ?, full_name = ?, email = ?, phone = ?, role = ? WHERE id = ?",
            (username, full_name, email, phone, role, user_id)
        )
        conn.commit()
        conn.close()
        increment_data_version()
        return True
    except sqlite3.IntegrityError:
        conn.close()
        return False

def update_password(user_id, new_password):
    conn = get_db_connection()
    conn.execute("UPDATE users SET password = ? WHERE id = ?", (new_password, user_id))
    conn.commit()
    conn.close()
    increment_data_version()
    return True

def verify_password(user_id, password):
    conn = get_db_connection()
    user = conn.execute("SELECT password FROM users WHERE id = ?", (user_id,)).fetchone()
    conn.close()
    return user and user['password'] == password

def delete_user(user_id):
    conn = get_db_connection()
    try:
        conn.execute("DELETE FROM attendance WHERE user_id = ?", (user_id,))
        conn.execute("DELETE FROM leaves WHERE user_id = ?", (user_id,))
        conn.execute("DELETE FROM missions WHERE user_id = ?", (user_id,))
        conn.execute("DELETE FROM attendance_settings WHERE user_id = ?", (user_id,))
        conn.execute("DELETE FROM user_attendance_lock WHERE user_id = ?", (user_id,))
        conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
        conn.commit()
        conn.close()
        increment_data_version()
        return True
    except Exception:
        conn.close()
        return False

def get_all_users():
    conn = get_db_connection()
    users = conn.execute("SELECT * FROM users ORDER BY id DESC").fetchall()
    conn.close()
    return users

# ============================================================
# ATTENDANCE SETTINGS FUNCTIONS
# ============================================================

def get_attendance_setting(user_id):
    conn = get_db_connection()
    setting = conn.execute(
        "SELECT * FROM attendance_settings WHERE user_id = ?",
        (user_id,)
    ).fetchone()
    conn.close()
    return dict(setting) if setting else None

def get_all_attendance_settings():
    conn = get_db_connection()
    settings = conn.execute('''
        SELECT s.*, u.username, u.full_name
        FROM attendance_settings s
        JOIN users u ON s.user_id = u.id
        ORDER BY u.full_name
    ''').fetchall()
    conn.close()
    return [dict(row) for row in settings]

def save_attendance_setting(user_id, check_in_deadline, is_active):
    conn = get_db_connection()
    existing = conn.execute(
        "SELECT id FROM attendance_settings WHERE user_id = ?",
        (user_id,)
    ).fetchone()

    if existing:
        conn.execute(
            "UPDATE attendance_settings SET check_in_deadline = ?, is_active = ?, updated_at = CURRENT_TIMESTAMP WHERE user_id = ?",
            (check_in_deadline, is_active, user_id)
        )
    else:
        conn.execute(
            "INSERT INTO attendance_settings (user_id, check_in_deadline, is_active) VALUES (?, ?, ?)",
            (user_id, check_in_deadline, is_active)
        )
    conn.commit()
    conn.close()
    increment_data_version()
    return True

def check_attendance_deadline(user_id):
    setting = get_attendance_setting(user_id)
    if not setting:
        return True, None

    if setting.get('is_active') != 1:
        return True, None

    deadline = setting.get('check_in_deadline')
    if not deadline:
        return True, None

    current_time = get_current_time_only()
    current_hhmm = current_time[:5]

    if current_hhmm > deadline:
        return False, deadline
    return True, deadline

# ============================================================
# ATTENDANCE FUNCTIONS
# ============================================================

def get_company_location():
    conn = get_db_connection()
    location = conn.execute("SELECT lat, lng FROM company_location ORDER BY id DESC LIMIT 1").fetchone()
    conn.close()
    return dict(location) if location else None

def save_company_location(lat, lng):
    conn = get_db_connection()
    conn.execute("INSERT INTO company_location (lat, lng) VALUES (?, ?)", (lat, lng))
    conn.commit()
    conn.close()
    increment_data_version()
    return True

def check_in(user_id, lat, lng, distance, shift):
    date = get_current_date()
    check_in_time = get_current_time()
    conn = get_db_connection()

    existing = conn.execute(
        """SELECT id FROM attendance
           WHERE user_id = ?
           AND check_out IS NULL""",
        (user_id,)
    ).fetchone()

    if existing:
        conn.close()
        return False, "អ្នកបានចូលធ្វើការរួចហើយ! សូមចុច 'ចេញពីធ្វើការ' មុនពេលចូលម្តងទៀត!"

    can_checkin, deadline = check_attendance_deadline(user_id)
    if not can_checkin:
        return False, f"⛔ អ្នកលើសម៉ោងដែល Admin បានកំណត់ (ម៉ោងកំណត់: {deadline})! សូមទាក់ទងទៅអ្នកគ្រប់គ្រង!"

    conn.execute(
        """INSERT INTO attendance
           (user_id, date, check_in, shift, check_in_lat, check_in_lng, check_in_distance)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (user_id, date, check_in_time, shift, lat, lng, distance)
    )
    conn.commit()
    conn.close()
    increment_data_version()
    return True, "ចូលធ្វើការជោគជ័យ!"

def check_out(user_id, lat, lng, distance):
    check_out_time = get_current_datetime()
    conn = get_db_connection()

    record = conn.execute(
        """SELECT id, check_in, shift, date
           FROM attendance
           WHERE user_id = ?
           AND check_out IS NULL
           ORDER BY id DESC LIMIT 1""",
        (user_id,)
    ).fetchone()

    if not record:
        conn.close()
        return False, "មិនមានការចូលធ្វើការដែលមិនទាន់ចេញ!"

    check_in_time = datetime.strptime(record['check_in'], '%Y-%m-%d %H:%M:%S')
    tz = pytz.timezone('Asia/Phnom_Penh')
    check_in_time = tz.localize(check_in_time)

    if check_out_time <= check_in_time:
        check_out_time = check_out_time + timedelta(days=1)

    diff = check_out_time - check_in_time
    total_hours = diff.total_seconds() / 3600

    conn.execute(
        """UPDATE attendance
           SET check_out = ?, total_hours = ?, check_out_lat = ?, check_out_lng = ?, check_out_distance = ?
           WHERE id = ?""",
        (check_out_time.strftime('%Y-%m-%d %H:%M:%S'), total_hours, lat, lng, distance, record['id'])
    )
    conn.commit()
    conn.close()
    increment_data_version()

    shift_names = {1: 'វគ្គ 1 (ព្រឹក)', 2: 'វគ្គ 2 (រសៀល)', 3: 'វគ្គ 3 (យប់)'}
    shift = record['shift'] if 'shift' in record.keys() else 1

    hours = int(total_hours)
    minutes = int((total_hours - hours) * 60)
    return True, f"ចេញធ្វើការជោគជ័យ! ({shift_names.get(shift, '')}) ម៉ោងសរុប: {hours:02d}:{minutes:02d}"

def get_attendance_stats():
    today = get_current_date()
    conn = get_db_connection()

    total_users = conn.execute("SELECT COUNT(*) as count FROM users WHERE role != 'admin'").fetchone()['count']

    present_today = conn.execute(
        "SELECT COUNT(DISTINCT user_id) as count FROM attendance WHERE check_out IS NULL"
    ).fetchone()['count']

    leave_today = conn.execute(
        "SELECT COUNT(DISTINCT user_id) as count FROM leaves WHERE status = 'approved' AND ? BETWEEN start_date AND end_date",
        (today,)
    ).fetchone()['count']

    mission_today = conn.execute(
        "SELECT COUNT(DISTINCT user_id) as count FROM missions WHERE status = 'approved' AND ? BETWEEN start_date AND end_date",
        (today,)
    ).fetchone()['count']

    conn.close()

    return {
        'total_users': total_users,
        'present_today': present_today,
        'leave_today': leave_today,
        'mission_today': mission_today
    }

def get_checkin_status(user_id):
    conn = get_db_connection()
    record = conn.execute(
        """SELECT id, check_in, COALESCE(shift, 1) as shift
           FROM attendance
           WHERE user_id = ?
           AND check_out IS NULL
           ORDER BY id DESC LIMIT 1""",
        (user_id,)
    ).fetchone()
    conn.close()

    if record:
        return {
            'has_checkin': True,
            'check_in_time': record['check_in'],
            'shift': record['shift']
        }
    return {'has_checkin': False, 'check_in_time': None, 'shift': None}

# ============================================================
# DASHBOARD REPORT FUNCTIONS
# ============================================================

def get_work_history_report(start_date=None, end_date=None, limit=200):
    if not start_date:
        today = datetime.now()
        start_date = today.replace(day=1).strftime('%Y-%m-%d')
    if not end_date:
        end_date = datetime.now().strftime('%Y-%m-%d')

    conn = get_db_connection()

    attendance_records = conn.execute('''
        SELECT
            a.id,
            a.user_id,
            u.full_name,
            a.date,
            CASE
                WHEN a.shift = 1 THEN 'វគ្គ 1'
                WHEN a.shift = 2 THEN 'វគ្គ 2'
                WHEN a.shift = 3 THEN 'វគ្គ 3'
                ELSE 'វគ្គ 1'
            END as shift,
            a.check_in,
            a.check_out,
            a.total_hours,
            'attendance' as type,
            NULL as days,
            NULL as reason,
            NULL as destination,
            NULL as start_date,
            NULL as end_date,
            CASE
                WHEN a.check_in IS NOT NULL AND a.check_out IS NOT NULL THEN 'បានបិទ'
                WHEN a.check_in IS NOT NULL AND a.check_out IS NULL THEN 'កំពុងធ្វើការ'
                ELSE 'មិនទាន់ចូល'
            END as status,
            a.date as sort_date
        FROM attendance a
        JOIN users u ON a.user_id = u.id
        WHERE a.date BETWEEN ? AND ?
            AND a.check_in IS NOT NULL
    ''', (start_date, end_date)).fetchall()

    leave_records = conn.execute('''
        SELECT
            l.id,
            l.user_id,
            u.full_name,
            l.start_date as date,
            'ច្បាប់' as shift,
            NULL as check_in,
            NULL as check_out,
            NULL as total_hours,
            'leave' as type,
            l.days as days,
            l.reason as reason,
            NULL as destination,
            l.start_date,
            l.end_date,
            l.status,
            l.start_date as sort_date
        FROM leaves l
        JOIN users u ON l.user_id = u.id
        WHERE l.status = 'approved'
            AND l.start_date BETWEEN ? AND ?
    ''', (start_date, end_date)).fetchall()

    mission_records = conn.execute('''
        SELECT
            m.id,
            m.user_id,
            u.full_name,
            m.start_date as date,
            'បេសកម្ម' as shift,
            NULL as check_in,
            NULL as check_out,
            NULL as total_hours,
            'mission' as type,
            m.days as days,
            NULL as reason,
            m.destination as destination,
            m.start_date,
            m.end_date,
            m.status,
            m.start_date as sort_date
        FROM missions m
        JOIN users u ON m.user_id = u.id
        WHERE m.status = 'approved'
            AND m.start_date BETWEEN ? AND ?
    ''', (start_date, end_date)).fetchall()

    conn.close()

    result = []

    for row in attendance_records:
        item = dict(row)
        if item.get('total_hours'):
            hours = int(item['total_hours'])
            minutes = int((item['total_hours'] - hours) * 60)
            if item['total_hours'] < 0:
                hours = abs(hours)
                minutes = abs(minutes)
                item['total_hours_formatted'] = f"-{hours:02d}:{minutes:02d}"
            else:
                item['total_hours_formatted'] = f"{hours:02d}:{minutes:02d}"
            item['display_value'] = item['total_hours_formatted']
        else:
            item['total_hours_formatted'] = ''
            item['display_value'] = ''
        if item.get('check_in'):
            item['check_in_time'] = item['check_in'][11:16] if ' ' in item['check_in'] else item['check_in'][:5]
        else:
            item['check_in_time'] = ''
        if item.get('check_out'):
            item['check_out_time'] = item['check_out'][11:16] if ' ' in item['check_out'] else item['check_out'][:5]
        else:
            item['check_out_time'] = ''
        item['display_info'] = ''
        result.append(item)

    for row in leave_records:
        item = dict(row)
        item['shift'] = 'ច្បាប់'
        item['check_in_time'] = ''
        item['check_out_time'] = ''
        item['total_hours_formatted'] = ''
        if item.get('days'):
            item['display_value'] = f"{item['days']} ថ្ងៃ"
        else:
            item['display_value'] = ''
        item['display_info'] = f"មូលហេតុ: {item.get('reason', 'មិនបានបញ្ជាក់')}"
        item['status'] = 'បានអនុម័ត'
        result.append(item)

    for row in mission_records:
        item = dict(row)
        item['shift'] = 'បេសកម្ម'
        item['check_in_time'] = ''
        item['check_out_time'] = ''
        item['total_hours_formatted'] = ''
        if item.get('days'):
            item['display_value'] = f"{item['days']} ថ្ងៃ"
        else:
            item['display_value'] = ''
        item['display_info'] = f"ទីតាំង: {item.get('destination', 'មិនបានបញ្ជាក់')}"
        item['status'] = 'បានអនុម័ត'
        result.append(item)

    result.sort(key=lambda x: x['sort_date'] if x['sort_date'] else '', reverse=True)
    return result[:limit]

def get_monthly_summary_report(start_date, end_date):
    conn = get_db_connection()

    attendance_summary = conn.execute('''
        SELECT
            u.id,
            u.full_name,
            COUNT(DISTINCT a.date) as days_worked,
            COALESCE(SUM(a.total_hours), 0) as total_hours,
            COUNT(DISTINCT CASE WHEN a.shift = 3 THEN a.date END) as night_shifts,
            COALESCE(SUM(CASE WHEN a.shift = 3 THEN a.total_hours ELSE 0 END), 0) as night_hours
        FROM users u
        LEFT JOIN attendance a ON u.id = a.user_id
            AND a.date BETWEEN ? AND ?
            AND a.check_in IS NOT NULL
            AND a.check_out IS NOT NULL
        WHERE u.role != 'admin'
        GROUP BY u.id
        ORDER BY u.full_name
    ''', (start_date, end_date)).fetchall()

    leave_summary = conn.execute('''
        SELECT
            user_id,
            SUM(days) as total_leave_days
        FROM leaves
        WHERE status = 'approved'
            AND start_date BETWEEN ? AND ?
        GROUP BY user_id
    ''', (start_date, end_date)).fetchall()

    mission_summary = conn.execute('''
        SELECT
            user_id,
            SUM(days) as total_mission_days
        FROM missions
        WHERE status = 'approved'
            AND start_date BETWEEN ? AND ?
        GROUP BY user_id
    ''', (start_date, end_date)).fetchall()

    conn.close()

    leave_dict = {row['user_id']: row['total_leave_days'] for row in leave_summary}
    mission_dict = {row['user_id']: row['total_mission_days'] for row in mission_summary}

    result = []
    for emp in attendance_summary:
        item = dict(emp)
        item['total_leave_days'] = leave_dict.get(emp['id'], 0)
        item['total_mission_days'] = mission_dict.get(emp['id'], 0)

        total_hours = item['total_hours']
        if total_hours < 0:
            total_hours = abs(total_hours)
            sign = "-"
        else:
            sign = ""
        hours = int(total_hours)
        minutes = int((total_hours - hours) * 60)
        item['total_hours_formatted'] = f"{sign}{hours:02d}:{minutes:02d}"
        item['days_worked_display'] = f"{item['days_worked']} ថ្ងៃ"
        result.append(item)

    return result

# ============================================================
# ATTENDANCE FUNCTIONS (continued)
# ============================================================

def get_all_attendance(limit=100):
    conn = get_db_connection()

    attendances = conn.execute('''
        SELECT
            a.id,
            a.user_id,
            u.full_name,
            u.username,
            a.date,
            a.shift,
            a.check_in,
            a.check_out,
            a.total_hours,
            'attendance' as type,
            NULL as reason,
            NULL as destination,
            NULL as start_date,
            NULL as end_date,
            NULL as days,
            a.date as sort_date
        FROM attendance a
        JOIN users u ON a.user_id = u.id
        ORDER BY a.date DESC, a.id DESC
        LIMIT ?
    ''', (limit,)).fetchall()

    leaves = conn.execute('''
        SELECT
            l.id,
            l.user_id,
            u.full_name,
            u.username,
            l.start_date as date,
            NULL as shift,
            NULL as check_in,
            NULL as check_out,
            l.days as total_hours,
            'leave' as type,
            l.reason,
            NULL as destination,
            l.start_date,
            l.end_date,
            l.days,
            l.start_date as sort_date
        FROM leaves l
        JOIN users u ON l.user_id = u.id
        WHERE l.status = 'approved'
        ORDER BY l.start_date DESC, l.id DESC
        LIMIT ?
    ''', (limit,)).fetchall()

    missions = conn.execute('''
        SELECT
            m.id,
            m.user_id,
            u.full_name,
            u.username,
            m.start_date as date,
            NULL as shift,
            NULL as check_in,
            NULL as check_out,
            m.days as total_hours,
            'mission' as type,
            NULL as reason,
            m.destination,
            m.start_date,
            m.end_date,
            m.days,
            m.start_date as sort_date
        FROM missions m
        JOIN users u ON m.user_id = u.id
        WHERE m.status = 'approved'
        ORDER BY m.start_date DESC, m.id DESC
        LIMIT ?
    ''', (limit,)).fetchall()

    conn.close()

    all_data = []

    for row in attendances:
        all_data.append(dict(row))

    for row in leaves:
        all_data.append(dict(row))

    for row in missions:
        all_data.append(dict(row))

    all_data.sort(key=lambda x: x['sort_date'] if x['sort_date'] else '', reverse=True)

    return all_data[:limit]

def get_attendance_report(start_date, end_date):
    conn = get_db_connection()

    daily_report = conn.execute('''
        SELECT
            a.id,
            a.user_id,
            u.full_name,
            u.username,
            a.date,
            COALESCE(a.shift, 1) as shift,
            a.check_in,
            a.check_out,
            a.total_hours,
            'attendance' as type,
            NULL as reason,
            NULL as destination,
            NULL as start_date,
            NULL as end_date,
            NULL as days,
            CASE
                WHEN a.shift = 1 THEN 'វគ្គ 1'
                WHEN a.shift = 2 THEN 'វគ្គ 2'
                WHEN a.shift = 3 THEN 'វគ្គ 3'
                ELSE 'វគ្គ 1'
            END as shift_name,
            CASE
                WHEN a.check_in IS NOT NULL AND a.check_out IS NOT NULL THEN 'បានបិទ'
                WHEN a.check_in IS NOT NULL AND a.check_out IS NULL THEN 'កំពុងធ្វើការ'
                ELSE 'មិនទាន់ចូល'
            END as status
        FROM attendance a
        JOIN users u ON a.user_id = u.id
        WHERE a.date BETWEEN ? AND ?
        ORDER BY a.date DESC, a.user_id, a.shift
    ''', (start_date, end_date)).fetchall()

    leave_report = conn.execute('''
        SELECT
            l.id,
            l.user_id,
            u.full_name,
            u.username,
            l.start_date as date,
            NULL as shift,
            NULL as check_in,
            NULL as check_out,
            l.days as total_hours,
            'leave' as type,
            l.reason,
            NULL as destination,
            l.start_date,
            l.end_date,
            l.days,
            NULL as shift_name,
            l.status
        FROM leaves l
        JOIN users u ON l.user_id = u.id
        WHERE l.status = 'approved'
            AND l.start_date >= ? AND l.end_date <= ?
        ORDER BY l.start_date DESC
    ''', (start_date, end_date)).fetchall()

    mission_report = conn.execute('''
        SELECT
            m.id,
            m.user_id,
            u.full_name,
            u.username,
            m.start_date as date,
            NULL as shift,
            NULL as check_in,
            NULL as check_out,
            m.days as total_hours,
            'mission' as type,
            NULL as reason,
            m.destination,
            m.start_date,
            m.end_date,
            m.days,
            NULL as shift_name,
            m.status
        FROM missions m
        JOIN users u ON m.user_id = u.id
        WHERE m.status = 'approved'
            AND m.start_date >= ? AND m.end_date <= ?
        ORDER BY m.start_date DESC
    ''', (start_date, end_date)).fetchall()

    all_reports = []

    for row in daily_report:
        all_reports.append(dict(row))

    for row in leave_report:
        all_reports.append(dict(row))

    for row in mission_report:
        all_reports.append(dict(row))

    all_reports.sort(key=lambda x: x['date'] if x['date'] else '', reverse=True)

    monthly_summary = conn.execute('''
        SELECT
            u.id,
            u.username,
            u.full_name,
            u.role,
            COUNT(DISTINCT a.date) as days_worked,
            COALESCE(SUM(a.total_hours), 0) as total_hours,
            COUNT(DISTINCT CASE WHEN a.shift = 3 THEN a.date END) as night_shifts,
            COALESCE(SUM(CASE WHEN a.shift = 3 THEN a.total_hours ELSE 0 END), 0) as night_hours
        FROM users u
        LEFT JOIN attendance a ON u.id = a.user_id
            AND a.date BETWEEN ? AND ?
            AND a.check_in IS NOT NULL
            AND a.check_out IS NOT NULL
        WHERE u.role != 'admin'
        GROUP BY u.id
        ORDER BY total_hours DESC
    ''', (start_date, end_date)).fetchall()

    leave_days = conn.execute('''
        SELECT
            user_id,
            SUM(days) as total_leave_days
        FROM leaves
        WHERE status = 'approved'
            AND start_date >= ? AND end_date <= ?
        GROUP BY user_id
    ''', (start_date, end_date)).fetchall()

    mission_days = conn.execute('''
        SELECT
            user_id,
            SUM(days) as total_mission_days
        FROM missions
        WHERE status = 'approved'
            AND start_date >= ? AND end_date <= ?
        GROUP BY user_id
    ''', (start_date, end_date)).fetchall()

    conn.close()

    summary = []
    for emp in monthly_summary:
        emp_dict = dict(emp)
        emp_dict['total_leave_days'] = 0
        emp_dict['total_mission_days'] = 0

        for leave in leave_days:
            if leave['user_id'] == emp['id']:
                emp_dict['total_leave_days'] = leave['total_leave_days']
                break

        for mission in mission_days:
            if mission['user_id'] == emp['id']:
                emp_dict['total_mission_days'] = mission['total_mission_days']
                break

        summary.append(emp_dict)

    return {
        'daily': all_reports,
        'summary': summary
    }

def get_monthly_summary(year, month):
    start_date = f"{year}-{month:02d}-01"
    if month == 12:
        end_date = f"{year}-12-31"
    else:
        next_month = month + 1
        end_dt = datetime(year, next_month, 1) - timedelta(days=1)
        end_date = end_dt.strftime('%Y-%m-%d')
    return get_attendance_report(start_date, end_date)

# ============================================================
# ATTENDANCE MANAGEMENT FUNCTIONS
# ============================================================

def get_attendance_by_id(attendance_id):
    try:
        conn = get_db_connection()
        data = conn.execute('''
            SELECT a.*, u.full_name, u.username
            FROM attendance a
            JOIN users u ON u.id = a.user_id
            WHERE a.id = ?
        ''', (attendance_id,)).fetchone()
        conn.close()
        return data
    except Exception as e:
        print(f"Error in get_attendance_by_id: {e}")
        return None

def update_attendance(attendance_id, check_in=None, check_out=None, date=None, shift=None):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        updates = []
        params = []

        if check_in is not None:
            updates.append("check_in = ?")
            params.append(check_in)
        if check_out is not None:
            updates.append("check_out = ?")
            params.append(check_out)
        if date is not None:
            updates.append("date = ?")
            params.append(date)
        if shift is not None:
            updates.append("shift = ?")
            params.append(shift)

        if check_in is not None and check_out is not None:
            try:
                from datetime import datetime
                check_in_dt = datetime.strptime(check_in, '%Y-%m-%d %H:%M:%S')
                check_out_dt = datetime.strptime(check_out, '%Y-%m-%d %H:%M:%S')
                total_hours = (check_out_dt - check_in_dt).total_seconds() / 3600
                updates.append("total_hours = ?")
                params.append(total_hours)
            except:
                pass

        if not updates:
            conn.close()
            return False

        params.append(attendance_id)
        sql = f"UPDATE attendance SET {', '.join(updates)} WHERE id = ?"
        cursor.execute(sql, params)
        conn.commit()
        conn.close()
        increment_data_version()
        return True
    except Exception as e:
        print(f"Error in update_attendance: {e}")
        return False

def delete_attendance(attendance_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM attendance WHERE id = ?", (attendance_id,))
        conn.commit()
        conn.close()
        increment_data_version()
        return True
    except Exception as e:
        print(f"Error in delete_attendance: {e}")
        return False

# ============================================================
# LEAVE FUNCTIONS
# ============================================================

def create_leave(user_id, start_date, end_date, days, reason='', attachment=None):
    conn = get_db_connection()
    conn.execute(
        "INSERT INTO leaves (user_id, start_date, end_date, days, reason, attachment, status) VALUES (?, ?, ?, ?, ?, ?, 'pending')",
        (user_id, start_date, end_date, days, reason, attachment)
    )
    conn.commit()
    conn.close()
    increment_data_version()
    return True

def get_pending_leaves():
    conn = get_db_connection()
    leaves = conn.execute('''
        SELECT l.*, u.username, u.full_name
        FROM leaves l
        JOIN users u ON l.user_id = u.id
        WHERE l.status = 'pending'
        ORDER BY l.created_at DESC
    ''').fetchall()
    conn.close()
    return [dict(row) for row in leaves]

def approve_leave(leave_id, admin_id):
    conn = get_db_connection()
    conn.execute(
        "UPDATE leaves SET status = 'approved', admin_id = ? WHERE id = ?",
        (admin_id, leave_id)
    )
    conn.commit()
    conn.close()
    increment_data_version()
    return True

def reject_leave(leave_id):
    conn = get_db_connection()
    conn.execute(
        "UPDATE leaves SET status = 'rejected' WHERE id = ?",
        (leave_id,)
    )
    conn.commit()
    conn.close()
    increment_data_version()
    return True

# ============================================================
# MISSION FUNCTIONS
# ============================================================

def create_mission(user_id, start_date, end_date, days, destination='', purpose='', attachment=None):
    conn = get_db_connection()
    conn.execute(
        "INSERT INTO missions (user_id, start_date, end_date, days, destination, purpose, attachment, status) VALUES (?, ?, ?, ?, ?, ?, ?, 'pending')",
        (user_id, start_date, end_date, days, destination, purpose, attachment)
    )
    conn.commit()
    conn.close()
    increment_data_version()
    return True

def get_pending_missions():
    conn = get_db_connection()
    missions = conn.execute('''
        SELECT m.*, u.username, u.full_name
        FROM missions m
        JOIN users u ON m.user_id = u.id
        WHERE m.status = 'pending'
        ORDER BY m.created_at DESC
    ''').fetchall()
    conn.close()
    return [dict(row) for row in missions]

def approve_mission(mission_id, admin_id):
    conn = get_db_connection()
    conn.execute(
        "UPDATE missions SET status = 'approved', admin_id = ? WHERE id = ?",
        (admin_id, mission_id)
    )
    conn.commit()
    conn.close()
    increment_data_version()
    return True

def reject_mission(mission_id):
    conn = get_db_connection()
    conn.execute(
        "UPDATE missions SET status = 'rejected' WHERE id = ?",
        (mission_id,)
    )
    conn.commit()
    conn.close()
    increment_data_version()
    return True

# ============================================================
# CLEAN DATA FUNCTIONS
# ============================================================

def clean_all_data():
    conn = get_db_connection()
    conn.execute("DELETE FROM attendance")
    conn.execute("DELETE FROM leaves")
    conn.execute("DELETE FROM missions")
    conn.commit()
    conn.close()
    increment_data_version()
    return True

def clean_attendance_only():
    conn = get_db_connection()
    conn.execute("DELETE FROM attendance")
    conn.commit()
    conn.close()
    increment_data_version()
    return True

def clean_leaves_only():
    conn = get_db_connection()
    conn.execute("DELETE FROM leaves")
    conn.commit()
    conn.close()
    increment_data_version()
    return True

def clean_missions_only():
    conn = get_db_connection()
    conn.execute("DELETE FROM missions")
    conn.commit()
    conn.close()
    increment_data_version()
    return True

# ============================================================
# HELPER FUNCTIONS
# ============================================================

def haversine_distance(lat1, lon1, lat2, lon2):
    R = 6371000
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    a = math.sin(delta_phi/2)**2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    return R * c

def save_uploaded_file(file, user_id, folder_type):
    if file and file.filename:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_id = str(uuid.uuid4())[:8]
        file_ext = os.path.splitext(file.filename)[1]
        new_filename = f"{folder_type}_{user_id}_{timestamp}_{unique_id}{file_ext}"

        if folder_type == 'leave':
            upload_folder = UPLOAD_FOLDER_LEAVES
        else:
            upload_folder = UPLOAD_FOLDER_MISSIONS

        file_path = os.path.join(upload_folder, new_filename)
        file.save(file_path)
        return f"/static/uploads/{folder_type}s/{new_filename}"
    return None

# ============================================================
# ROUTES
# ============================================================

@app.route('/')
def home():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    return redirect(url_for('dashboard'))

@app.route('/dashboard')
def dashboard():
    if not session.get('logged_in'):
        return redirect(url_for('login'))

    stats = get_attendance_stats()

    today = datetime.now()
    start_date = today.replace(day=1).strftime('%Y-%m-%d')
    end_date = today.strftime('%Y-%m-%d')
    work_history = get_work_history_report(start_date, end_date, 200)
    monthly_summary = get_monthly_summary_report(start_date, end_date)
    company_loc = get_company_location()

    pending_leaves = get_pending_leaves()
    pending_missions = get_pending_missions()
    pending_count = len(pending_leaves) + len(pending_missions)

    data_version = get_data_version()
    current_month = today.strftime('%m')
    current_year = today.strftime('%Y')
    allowed_distance = app.config['ALLOWED_DISTANCE']

    lock_status = get_system_lock_status()
    user_lock = get_user_lock_status(session.get('user_id'))

    return render_template_string(DASHBOARD_HTML,
                                   session=session,
                                   stats=stats,
                                   work_history=work_history,
                                   monthly_summary=monthly_summary,
                                   company_lat=company_loc['lat'] if company_loc else '',
                                   company_lng=company_loc['lng'] if company_loc else '',
                                   pending_count=pending_count,
                                   data_version=data_version,
                                   current_month=current_month,
                                   current_year=current_year,
                                   allowed_distance=allowed_distance,
                                   lock_status=lock_status,
                                   user_lock=user_lock)

@app.route('/get_data_version')
def get_data_version_route():
    if not session.get('logged_in'):
        return jsonify({'version': 0})
    version = get_data_version()
    return jsonify({'version': version})

@app.route('/get_checkin_status')
def get_checkin_status_route():
    if not session.get('logged_in'):
        return jsonify({'has_checkin': False, 'check_in_time': None, 'shift': None})

    user_id = session.get('user_id')
    if not user_id:
        return jsonify({'has_checkin': False, 'check_in_time': None, 'shift': None})

    status = get_checkin_status(user_id)
    return jsonify(status)

@app.route('/get_system_lock_status')
def get_system_lock_status_route():
    if not session.get('logged_in'):
        return jsonify({'is_locked': 0, 'auto_unlock_time': '06:00'})
    lock = get_system_lock_status()
    return jsonify(lock)

@app.route('/toggle_system_lock', methods=['POST'])
def toggle_system_lock_route():
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'អ្នកមិនមែនជា Admin!'})

    data = request.get_json()
    lock_state = data.get('lock_state', 0)
    auto_unlock_time = data.get('auto_unlock_time', '06:00')
    user_id = session.get('user_id')

    result = toggle_system_lock(lock_state, auto_unlock_time, user_id)
    if result:
        status = "បិទ" if lock_state == 1 else "បើក"
        return jsonify({'success': True, 'message': f'✅ បាន{status}ប្រព័ន្ធចូលធ្វើការជោគជ័យ!'})
    return jsonify({'success': False, 'message': '❌ មិនអាចប្តូរស្ថានភាពបាន!'})

# ============================================================
# USER LOCK ROUTES
# ============================================================

@app.route('/get_user_lock/<int:user_id>')
def get_user_lock_route(user_id):
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    lock = get_user_lock_status(user_id)
    return jsonify(lock)

@app.route('/get_all_user_locks')
def get_all_user_locks_route():
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify([]), 403
    locks = get_all_user_lock_status()
    return jsonify(locks)

@app.route('/toggle_user_lock', methods=['POST'])
def toggle_user_lock_route():
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'អ្នកមិនមែនជា Admin!'})

    data = request.get_json()
    user_id = data.get('user_id')
    lock_state = data.get('lock_state', 0)
    auto_unlock_time = data.get('auto_unlock_time', '')
    admin_id = session.get('user_id')

    if not user_id:
        return jsonify({'success': False, 'message': 'មិនមាន user_id!'})

    if auto_unlock_time:
        try:
            datetime.strptime(auto_unlock_time, '%H:%M')
        except ValueError:
            return jsonify({'success': False, 'message': 'ទ្រង់ទ្រាយម៉ោងមិនត្រឹមត្រូវ! សូមប្រើ HH:MM'})

    if user_id == admin_id:
        return jsonify({'success': False, 'message': '❌ អ្នកមិនអាចបិទគណនីរបស់ខ្លួនឯងបានទេ!'})

    result = toggle_user_lock(user_id, lock_state, auto_unlock_time if auto_unlock_time else None, admin_id)
    if result:
        user = get_user_by_id(user_id)
        status = "បិទ" if lock_state == 1 else "បើក"
        return jsonify({'success': True, 'message': f'✅ បាន{status}ការចូលធ្វើការរបស់ {user["full_name"]} ជោគជ័យ!'})
    return jsonify({'success': False, 'message': '❌ មិនអាចប្តូរស្ថានភាពបាន!'})

@app.route('/check_in', methods=['POST'])
def check_in_route():
    try:
        if not session.get('logged_in'):
            return jsonify({'success': False, 'message': 'មិនទាន់ Login'})

        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'success': False, 'message': 'មិនមាន user_id ក្នុង Session'})

        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'មិនមានទិន្នន័យផ្ញើមក'})

        user_lat = data.get('user_lat')
        user_lng = data.get('user_lng')
        company_lat = data.get('company_lat')
        company_lng = data.get('company_lng')
        shift = data.get('shift', 1)

        if user_lat is None or user_lng is None:
            return jsonify({'success': False, 'message': 'មិនមានទីតាំងរបស់អ្នក!'})

        if company_lat is None or company_lng is None:
            return jsonify({'success': False, 'message': 'សូមឲ្យ Admin កំណត់ទីតាំងក្រុមហ៊ុនជាមុនសិន!'})

        allowed, lock_message = check_system_lock_for_user(user_id)
        if not allowed:
            return jsonify({'success': False, 'message': lock_message})

        allowed, user_lock_message = check_user_lock(user_id)
        if not allowed:
            return jsonify({'success': False, 'message': user_lock_message})

        can_checkin, deadline = check_attendance_deadline(user_id)
        if not can_checkin:
            system_lock = get_system_lock_status()
            auto_unlock = system_lock.get('auto_unlock_time', '06:00')
            update_user_lock(user_id, 1, auto_unlock_time=auto_unlock)
            increment_data_version()
            return jsonify({'success': False, 'message': f'⛔ អ្នកលើសម៉ោងដែល Admin បានកំណត់ (ម៉ោងកំណត់: {deadline})! ប្រព័ន្ធបានបិទការចូលធ្វើការរបស់អ្នកដោយស្វ័យប្រវត្តិ!'})

        distance = haversine_distance(user_lat, user_lng, company_lat, company_lng)
        allowed_distance = app.config['ALLOWED_DISTANCE']

        if distance > allowed_distance:
            return jsonify({
                'success': False,
                'message': f'អ្នកនៅឆ្ងាយពីទីតាំងក្រុមហ៊ុន {round(distance, 2)} ម៉ែត្រ (អនុញ្ញាតត្រឹម {allowed_distance} ម៉ែត្រ)! មិនអាចចូលធ្វើការបាន!'
            })

        status = get_checkin_status(user_id)
        if status['has_checkin']:
            return jsonify({'success': False, 'message': 'អ្នកបានចូលធ្វើការរួចហើយ! សូមចុច "ចេញពីធ្វើការ" មុនពេលចូលម្តងទៀត!'})

        result, message = check_in(user_id, user_lat, user_lng, distance, shift)
        if result:
            shift_names = {1: 'វគ្គ 1 (ព្រឹក)', 2: 'វគ្គ 2 (រសៀល)', 3: 'វគ្គ 3 (យប់)'}
            return jsonify({
                'success': True,
                'message': f'✅ {message} ({shift_names.get(shift, "វគ្គ " + str(shift))}) (ចម្ងាយ {round(distance, 2)} ម៉ែត្រ)'
            })
        else:
            return jsonify({'success': False, 'message': '❌ ' + message})

    except Exception as e:
        print(f"Error in check_in: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'កំហុស: {str(e)}'})

@app.route('/check_out', methods=['POST'])
def check_out_route():
    try:
        if not session.get('logged_in'):
            return jsonify({'success': False, 'message': 'មិនទាន់ Login'})

        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'success': False, 'message': 'មិនមាន user_id ក្នុង Session'})

        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'មិនមានទិន្នន័យផ្ញើមក'})

        user_lat = data.get('user_lat')
        user_lng = data.get('user_lng')
        company_lat = data.get('company_lat')
        company_lng = data.get('company_lng')

        if user_lat is None or user_lng is None:
            return jsonify({'success': False, 'message': 'មិនមានទីតាំងរបស់អ្នក!'})

        if company_lat is None or company_lng is None:
            return jsonify({'success': False, 'message': 'សូមឲ្យ Admin កំណត់ទីតាំងក្រុមហ៊ុនជាមុនសិន!'})

        allowed, lock_message = check_system_lock_for_user(user_id)
        if not allowed:
            return jsonify({'success': False, 'message': lock_message})

        allowed, user_lock_message = check_user_lock(user_id)
        if not allowed:
            return jsonify({'success': False, 'message': user_lock_message})

        distance = haversine_distance(user_lat, user_lng, company_lat, company_lng)
        allowed_distance = app.config['ALLOWED_DISTANCE']

        if distance > allowed_distance:
            return jsonify({
                'success': False,
                'message': f'អ្នកនៅឆ្ងាយពីទីតាំងក្រុមហ៊ុន {round(distance, 2)} ម៉ែត្រ (អនុញ្ញាតត្រឹម {allowed_distance} ម៉ែត្រ)! មិនអាចចេញធ្វើការបាន!'
            })

        status = get_checkin_status(user_id)
        if not status['has_checkin']:
            return jsonify({'success': False, 'message': 'អ្នកមិនទាន់ចូលធ្វើការនៅថ្ងៃនេះទេ!'})

        result, message = check_out(user_id, user_lat, user_lng, distance)
        if result:
            return jsonify({'success': True, 'message': '✅ ' + message})
        else:
            return jsonify({'success': False, 'message': '❌ ' + message})

    except Exception as e:
        print(f"Error in check_out: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'កំហុស: {str(e)}'})

@app.route('/get_company_location')
def get_company_location_route():
    location = get_company_location()
    if location:
        return jsonify({'lat': location['lat'], 'lng': location['lng']})
    return jsonify({'lat': None, 'lng': None})

@app.route('/save_location', methods=['POST'])
def save_location():
    if not session.get('logged_in'):
        return jsonify({'success': False, 'message': 'មិនទាន់ Login'})
    data = request.get_json()
    lat = data.get('lat')
    lng = data.get('lng')
    if lat is None or lng is None:
        return jsonify({'success': False, 'message': 'ទិន្នន័យមិនត្រឹមត្រូវ'})
    result = save_company_location(lat, lng)
    if result:
        return jsonify({'success': True, 'message': 'រក្សាទុកជោគជ័យ'})
    return jsonify({'success': False, 'message': 'មិនអាចរក្សាទុកបាន'})

@app.route('/register', methods=['GET', 'POST'])
def register():
    if not session.get('logged_in') or session.get('role') != 'admin':
        return redirect(url_for('dashboard'))
    message = ''
    message_type = ''
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        full_name = request.form.get('full_name')
        email = request.form.get('email')
        phone = request.form.get('phone')
        role = request.form.get('role')
        if not username or not password or not full_name:
            message = 'សូមបំពេញព័ត៌មានឲ្យបានពេញលេញ!'
            message_type = 'error'
        elif password != confirm_password:
            message = 'ពាក្យសម្ងាត់មិនត្រូវគ្នា!'
            message_type = 'error'
        elif len(password) < 4:
            message = 'ពាក្យសម្ងាត់ត្រូវមានយ៉ាងតិច 4 តួ!'
            message_type = 'error'
        else:
            if create_user(username, password, full_name, email, phone, role):
                message = f'✅ ចុះឈ្មោះអ្នកប្រើ "{username}" ជោគជ័យ!'
                message_type = 'success'
            else:
                message = f'❌ ឈ្មោះអ្នកប្រើ "{username}" មានរួចហើយ!'
                message_type = 'error'
    return render_template_string(REGISTER_HTML, session=session, message=message, message_type=message_type)

@app.route('/manage_users')
def manage_users():
    if not session.get('logged_in') or session.get('role') != 'admin':
        return redirect(url_for('dashboard'))
    users = get_all_users()
    settings = get_all_attendance_settings()
    settings_dict = {s['user_id']: s for s in settings}
    user_locks = get_all_user_lock_status()
    user_locks_dict = {u['id']: u for u in user_locks}
    return render_template_string(USER_MANAGEMENT_HTML,
                                   session=session,
                                   users=users,
                                   settings=settings_dict,
                                   user_locks=user_locks_dict,
                                   message='',
                                   message_type='')

@app.route('/get_user/<int:user_id>')
def get_user(user_id):
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    user = get_user_by_id(user_id)
    if not user:
        return jsonify({'error': 'User not found'}), 404
    return jsonify({'id': user['id'], 'username': user['username'], 'full_name': user['full_name'], 'email': user['email'] or '', 'phone': user['phone'] or '', 'role': user['role']})

@app.route('/get_attendance_setting/<int:user_id>')
def get_attendance_setting_route(user_id):
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    setting = get_attendance_setting(user_id)
    if setting:
        return jsonify(setting)
    return jsonify({'user_id': user_id, 'check_in_deadline': '', 'is_active': 0})

@app.route('/save_attendance_setting', methods=['POST'])
def save_attendance_setting_route():
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'អ្នកមិនមែនជា Admin!'})

    data = request.get_json()
    user_id = data.get('user_id')
    check_in_deadline = data.get('check_in_deadline', '').strip()
    is_active = data.get('is_active', 0)

    if not user_id:
        return jsonify({'success': False, 'message': 'មិនមាន user_id!'})

    if check_in_deadline:
        try:
            datetime.strptime(check_in_deadline, '%H:%M')
        except ValueError:
            return jsonify({'success': False, 'message': 'ទ្រង់ទ្រាយម៉ោងមិនត្រឹមត្រូវ! សូមប្រើ HH:MM (ឧទាហរណ៍: 08:00)'})

    result = save_attendance_setting(user_id, check_in_deadline, int(is_active))
    if result:
        return jsonify({'success': True, 'message': '✅ រក្សាទុកការកំណត់ជោគជ័យ!'})
    return jsonify({'success': False, 'message': '❌ មិនអាចរក្សាទុកបាន!'})

@app.route('/update_user/<int:user_id>', methods=['POST'])
def update_user_route(user_id):
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'អ្នកមិនមែនជា Admin!'})
    data = request.get_json()
    result = update_user(user_id, data.get('username'), data.get('full_name'), data.get('email'), data.get('phone'), data.get('role'))
    if result:
        return jsonify({'success': True, 'message': '✅ កែប្រែអ្នកប្រើជោគជ័យ!'})
    return jsonify({'success': False, 'message': '❌ មិនអាចកែប្រែបាន!'})

@app.route('/add_user', methods=['POST'])
def add_user_route():
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'អ្នកមិនមែនជា Admin!'})
    data = request.get_json()
    if create_user(data.get('username'), data.get('password'), data.get('full_name'), data.get('email'), data.get('phone'), data.get('role')):
        return jsonify({'success': True, 'message': f'✅ បានបន្ថែមអ្នកប្រើ "{data.get("username")}" ជោគជ័យ!'})
    return jsonify({'success': False, 'message': '❌ ឈ្មោះអ្នកប្រើមានរួចហើយ!'})

@app.route('/delete_user/<int:user_id>', methods=['POST'])
def delete_user_route(user_id):
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'អ្នកមិនមែនជា Admin!'})
    if user_id == session.get('user_id'):
        return jsonify({'success': False, 'message': '❌ អ្នកមិនអាចលុបគណនីរបស់ខ្លួនឯងបានទេ!'})
    if delete_user(user_id):
        return jsonify({'success': True, 'message': '✅ លុបអ្នកប្រើជោគជ័យ!'})
    return jsonify({'success': False, 'message': '❌ មិនអាចលុបបាន!'})

@app.route('/change_password', methods=['GET', 'POST'])
def change_password():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    message = ''
    message_type = ''
    user_id = session.get('user_id')
    if request.method == 'POST':
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        if not current_password or not new_password or not confirm_password:
            message = 'សូមបំពេញព័ត៌មានឲ្យបានពេញលេញ!'
            message_type = 'error'
        elif new_password != confirm_password:
            message = 'ពាក្យសម្ងាត់ថ្មីមិនត្រូវគ្នា!'
            message_type = 'error'
        elif len(new_password) < 4:
            message = 'ពាក្យសម្ងាត់ថ្មីត្រូវមានយ៉ាងតិច 4 តួ!'
            message_type = 'error'
        elif not verify_password(user_id, current_password):
            message = 'ពាក្យសម្ងាត់បច្ចុប្បន្នមិនត្រឹមត្រូវ!'
            message_type = 'error'
        else:
            if update_password(user_id, new_password):
                message = '✅ បានប្តូរពាក្យសម្ងាត់ជោគជ័យ!'
                message_type = 'success'
            else:
                message = '❌ មិនអាចប្តូរពាក្យសម្ងាត់បាន!'
                message_type = 'error'
    return render_template_string(CHANGE_PASSWORD_HTML, session=session, message=message, message_type=message_type)

@app.route('/admin_reset_password/<int:user_id>', methods=['POST'])
def admin_reset_password(user_id):
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'អ្នកមិនមែនជា Admin!'})
    data = request.get_json()
    new_password = data.get('new_password')
    if not new_password or len(new_password) < 4:
        return jsonify({'success': False, 'message': 'ពាក្យសម្ងាត់ត្រូវមានយ៉ាងតិច 4 តួ!'})
    if update_password(user_id, new_password):
        return jsonify({'success': True, 'message': '✅ បានកំណត់ពាក្យសម្ងាត់ថ្មីជោគជ័យ!'})
    return jsonify({'success': False, 'message': '❌ មិនអាចកំណត់ពាក្យសម្ងាត់បាន!'})

@app.route('/clean_data', methods=['POST'])
def clean_data():
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'អ្នកមិនមែនជា Admin!'})
    data = request.get_json()
    clean_type = data.get('type', 'all')
    if clean_type == 'all':
        result = clean_all_data()
        message = '✅ បានសម្អាតទិន្នន័យទាំងអស់ជោគជ័យ!'
    elif clean_type == 'attendance':
        result = clean_attendance_only()
        message = '✅ បានសម្អាតទិន្នន័យវត្តមានជោគជ័យ!'
    elif clean_type == 'leaves':
        result = clean_leaves_only()
        message = '✅ បានសម្អាតទិន្នន័យសុំច្បាប់ជោគជ័យ!'
    elif clean_type == 'missions':
        result = clean_missions_only()
        message = '✅ បានសម្អាតទិន្នន័យបេសកម្មជោគជ័យ!'
    else:
        return jsonify({'success': False, 'message': 'ប្រភេទមិនត្រឹមត្រូវ!'})
    if result:
        return jsonify({'success': True, 'message': message})
    return jsonify({'success': False, 'message': '❌ មិនអាចសម្អាតទិន្នន័យបាន!'})

@app.route('/request_leave', methods=['POST'])
def request_leave():
    if not session.get('logged_in'):
        return jsonify({'success': False, 'message': 'មិនទាន់ Login'})

    user_id = session.get('user_id')

    attachment = None
    if 'attachment' in request.files:
        file = request.files['attachment']
        attachment = save_uploaded_file(file, user_id, 'leave')

    if request.is_json:
        data = request.get_json()
        result = create_leave(
            user_id,
            data.get('start_date'),
            data.get('end_date'),
            data.get('days'),
            data.get('reason', ''),
            attachment
        )
        days = data.get('days')
    else:
        result = create_leave(
            user_id,
            request.form.get('start_date'),
            request.form.get('end_date'),
            request.form.get('days'),
            request.form.get('reason', ''),
            attachment
        )
        days = request.form.get('days')

    if result:
        return jsonify({'success': True, 'message': f'✅ បានសុំច្បាប់ {days} ថ្ងៃរួចហើយ! រង់ចាំការអនុម័តពី Admin'})
    return jsonify({'success': False, 'message': '❌ មិនអាចសុំច្បាប់បាន!'})

@app.route('/request_mission', methods=['POST'])
def request_mission():
    if not session.get('logged_in'):
        return jsonify({'success': False, 'message': 'មិនទាន់ Login'})

    user_id = session.get('user_id')

    attachment = None
    if 'attachment' in request.files:
        file = request.files['attachment']
        attachment = save_uploaded_file(file, user_id, 'mission')

    if request.is_json:
        data = request.get_json()
        result = create_mission(
            user_id,
            data.get('start_date'),
            data.get('end_date'),
            data.get('days'),
            data.get('destination', ''),
            data.get('purpose', ''),
            attachment
        )
        days = data.get('days')
    else:
        result = create_mission(
            user_id,
            request.form.get('start_date'),
            request.form.get('end_date'),
            request.form.get('days'),
            request.form.get('destination', ''),
            request.form.get('purpose', ''),
            attachment
        )
        days = request.form.get('days')

    if result:
        return jsonify({'success': True, 'message': f'✅ បានសុំបេសកម្ម {days} ថ្ងៃរួចហើយ! រង់ចាំការអនុម័តពី Admin'})
    return jsonify({'success': False, 'message': '❌ មិនអាចសុំបេសកម្មបាន!'})

@app.route('/get_pending_requests')
def get_pending_requests():
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'leaves': [], 'missions': []})
    leaves = get_pending_leaves()
    missions = get_pending_missions()
    return jsonify({'leaves': leaves, 'missions': missions})

@app.route('/check_new_requests')
def check_new_requests():
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'new': False, 'count': 0, 'message': ''})
    pending_leaves = get_pending_leaves()
    pending_missions = get_pending_missions()
    total_pending = len(pending_leaves) + len(pending_missions)
    last_count = session.get('last_pending_count', 0)
    is_new = total_pending > last_count
    session['last_pending_count'] = total_pending
    message = ''
    if is_new and total_pending > 0:
        new_count = total_pending - last_count
        message = f'📬 មានសំណើថ្មី {new_count} ករណី! សូមចុច "📬 សំណើ" ដើម្បីអនុម័ត'
    return jsonify({'new': is_new, 'count': total_pending, 'message': message})

@app.route('/approve_request', methods=['POST'])
def approve_request():
    try:
        if not session.get('logged_in'):
            return jsonify({'success': False, 'message': 'សូម Login ជាមុន!'})

        if session.get('role') != 'admin':
            return jsonify({'success': False, 'message': 'អ្នកមិនមែនជា Admin!'})

        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'មិនមានទិន្នន័យផ្ញើមក!'})

        req_type = data.get('type')
        req_id = data.get('id')
        admin_id = session.get('user_id')

        print(f"🔄 Approving: type={req_type}, id={req_id}, admin={admin_id}")

        if req_type == 'leave':
            result = approve_leave(req_id, admin_id)
            print(f"📝 Approve leave result: {result}")
        elif req_type == 'mission':
            result = approve_mission(req_id, admin_id)
            print(f"📝 Approve mission result: {result}")
        else:
            return jsonify({'success': False, 'message': 'ប្រភេទមិនត្រឹមត្រូវ'})

        if result:
            return jsonify({'success': True, 'message': '✅ បានអនុម័តជោគជ័យ!'})
        return jsonify({'success': False, 'message': '❌ មិនអាចអនុម័តបាន!'})
    except Exception as e:
        print(f"❌ Error in approve_request: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'កំហុស: {str(e)}'})

@app.route('/reject_request', methods=['POST'])
def reject_request():
    try:
        if not session.get('logged_in'):
            return jsonify({'success': False, 'message': 'សូម Login ជាមុន!'})

        if session.get('role') != 'admin':
            return jsonify({'success': False, 'message': 'អ្នកមិនមែនជា Admin!'})

        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'មិនមានទិន្នន័យផ្ញើមក!'})

        req_type = data.get('type')
        req_id = data.get('id')

        print(f"🔄 Rejecting: type={req_type}, id={req_id}")

        if req_type == 'leave':
            result = reject_leave(req_id)
            print(f"📝 Reject leave result: {result}")
        elif req_type == 'mission':
            result = reject_mission(req_id)
            print(f"📝 Reject mission result: {result}")
        else:
            return jsonify({'success': False, 'message': 'ប្រភេទមិនត្រឹមត្រូវ'})

        if result:
            return jsonify({'success': True, 'message': '✅ បានបដិសេធជោគជ័យ!'})
        return jsonify({'success': False, 'message': '❌ មិនអាចបដិសេធបាន!'})
    except Exception as e:
        print(f"❌ Error in reject_request: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'កំហុស: {str(e)}'})

@app.route('/report')
def report():
    if not session.get('logged_in'):
        return redirect(url_for('login'))

    start_date = request.args.get('start')
    end_date = request.args.get('end')
    month = request.args.get('month')
    year = request.args.get('year')
    current_year = datetime.now().year

    if month and year:
        month = int(month)
        year = int(year)
        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year}-12-31"
        else:
            next_month = month + 1
            end_dt = datetime(year, next_month, 1) - timedelta(days=1)
            end_date = end_dt.strftime('%Y-%m-%d')
    elif start_date and end_date:
        pass
    else:
        today = datetime.now()
        start_date = today.replace(day=1).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')

    result = get_attendance_report(start_date, end_date)
    daily_data = result.get('daily', [])
    summary_data = result.get('summary', [])

    return render_template_string(REPORT_HTML,
                                   session=session,
                                   daily=daily_data,
                                   summary=summary_data,
                                   start_date=start_date,
                                   end_date=end_date,
                                   current_year=current_year,
                                   user_role=session.get('role', 'user'))

@app.route('/export_excel')
def export_excel():
    if not session.get('logged_in'):
        return redirect(url_for('login'))

    start_date = request.args.get('start')
    end_date = request.args.get('end')
    month = request.args.get('month')
    year = request.args.get('year')

    if month and year:
        month = int(month)
        year = int(year)
        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year}-12-31"
        else:
            next_month = month + 1
            end_dt = datetime(year, next_month, 1) - timedelta(days=1)
            end_date = end_dt.strftime('%Y-%m-%d')
    elif start_date and end_date:
        pass
    else:
        today = datetime.now()
        start_date = today.replace(day=1).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')

    result = get_attendance_report(start_date, end_date)
    daily_data = result.get('daily', [])
    summary_data = result.get('summary', [])

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "ប្រចាំថ្ងៃ"

    ws1.merge_cells('A1:H1')
    ws1['A1'] = f"របាយការណ៍ប្រចាំថ្ងៃ\nចាប់ពី {start_date} ដល់ {end_date}"
    ws1['A1'].font = Font(size=16, bold=True)
    ws1['A1'].alignment = Alignment(horizontal='center')

    headers = ['ល.រ', 'ឈ្មោះបុគ្គលិក', 'កាលបរិច្ឆេទ', 'ប្រភេទ', 'ចាប់ផ្តើម', 'បញ្ចប់', 'ចំនួន', 'ស្ថានភាព/ព័ត៌មាន']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1a73e8', end_color='1a73e8', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    for row_idx, item in enumerate(daily_data, 4):
        ws1.cell(row=row_idx, column=1, value=row_idx-3)
        ws1.cell(row=row_idx, column=2, value=item.get('full_name', ''))
        ws1.cell(row=row_idx, column=3, value=item.get('date', ''))

        if item.get('type') == 'attendance':
            ws1.cell(row=row_idx, column=4, value='វត្តមាន')
        elif item.get('type') == 'leave':
            ws1.cell(row=row_idx, column=4, value='ច្បាប់')
        elif item.get('type') == 'mission':
            ws1.cell(row=row_idx, column=4, value='បេសកម្ម')
        else:
            ws1.cell(row=row_idx, column=4, value='-')

        if item.get('type') == 'attendance':
            ws1.cell(row=row_idx, column=5, value=item.get('check_in', '')[:16] if item.get('check_in') else '')
            ws1.cell(row=row_idx, column=6, value=item.get('check_out', '')[:16] if item.get('check_out') else '')
            if item.get('total_hours'):
                hours = int(item['total_hours'])
                minutes = int((item['total_hours'] - hours) * 60)
                ws1.cell(row=row_idx, column=7, value=f"{hours:02d}:{minutes:02d}")
            else:
                ws1.cell(row=row_idx, column=7, value='-')
            if item.get('check_in') and item.get('check_out'):
                ws1.cell(row=row_idx, column=8, value='បានបិទ')
            elif item.get('check_in') and not item.get('check_out'):
                ws1.cell(row=row_idx, column=8, value='កំពុងធ្វើការ')
            else:
                ws1.cell(row=row_idx, column=8, value='មិនទាន់ចូល')
        else:
            ws1.cell(row=row_idx, column=5, value=item.get('start_date', ''))
            ws1.cell(row=row_idx, column=6, value=item.get('end_date', ''))
            ws1.cell(row=row_idx, column=7, value=f"{item.get('days', 0)} ថ្ងៃ")
            if item.get('type') == 'leave':
                ws1.cell(row=row_idx, column=8, value=f"ច្បាប់: {item.get('reason', '')}")
            else:
                ws1.cell(row=row_idx, column=8, value=f"បេសកម្ម: {item.get('destination', '')}")

    for col in range(1, 9):
        ws1.column_dimensions[chr(64 + col)].width = 18

    ws2 = wb.create_sheet("សង្ខេបប្រចាំខែ")

    ws2.merge_cells('A1:H1')
    ws2['A1'] = f"របាយការណ៍សង្ខេបប្រចាំខែ\nចាប់ពី {start_date} ដល់ {end_date}"
    ws2['A1'].font = Font(size=16, bold=True)
    ws2['A1'].alignment = Alignment(horizontal='center')

    headers2 = ['ល.រ', 'ឈ្មោះបុគ្គលិក', 'ថ្ងៃធ្វើការ', 'ម៉ោងធ្វើការសរុប', 'វគ្គយប់', 'ម៉ោងយប់', 'ចំនួនថ្ងៃសុំច្បាប់', 'ចំនួនថ្ងៃបេសកម្ម']
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1a73e8', end_color='1a73e8', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    for row_idx, item in enumerate(summary_data, 4):
        ws2.cell(row=row_idx, column=1, value=row_idx-3)
        ws2.cell(row=row_idx, column=2, value=item.get('full_name', ''))
        ws2.cell(row=row_idx, column=3, value=item.get('days_worked', 0))

        total_hours = item.get('total_hours', 0)
        hours = int(total_hours)
        minutes = int((total_hours - hours) * 60)
        ws2.cell(row=row_idx, column=4, value=f"{hours:02d}:{minutes:02d}")

        ws2.cell(row=row_idx, column=5, value=item.get('night_shifts', 0))

        night_hours = item.get('night_hours', 0)
        nhours = int(night_hours)
        nminutes = int((night_hours - nhours) * 60)
        ws2.cell(row=row_idx, column=6, value=f"{nhours:02d}:{nminutes:02d}")

        ws2.cell(row=row_idx, column=7, value=item.get('total_leave_days', 0))
        ws2.cell(row=row_idx, column=8, value=item.get('total_mission_days', 0))

    for col in range(1, 9):
        ws2.column_dimensions[chr(64 + col)].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"របាយការណ៍_{start_date}_ដល់_{end_date}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/check_auto_unlock')
def check_auto_unlock():
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'})

    try:
        current_time = get_current_time_only()[:5]
        unlocked_items = []

        lock = get_system_lock_status()
        if lock.get('is_locked', 0) == 1:
            auto_unlock = lock.get('auto_unlock_time')
            if auto_unlock and current_time >= auto_unlock:
                update_system_lock(0, locked_by=None)
                unlocked_items.append('system')
                print(f"✅ System auto-unlocked via API at {current_time}")

        conn = get_db_connection()
        locked_users = conn.execute('''
            SELECT user_id, auto_unlock_time
            FROM user_attendance_lock
            WHERE is_locked = 1
            AND auto_unlock_time IS NOT NULL
            AND auto_unlock_time != ''
        ''').fetchall()
        conn.close()

        for user in locked_users:
            auto_unlock = user['auto_unlock_time']
            if auto_unlock and current_time >= auto_unlock:
                update_user_lock(user['user_id'], 0)
                unlocked_items.append(f"user_{user['user_id']}")
                print(f"✅ User {user['user_id']} auto-unlocked via API at {current_time}")

        if unlocked_items:
            increment_data_version()
            return jsonify({
                'success': True,
                'message': f'Auto-unlocked: {", ".join(unlocked_items)}',
                'unlocked': unlocked_items
            })
        else:
            return jsonify({
                'success': True,
                'message': 'No auto-unlock needed at this time',
                'unlocked': []
            })

    except Exception as e:
        print(f"❌ Error in check_auto_unlock: {e}")
        return jsonify({'success': False, 'message': str(e)})

# ============================================================
# PWA & STATIC ROUTES
# ============================================================

@app.route('/static/manifest.json')
def manifest():
    return {
        "name": "ប្រព័ន្ធគ្រប់គ្រងបុគ្គលិក",
        "short_name": "HR System",
        "start_url": "/dashboard",
        "display": "standalone",
        "background_color": "#1a73e8",
        "theme_color": "#1a73e8",
        "orientation": "portrait",
        "icons": [
            {
                "src": "/static/icon-192.png",
                "sizes": "192x192",
                "type": "image/png",
                "purpose": "any maskable"
            },
            {
                "src": "/static/icon-512.png",
                "sizes": "512x512",
                "type": "image/png",
                "purpose": "any maskable"
            }
        ]
    }

@app.route('/static/sw.js')
def service_worker():
    return '''
const CACHE_NAME = 'hr-system-v1';
const ASSETS = [
    '/',
    '/dashboard',
    '/static/manifest.json'
];

self.addEventListener('install', function(e) {
    e.waitUntil(
        caches.open(CACHE_NAME).then(function(cache) {
            return cache.addAll(ASSETS);
        }).then(function() {
            return self.skipWaiting();
        })
    );
});

self.addEventListener('activate', function(e) {
    e.waitUntil(
        caches.keys().then(function(cacheNames) {
            return Promise.all(
                cacheNames.filter(function(name) {
                    return name !== CACHE_NAME;
                }).map(function(name) {
                    return caches.delete(name);
                })
            );
        }).then(function() {
            return self.clients.claim();
        })
    );
});

self.addEventListener('fetch', function(e) {
    e.respondWith(
        caches.match(e.request).then(function(response) {
            return response || fetch(e.request).then(function(fetchResponse) {
                return caches.open(CACHE_NAME).then(function(cache) {
                    cache.put(e.request, fetchResponse.clone());
                    return fetchResponse;
                });
            });
        }).catch(function() {
            return new Response('Offline', {
                status: 503,
                statusText: 'Service Unavailable'
            });
        })
    );
});
    ''', 200, {'Content-Type': 'application/javascript'}

from flask import send_from_directory

@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory('static', filename)

@app.route('/edit_attendance/<int:attendance_id>', methods=['GET', 'POST'])
def edit_attendance(attendance_id):
    if not session.get('logged_in') or session.get('role') != 'admin':
        return redirect(url_for('dashboard'))

    attendance = get_attendance_by_id(attendance_id)
    if not attendance:
        return "មិនមានទិន្នន័យ", 404

    if request.method == 'POST':
        check_in_date = request.form.get('check_in_date')
        check_in_time = request.form.get('check_in_time')
        check_out_date = request.form.get('check_out_date')
        check_out_time = request.form.get('check_out_time')
        shift = request.form.get('shift')

        check_in_full = None
        check_out_full = None

        if check_in_date and check_in_time:
            check_in_full = f"{check_in_date} {check_in_time}:00"
        if check_out_date and check_out_time:
            check_out_full = f"{check_out_date} {check_out_time}:00"

        result = update_attendance(attendance_id, check_in_full, check_out_full, check_in_date, shift)
        if result:
            return redirect(url_for('dashboard'))
        else:
            return "មិនអាចកែប្រែបាន!", 500

    edit_form_html = '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>កែប្រែទិន្នន័យវត្តមាន</title>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <style>
            * { margin: 0; padding: 0; box-sizing: border-box; }
            body { font-family: 'Khmer OS', 'Arial', sans-serif; background: #f0f2f5; padding: 20px; }
            .container { max-width: 600px; margin: 0 auto; background: white; padding: 30px; border-radius: 12px; box-shadow: 0 2px 10px rgba(0,0,0,0.08); }
            h2 { color: #1a73e8; margin-bottom: 20px; }
            .info { background: #f8f9fa; padding: 12px; border-radius: 8px; margin-bottom: 20px; }
            .info p { margin: 5px 0; }
            label { display: block; margin: 12px 0 5px; font-weight: 600; color: #555; }
            input, select { width: 100%; padding: 10px 14px; border: 2px solid #e8ecf1; border-radius: 8px; font-size: 14px; font-family: 'Khmer OS', 'Arial', sans-serif; }
            input:focus, select:focus { outline: none; border-color: #1a73e8; }
            .row { display: grid; grid-template-columns: 1fr 1fr; gap: 10px; }
            .btn-group { display: flex; gap: 10px; margin-top: 20px; flex-wrap: wrap; }
            .btn-save { flex: 1; background: #1a73e8; color: white; padding: 12px 20px; border: none; border-radius: 8px; cursor: pointer; font-size: 16px; font-family: 'Khmer OS', 'Arial', sans-serif; }
            .btn-save:hover { background: #1557b0; }
            .btn-cancel { flex: 1; background: #dc3545; color: white; padding: 12px 20px; border: none; border-radius: 8px; cursor: pointer; font-size: 16px; font-family: 'Khmer OS', 'Arial', sans-serif; text-decoration: none; text-align: center; }
            .btn-cancel:hover { background: #b02a37; }
            .hint { font-size: 12px; color: #888; margin-top: 3px; }
            @media (max-width: 600px) {
                .container { padding: 20px; }
                .row { grid-template-columns: 1fr; }
                .btn-group { flex-direction: column; }
            }
        </style>
    </head>
    <body>
        <div class="container">
            <h2>✏️ កែប្រែទិន្នន័យវត្តមាន</h2>
            <div class="info">
                <p><strong>ឈ្មោះបុគ្គលិក:</strong> {{ attendance.full_name }}</p>
                <p><strong>ឈ្មោះអ្នកប្រើ:</strong> {{ attendance.username }}</p>
                <p><strong>វគ្គបច្ចុប្បន្ន:</strong>
                    {% if attendance.shift == 1 %}វគ្គ 1 (ព្រឹក 07:00-11:00)
                    {% elif attendance.shift == 2 %}វគ្គ 2 (រសៀល 13:00-17:00)
                    {% elif attendance.shift == 3 %}វគ្គ 3 (យប់ 18:00-08:00)
                    {% else %}វគ្គ 1{% endif %}
                </p>
            </div>
            <form method="POST">
                <label>📅 កាលបរិច្ឆេទចូល</label>
                <input type="date" name="check_in_date" value="{{ attendance.date }}" required>

                <label>⏰ ម៉ោងចូល</label>
                <input type="time" name="check_in_time" value="{{ attendance.check_in.split(' ')[1][:5] if attendance.check_in else '' }}">

                <label>📅 កាលបរិច្ឆេទចេញ</label>
                <input type="date" name="check_out_date" value="{{ attendance.check_out.split(' ')[0] if attendance.check_out else attendance.date }}">

                <label>⏰ ម៉ោងចេញ</label>
                <input type="time" name="check_out_time" value="{{ attendance.check_out.split(' ')[1][:5] if attendance.check_out else '' }}">
                <div class="hint">💡 សម្រាប់វគ្គយប់ (វគ្គ 3) ប្រសិនបើចេញព្រឹកថ្ងៃបន្ទាប់ សូមកំណត់កាលបរិច្ឆេទចេញជាថ្ងៃបន្ទាប់</div>

                <label>វគ្គ</label>
                <select name="shift">
                    <option value="1" {% if attendance.shift == 1 %}selected{% endif %}>វគ្គ 1 (ព្រឹក 07:00-11:00)</option>
                    <option value="2" {% if attendance.shift == 2 %}selected{% endif %}>វគ្គ 2 (រសៀល 13:00-17:00)</option>
                    <option value="3" {% if attendance.shift == 3 %}selected{% endif %}>វគ្គ 3 (យប់ 18:00-08:00)</option>
                </select>

                <div class="btn-group">
                    <button type="submit" class="btn-save">💾 រក្សាទុក</button>
                    <a href="/dashboard" class="btn-cancel">❌ បោះបង់</a>
                </div>
            </form>
        </div>
    </body>
    </html>
    '''
    return render_template_string(edit_form_html, attendance=attendance)

@app.route('/delete_attendance/<int:attendance_id>', methods=['POST'])
def delete_attendance_route(attendance_id):
    if not session.get('logged_in') or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'អ្នកមិនមែនជា Admin!'})

    result = delete_attendance(attendance_id)
    if result:
        return jsonify({'success': True, 'message': '✅ លុបទិន្នន័យជោគជ័យ!'})
    else:
        return jsonify({'success': False, 'message': '❌ មិនអាចលុបទិន្នន័យបាន!'})

@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory(os.path.join('static', 'uploads'), filename)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = get_user_by_username(username)
        if user and user['password'] == password:
            session['logged_in'] = True
            session['username'] = username
            session['user_id'] = user['id']
            session['role'] = user['role']
            session['full_name'] = user['full_name']
            return redirect(url_for('dashboard'))
        else:
            return '''
            <!DOCTYPE html>
            <html>
            <head><title>កំហុស</title><meta charset="UTF-8">
            <style>
                body { font-family:'Khmer OS',Arial; text-align:center; padding:50px; background:#f0f2f5; }
                .box { background:white; padding:40px; border-radius:16px; max-width:400px; margin:0 auto; box-shadow:0 4px 20px rgba(0,0,0,0.08); }
                h3 { color:#dc3545; }
                .back-link { display:inline-block; margin-top:20px; color:#1a73e8; text-decoration:none; padding:10px 25px; border:2px solid #1a73e8; border-radius:10px; }
                .back-link:hover { background:#1a73e8; color:white; }
            </style>
            </head>
            <body>
                <div class="box">
                    <h3>❌ ឈ្មោះអ្នកប្រើ ឬ ពាក្យសម្ងាត់មិនត្រឹមត្រូវ!</h3>
                    <a href="/login" class="back-link">← ត្រលប់មកវិញ</a>
                </div>
            </body>
            </html>
            '''
    return '''
<!DOCTYPE html>
<html>
<head>
    <title>ចូលប្រើប្រព័ន្ធ</title>
    <meta charset="UTF-8">
    <link rel="manifest" href="/static/manifest.json">
    <meta name="theme-color" content="#1a73e8">
    <meta name="apple-mobile-web-app-capable" content="yes">
    <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
    <meta name="apple-mobile-web-app-title" content="HR System">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        html, body {
            height: 100%;
            width: 100%;
            font-family: 'Khmer OS', Arial, sans-serif;
            background: #ffffff;
            margin: 0;
            padding: 0;
            overflow: hidden;
        }
        .login-container {
            display: flex;
            justify-content: center;
            align-items: center;
            height: 100vh;
            width: 100vw;
            padding: 15px;
            background: #ffffff;
        }
        .login-box {
            width: 100%;
            height: 100%;
            max-width: 480px;
            max-height: 600px;
            background: #ffffff;
            padding: 40px 30px;
            border-radius: 24px;
            box-shadow: 0 10px 40px rgba(0, 0, 0, 0.06);
            border: 1px solid #f0f0f0;
            display: flex;
            flex-direction: column;
            justify-content: center;
            animation: fadeInUp 0.5s ease;
        }
        @keyframes fadeInUp {
            from { opacity: 0; transform: translateY(20px); }
            to { opacity: 1; transform: translateY(0); }
        }
        .logo-icon {
            text-align: center;
            font-size: 65px;
            margin-bottom: 8px;
        }
        .login-box h2 {
            text-align: center;
            color: #1a1a2e;
            font-size: 30px;
            font-weight: 700;
            margin-bottom: 4px;
        }
        .login-box .sub-title {
            text-align: center;
            color: #888;
            font-size: 15px;
            margin-bottom: 28px;
        }
        .login-box .form-group {
            margin-bottom: 16px;
        }
        .login-box input {
            width: 100%;
            padding: 16px 18px;
            border: 2px solid #e8ecf1;
            border-radius: 14px;
            font-size: 17px;
            box-sizing: border-box;
            font-family: 'Khmer OS', Arial, sans-serif;
            transition: all 0.3s;
            background: #f8f9fa;
        }
        .login-box input:focus {
            outline: none;
            border-color: #1a73e8;
            background: white;
            box-shadow: 0 0 0 4px rgba(26, 115, 232, 0.08);
        }
        .login-box button {
            width: 100%;
            padding: 16px;
            background: #1a73e8;
            color: white;
            border: none;
            border-radius: 14px;
            font-size: 18px;
            cursor: pointer;
            font-family: 'Khmer OS', Arial, sans-serif;
            font-weight: 600;
            transition: all 0.3s;
            margin-top: 4px;
        }
        .login-box button:hover {
            transform: translateY(-2px);
            box-shadow: 0 8px 25px rgba(26, 115, 232, 0.35);
        }
        .login-box button:active { transform: scale(0.98); }
        .login-box .hint {
            text-align: center;
            margin-top: 16px;
            color: #999;
            font-size: 13px;
        }
        .login-box .hint b { color: #1a73e8; }
        .login-box .footer-text {
            text-align: center;
            margin-top: 20px;
            padding-top: 16px;
            border-top: 1px solid #eee;
            color: #ccc;
            font-size: 12px;
        }
        @media (max-width: 600px) {
            .login-box {
                padding: 30px 22px;
                border-radius: 18px;
                max-height: none;
            }
            .login-box h2 {
                font-size: 26px;
            }
            .login-box input {
                padding: 15px 16px;
                font-size: 16px;
            }
            .login-box button {
                padding: 15px;
                font-size: 17px;
            }
            .logo-icon {
                font-size: 55px;
            }
        }
        @media (max-width: 400px) {
            .login-container {
                padding: 10px;
            }
            .login-box {
                padding: 22px 16px;
                border-radius: 14px;
            }
            .login-box h2 {
                font-size: 22px;
            }
            .login-box input {
                padding: 13px 14px;
                font-size: 15px;
            }
            .login-box button {
                padding: 13px;
                font-size: 16px;
            }
            .logo-icon {
                font-size: 45px;
            }
            .login-box .sub-title {
                font-size: 13px;
                margin-bottom: 20px;
            }
        }
        @media (max-height: 600px) {
            .login-box {
                padding: 20px 20px;
            }
            .logo-icon {
                font-size: 40px;
                margin-bottom: 4px;
            }
            .login-box h2 {
                font-size: 22px;
                margin-bottom: 2px;
            }
            .login-box .sub-title {
                font-size: 13px;
                margin-bottom: 16px;
            }
            .login-box .form-group {
                margin-bottom: 10px;
            }
            .login-box input {
                padding: 12px 14px;
                font-size: 15px;
            }
            .login-box button {
                padding: 12px;
                font-size: 16px;
            }
            .login-box .hint {
                margin-top: 10px;
                font-size: 12px;
            }
            .login-box .footer-text {
                margin-top: 12px;
                padding-top: 10px;
                font-size: 11px;
            }
        }
    </style>
</head>
<body>
    <div class="login-container">
        <div class="login-box">
            <div class="logo-icon">🏢</div>
            <h2 style="font-family: 'Khmer', 'Khmer OS Muol Light', 'Khmer OS', Arial, sans-serif; font-weight: 300; color: #1a73e8;">ប្រព័ន្ធគ្រប់គ្រងបុគ្គលិក</h2>
            <div class="sub-title">សូមបញ្ចូលឈ្មោះ និងពាក្យសម្ងាត់</div>
            <form method="POST" style="flex:1;display:flex;flex-direction:column;justify-content:center;">
                <div class="form-group">
                    <input type="text" name="username" placeholder="ឈ្មោះអ្នកប្រើ" required>
                </div>
                <div class="form-group">
                    <input type="password" name="password" placeholder="ពាក្យសម្ងាត់" required>
                </div>
                <button type="submit">🔐 Login</button>
            </form>
            <div class="hint"><b>ពត៌មានបន្ថែមៈ ទំនាក់ទំនងលោក YEN SONY</b></div>
            <div class="hint"><b>ទូរស័ព្ទ៖ +855 92 740 067</b></div>
            <div class="footer-text">© 2026 ប្រព័ន្ធគ្រប់គ្រងបុគ្គលិក</div>
        </div>
    </div>
    <script>
        let deferredPrompt;
        const installBtn = document.createElement('button');
        installBtn.id = 'pwaInstallBtn';
        installBtn.style.cssText = `
            position: fixed;
            bottom: 20px;
            left: 50%;
            transform: translateX(-50%);
            background: #34a853;
            color: white;
            border: none;
            border-radius: 50px;
            padding: 14px 30px;
            font-size: 16px;
            font-family: 'Khmer OS', Arial, sans-serif;
            font-weight: 600;
            cursor: pointer;
            box-shadow: 0 4px 20px rgba(52, 168, 83, 0.4);
            z-index: 9999;
            display: none;
            animation: slideUp 0.5s ease;
        `;
        installBtn.innerHTML = '📲 ដំឡើងកម្មវិធី';
        document.body.appendChild(installBtn);

        const stylePwa = document.createElement('style');
        stylePwa.textContent = `
            @keyframes slideUp {
                from { transform: translateX(-50%) translateY(100px); opacity: 0; }
                to { transform: translateX(-50%) translateY(0); opacity: 1; }
            }
        `;
        document.head.appendChild(stylePwa);

        window.addEventListener('beforeinstallprompt', function(e) {
            e.preventDefault();
            deferredPrompt = e;
            installBtn.style.display = 'block';
        });

        installBtn.addEventListener('click', function() {
            if (deferredPrompt) {
                deferredPrompt.prompt();
                deferredPrompt.userChoice.then(function(choiceResult) {
                    if (choiceResult.outcome === 'accepted') {
                        console.log('User accepted the install prompt');
                        installBtn.style.display = 'none';
                    } else {
                        console.log('User dismissed the install prompt');
                    }
                    deferredPrompt = null;
                });
            }
        });

        window.addEventListener('appinstalled', function() {
            console.log('App installed successfully!');
            installBtn.style.display = 'none';
        });

        if ('serviceWorker' in navigator) {
            navigator.serviceWorker.register('/static/sw.js')
            .then(function(reg) {
                console.log('Service Worker registered successfully!');
            })
            .catch(function(err) {
                console.log('Service Worker registration failed:', err);
            });
        }

        if (window.matchMedia('(display-mode: standalone)').matches) {
            installBtn.style.display = 'none';
        }
    </script>
</body>
</html>
'''

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ============================================================
# DASHBOARD_HTML - កែសម្រួលផ្នែក JavaScript សម្រាប់ toggleSystemLock
# ============================================================

DASHBOARD_HTML = r'''
<!DOCTYPE html>
<html>
<head>
    <title>ប្រព័ន្ធគ្រប់គ្រងបុគ្គលិក</title>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
    <link rel="manifest" href="/static/manifest.json">
    <meta name="theme-color" content="#1a73e8">
    <meta name="apple-mobile-web-app-capable" content="yes">
    <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
    <meta name="apple-mobile-web-app-title" content="HR System">
    <link href="https://fonts.googleapis.com/css2?family=Khmer&display=swap" rel="stylesheet">
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif; background: #f0f2f5; padding: 15px; min-height: 100vh; }
        .container { max-width: 1200px; margin: 0 auto; }
        .header {
            background: linear-gradient(135deg, #1a73e8, #0d47a1);
            color: white;
            padding: 12px 20px;
            border-radius: 12px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            box-shadow: 0 4px 15px rgba(26, 115, 232, 0.3);
            margin-bottom: 15px;
            flex-wrap: wrap;
            gap: 10px;
        }
        .header-left h2 { font-size: 18px; font-weight: 600; white-space: nowrap; }
        .header-right {
            display: flex;
            align-items: center;
            gap: 8px;
            flex-wrap: wrap;
        }
        .header-right .nav-btn {
            color: white;
            text-decoration: none;
            padding: 6px 16px;
            border-radius: 20px;
            font-size: 13px;
            font-weight: 500;
            transition: all 0.3s;
            background: rgba(255,255,255,0.1);
            border: 1px solid rgba(255,255,255,0.15);
            white-space: nowrap;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
            cursor: pointer;
        }
        .header-right .nav-btn:hover { background: rgba(255,255,255,0.25); transform: translateY(-2px); }
        .header-right .nav-btn.location-btn {
            background: rgba(255, 193, 7, 0.3);
            border-color: #ffc107;
        }
        .header-right .nav-btn.location-btn:hover {
            background: rgba(255, 193, 7, 0.5);
        }
        .header-right .user-name-btn {
            color: white;
            padding: 6px 16px;
            border-radius: 20px;
            font-size: 13px;
            font-weight: 500;
            background: rgba(255,255,255,0.2);
            border: 1px solid rgba(255,255,255,0.25);
            white-space: nowrap;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
            cursor: default;
            display: flex;
            align-items: center;
            gap: 6px;
        }
        .header-right .logout-link {
            color: white;
            text-decoration: none;
            background: rgba(255,255,255,0.2);
            padding: 6px 16px;
            border-radius: 20px;
            font-size: 13px;
            transition: all 0.3s;
            font-weight: 500;
            border: 1px solid rgba(255,255,255,0.15);
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
        }
        .header-right .logout-link:hover { background: rgba(255,255,255,0.35); transform: scale(1.05); }
        .stats {
            display: grid;
            grid-template-columns: repeat(4, 1fr);
            gap: 15px;
            margin-bottom: 20px;
        }
        .stat-card {
            background: white;
            padding: 18px 15px;
            border-radius: 12px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.06);
            text-align: center;
            transition: all 0.3s;
        }
        .stat-card:hover { transform: translateY(-2px); box-shadow: 0 6px 20px rgba(0,0,0,0.1); }
        .stat-card .stat-icon { font-size: 28px; display: block; margin-bottom: 4px; }
        .stat-card .number { font-size: 30px; font-weight: 700; color: #1a73e8; }
        .stat-card .number.green { color: #34a853; }
        .stat-card .number.orange { color: #fbbc04; }
        .stat-card .number.purple { color: #7c3aed; }
        .stat-card .label { font-size: 13px; color: #888; margin-top: 2px; }
        .action-buttons {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 12px;
            margin-bottom: 20px;
        }
        .action-buttons .btn {
            padding: 14px 10px;
            border: none;
            border-radius: 12px;
            cursor: pointer;
            font-size: 16px;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
            transition: all 0.3s;
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            gap: 4px;
            font-weight: 500;
            min-height: 65px;
            color: white;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        }
        .action-buttons .btn .icon { font-size: 28px; }
        .action-buttons .btn .sub-text { font-size: 12px; opacity: 0.85; font-weight: 400; }
        .action-buttons .btn:hover { transform: translateY(-3px); box-shadow: 0 6px 25px rgba(0,0,0,0.15); }
        .action-buttons .btn:active { transform: scale(0.95); }
        .action-buttons .btn.disabled {
            opacity: 0.5;
            cursor: not-allowed;
            transform: none !important;
        }
        .btn-checkin { background: linear-gradient(135deg, #34a853, #1e7e34); }
        .btn-checkout { background: linear-gradient(135deg, #dc3545, #b02a37); }
        .btn-leave { background: linear-gradient(135deg, #fbbc04, #e5a800); color: #333; }
        .btn-mission { background: linear-gradient(135deg, #1a73e8, #0d47a1); }
        .table-container {
            background: white;
            border-radius: 12px;
            padding: 20px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.06);
            overflow-x: auto;
        }
        .table-container .table-title {
            font-size: 18px;
            font-weight: 600;
            color: #333;
            margin-bottom: 15px;
            display: flex;
            align-items: center;
            gap: 10px;
        }
        .table-container .table-title .badge-count {
            background: #1a73e8;
            color: white;
            padding: 2px 12px;
            border-radius: 20px;
            font-size: 14px;
        }
        table { width: 100%; border-collapse: collapse; }
        table thead th {
            background: #f8f9fa;
            color: #555;
            padding: 12px 15px;
            text-align: left;
            font-size: 13px;
            font-weight: 600;
            border-bottom: 2px solid #e8ecf1;
        }
        table tbody td {
            padding: 12px 15px;
            border-bottom: 1px solid #f0f2f5;
            color: #333;
            font-size: 14px;
        }
        table tbody tr:hover { background: #f8f9fa; }
        .status-badge {
            display: inline-block;
            padding: 4px 14px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: 600;
        }
        .status-badge.present { background: #e8f5e9; color: #2e7d32; }
        .status-badge.outside { background: #ffebee; color: #c62828; }
        .status-badge.working { background: #fff3cd; color: #856404; }
        .status-badge.leave { background: #fff3cd; color: #856404; }
        .status-badge.mission { background: #d1ecf1; color: #0c5460; }
        .status-badge.approved { background: #d4edda; color: #155724; }
        .shift-badge {
            display: inline-block;
            padding: 2px 10px;
            border-radius: 12px;
            font-size: 11px;
            font-weight: 600;
        }
        .shift-1 { background: #e3f2fd; color: #0d47a1; }
        .shift-2 { background: #fff3e0; color: #e65100; }
        .shift-3 { background: #f3e5f5; color: #4a148c; }
        .shift-leave { background: #fff3cd; color: #856404; }
        .shift-mission { background: #d1ecf1; color: #0c5460; }
        .footer { text-align: center; padding: 20px 0 5px; color: #aaa; font-size: 12px; }
        .modal {
            display: none;
            position: fixed;
            top: 0; left: 0; width: 100%; height: 100%;
            background: rgba(0,0,0,0.5);
            justify-content: center;
            align-items: center;
            z-index: 999;
        }
        .modal.show { display: flex; }
        .modal-content {
            background: white;
            padding: 30px;
            border-radius: 16px;
            max-width: 500px;
            width: 90%;
            box-shadow: 0 10px 40px rgba(0,0,0,0.3);
            max-height: 90vh;
            overflow-y: auto;
        }
        .modal-content h3 { color: #333; margin-bottom: 10px; }
        .modal-content p { color: #666; font-size: 14px; margin-bottom: 15px; }
        .modal-content input, .modal-content textarea, .modal-content select {
            width: 100%;
            padding: 10px 14px;
            border: 2px solid #e8ecf1;
            border-radius: 10px;
            font-size: 14px;
            margin-bottom: 10px;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
        }
        .modal-content input:focus, .modal-content textarea:focus, .modal-content select:focus {
            outline: none;
            border-color: #1a73e8;
        }
        .modal-content .btn-group {
            display: flex;
            gap: 10px;
            flex-wrap: wrap;
        }
        .modal-content .btn-group button {
            flex: 1;
            padding: 12px;
            border: none;
            border-radius: 10px;
            cursor: pointer;
            font-size: 15px;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
            font-weight: 500;
            min-width: 80px;
        }
        .btn-save { background: #1a73e8; color: white; }
        .btn-save:hover { background: #1557b0; }
        .btn-success { background: #34a853; color: white; }
        .btn-success:hover { background: #1e7e34; }
        .btn-danger { background: #dc3545; color: white; }
        .btn-danger:hover { background: #b02a37; }
        .btn-cancel { background: #e8ecf1; color: #333; }
        .btn-cancel:hover { background: #d5d8dd; }
        .btn-get-location {
            background: #34a853;
            color: white;
            padding: 10px;
            border: none;
            border-radius: 10px;
            cursor: pointer;
            font-size: 14px;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
            width: 100%;
            margin-bottom: 10px;
        }
        .btn-get-location:hover { background: #1e7e34; }
        .location-info {
            background: #f8f9fa;
            padding: 10px;
            border-radius: 8px;
            font-size: 13px;
            color: #555;
            margin-bottom: 10px;
        }
        .request-item {
            background: #f8f9fa;
            padding: 12px 15px;
            border-radius: 10px;
            margin-bottom: 10px;
            border-left: 4px solid #ffc107;
        }
        .request-item .request-user { font-weight: 600; color: #333; }
        .request-item .request-detail { font-size: 13px; color: #666; margin-top: 3px; }
        .request-item .request-actions { margin-top: 8px; display: flex; gap: 8px; }
        .request-item .request-actions button {
            padding: 5px 15px;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            font-size: 13px;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
        }
        .request-item .request-actions .approve-btn {
            background: #34a853;
            color: white;
        }
        .request-item .request-actions .approve-btn:hover {
            background: #1e7e34;
        }
        .request-item .request-actions .reject-btn {
            background: #dc3545;
            color: white;
        }
        .request-item .request-actions .reject-btn:hover {
            background: #b02a37;
        }
        .status-message {
            padding: 10px 15px;
            border-radius: 8px;
            margin-bottom: 10px;
            font-size: 14px;
            display: none;
        }
        .status-message.info {
            background: #d1ecf1;
            color: #0c5460;
            border: 1px solid #bee5eb;
            display: block;
        }
        .status-message.success {
            background: #d4edda;
            color: #155724;
            border: 1px solid #c3e6cb;
            display: block;
        }
        .status-message.warning {
            background: #fff3cd;
            color: #856404;
            border: 1px solid #ffc107;
            display: block;
        }
        .status-message.danger {
            background: #f8d7da;
            color: #721c24;
            border: 1px solid #f5c6cb;
            display: block;
        }
        .status-message.locked {
            background: #f8d7da;
            color: #721c24;
            border: 2px solid #dc3545;
            display: block;
            font-weight: 600;
        }
        .install-btn {
            background: #34a853;
            color: white;
            border: none;
            border-radius: 20px;
            padding: 6px 16px;
            cursor: pointer;
            font-size: 13px;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
            font-weight: 500;
            display: none;
            animation: pulse-green 2s infinite;
        }
        @keyframes pulse-green {
            0% { box-shadow: 0 0 0 0 rgba(52, 168, 83, 0.4); }
            70% { box-shadow: 0 0 0 10px rgba(52, 168, 83, 0); }
            100% { box-shadow: 0 0 0 0 rgba(52, 168, 83, 0); }
        }
        .install-btn:hover { background: #1e7e34; transform: scale(1.05); }
        .summary-grid {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 15px;
            margin-top: 15px;
        }
        .summary-card {
            background: white;
            border-radius: 12px;
            padding: 15px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.06);
        }
        .summary-card .card-title {
            font-size: 15px;
            font-weight: 600;
            color: #555;
            margin-bottom: 10px;
            padding-bottom: 8px;
            border-bottom: 2px solid #f0f2f5;
        }
        .summary-item {
            display: flex;
            justify-content: space-between;
            padding: 6px 0;
            border-bottom: 1px solid #f5f6f8;
            font-size: 14px;
        }
        .summary-item .name { color: #333; }
        .summary-item .value { font-weight: 600; color: #1a73e8; }
        .summary-item .value.leave { color: #fbbc04; }
        .summary-item .value.mission { color: #7c3aed; }
        .summary-item .value.hours { color: #34a853; }
        .deadline-badge {
            display: inline-block;
            padding: 2px 10px;
            border-radius: 12px;
            font-size: 11px;
            font-weight: 600;
            background: #ff9800;
            color: white;
        }
        .deadline-badge.active {
            background: #4caf50;
        }
        .deadline-badge.inactive {
            background: #9e9e9e;
        }
        .user-lock-badge {
            display: inline-block;
            padding: 2px 10px;
            border-radius: 12px;
            font-size: 11px;
            font-weight: 600;
            background: #dc3545;
            color: white;
        }
        .user-lock-badge.unlocked {
            background: #4caf50;
        }
        .lock-banner {
            background: linear-gradient(135deg, #dc3545, #b02a37);
            color: white;
            padding: 12px 20px;
            border-radius: 12px;
            display: none;
            align-items: center;
            justify-content: space-between;
            margin-bottom: 15px;
            flex-wrap: wrap;
            gap: 10px;
        }
        .lock-banner.show { display: flex; }
        .lock-banner .lock-icon { font-size: 24px; margin-right: 10px; }
        .lock-banner .lock-text {
            font-size: 16px;
            font-weight: 500;
            flex: 1;
        }
        .lock-banner .lock-time {
            font-size: 13px;
            opacity: 0.8;
        }
        .lock-banner .lock-btn {
            background: rgba(255,255,255,0.2);
            border: 1px solid rgba(255,255,255,0.3);
            color: white;
            padding: 6px 16px;
            border-radius: 20px;
            cursor: pointer;
            font-size: 13px;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
            transition: all 0.3s;
        }
        .lock-banner .lock-btn:hover {
            background: rgba(255,255,255,0.35);
        }
        .lock-banner.unlocked {
            background: linear-gradient(135deg, #34a853, #1e7e34);
        }
        .user-lock-banner {
            background: linear-gradient(135deg, #ff9800, #e65100);
            color: white;
            padding: 10px 16px;
            border-radius: 12px;
            display: none;
            align-items: center;
            justify-content: space-between;
            margin-bottom: 15px;
            flex-wrap: wrap;
            gap: 8px;
        }
        .user-lock-banner.show { display: flex; }
        .user-lock-banner .lock-icon { font-size: 20px; margin-right: 8px; }
        .user-lock-banner .lock-text {
            font-size: 14px;
            font-weight: 500;
            flex: 1;
        }
        .user-lock-banner .lock-time {
            font-size: 12px;
            opacity: 0.8;
        }
        @media (max-width: 900px) {
            .header { flex-direction: column; align-items: center; gap: 12px; padding: 15px; }
            .header-left h2 { font-size: 16px; text-align: center; }
            .header-right { justify-content: center; }
            .header-right .nav-btn { font-size: 12px; padding: 5px 12px; }
            .header-right .user-name-btn { font-size: 12px; padding: 5px 12px; }
            .header-right .logout-link { font-size: 12px; padding: 5px 12px; }
        }
        @media (max-width: 768px) {
            .stats { grid-template-columns: repeat(2, 1fr); gap: 10px; }
            .stat-card .number { font-size: 24px; }
            .action-buttons { gap: 10px; }
            .action-buttons .btn { padding: 12px 8px; font-size: 14px; min-height: 55px; }
            .action-buttons .btn .icon { font-size: 22px; }
            .summary-grid { grid-template-columns: 1fr; }
            .lock-banner { flex-direction: column; text-align: center; }
            .lock-banner .lock-text { font-size: 14px; }
            .user-lock-banner { flex-direction: column; text-align: center; }
            .user-lock-banner .lock-text { font-size: 13px; }
        }
        @media (max-width: 600px) {
            .header-right .nav-btn { font-size: 11px; padding: 4px 10px; }
            .header-right .user-name-btn { font-size: 11px; padding: 4px 10px; }
            .header-right .logout-link { font-size: 11px; padding: 4px 10px; }
            .action-buttons { gap: 8px; }
            .action-buttons .btn { padding: 10px 6px; font-size: 13px; min-height: 50px; }
            .action-buttons .btn .icon { font-size: 20px; }
            .action-buttons .btn .sub-text { font-size: 10px; }
        }
        @media (max-width: 400px) {
            .stats { grid-template-columns: 1fr 1fr; gap: 8px; }
            .stat-card .number { font-size: 20px; }
            .action-buttons { gap: 6px; }
            .action-buttons .btn { padding: 8px 5px; font-size: 12px; min-height: 45px; }
            .action-buttons .btn .icon { font-size: 18px; }
            .action-buttons .btn .sub-text { font-size: 9px; }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <div class="header-left">
                <h2>🏢 ប្រព័ន្ធគ្រប់គ្រងបុគ្គលិក</h2>
            </div>
            <div class="header-right">
                <button id="installBtn" class="install-btn">📲 ដំឡើងកម្មវិធី</button>
                <a href="/report" class="nav-btn">📊 របាយការណ៍</a>
                {% if session.role == 'admin' %}
                <a href="/register" class="nav-btn">📝 ចុះឈ្មោះ</a>
                <a href="/manage_users" class="nav-btn">👤 គ្រប់គ្រងអ្នកប្រើ</a>
                <button class="nav-btn location-btn" onclick="openLocationModal()">📍 កំណត់ទីតាំង</button>
                <button class="nav-btn" onclick="openCleanModal()" style="background:rgba(220,53,69,0.2);border-color:#dc3545;">🧹 សម្អាតទិន្នន័យ</button>
                <button class="nav-btn" onclick="openRequestsModal()" style="background:rgba(255,193,7,0.2);border-color:#ffc107;">
                    📬 សំណើ
                    {% if pending_count > 0 %}
                    <span style="background:red;color:white;border-radius:50%;padding:0 6px;font-size:11px;margin-left:3px;">{{ pending_count }}</span>
                    {% endif %}
                </button>
                <button class="nav-btn" onclick="openSystemLockModal()" style="background:rgba(33,150,243,0.2);border-color:#2196f3;">
                    🔒 បិទ/បើកប្រព័ន្ធ
                </button>
                {% endif %}
                <a href="/change_password" class="nav-btn">🔑 ប្តូរពាក្យសម្ងាត់</a>
                <span class="user-name-btn">👤 {{ session.username }}</span>
                <a href="/logout" class="logout-link">🚪 ចាកចេញ</a>
            </div>
        </div>

        <!-- User Lock Banner -->
        <div id="userLockBanner" class="user-lock-banner {% if user_lock.is_locked == 1 %}show{% endif %}">
            <div style="display:flex;align-items:center;gap:8px;">
                <span class="lock-icon">🔒</span>
                <span class="lock-text" id="userLockText">
                    {% if user_lock.is_locked == 1 %}
                    អ្នកត្រូវបានបិទការចូលធ្វើការដោយ Admin
                    {% if user_lock.auto_unlock_time %}
                    (បើកដោយស្វ័យប្រវត្តិនៅម៉ោង {{ user_lock.auto_unlock_time }})
                    {% endif %}
                    {% endif %}
                </span>
            </div>
        </div>

        <!-- System Lock Banner -->
        <div id="lockBanner" class="lock-banner show">
            <div style="display:flex;align-items:center;gap:10px;">
                <span class="lock-icon">🔒</span>
                <span class="lock-text" id="lockText">ប្រព័ន្ធកំពុងបិទការចូលធ្វើការ</span>
                <span class="lock-time" id="lockTime">បើកដោយស្វ័យប្រវត្តិនៅម៉ោង 06:00</span>
            </div>
            {% if session.role == 'admin' %}
            <button class="lock-btn" id="lockToggleBtn" onclick="openSystemLockModal()">🔒 បិទ/បើកប្រព័ន្ធ</button>
            {% endif %}
        </div>

        <div id="statusMessage" class="status-message"></div>

        <div class="stats">
            <div class="stat-card">
                <span class="stat-icon">👥</span>
                <div class="number">{{ stats.total_users }}</div>
                <div class="label">បុគ្គលិកសរុប</div>
            </div>
            <div class="stat-card">
                <span class="stat-icon">✅</span>
                <div class="number green">{{ stats.present_today }}</div>
                <div class="label">កំពុងធ្វើការ</div>
            </div>
            <div class="stat-card">
                <span class="stat-icon">📋</span>
                <div class="number orange">{{ stats.leave_today }}</div>
                <div class="label">សុំច្បាប់ (Approved)</div>
            </div>
            <div class="stat-card">
                <span class="stat-icon">🚗</span>
                <div class="number purple">{{ stats.mission_today }}</div>
                <div class="label">ចុះបេសកម្ម (Approved)</div>
            </div>
        </div>

        <div class="action-buttons" id="actionButtons">
            <button class="btn btn-checkin" onclick="openCheckInModal()" id="checkinBtn">
                <span class="icon">✅</span>
                ចូលធ្វើការ
                <span class="sub-text">ជ្រើសរើសវគ្គ</span>
            </button>
            <button class="btn btn-checkout" onclick="handleCheckOut()" id="checkoutBtn">
                <span class="icon">🚪</span>
                ចេញពីធ្វើការ
                <span class="sub-text">ចុចដើម្បីចេញ</span>
            </button>
            <button class="btn btn-leave" onclick="handleLeave()">
                <span class="icon">📋</span>
                សុំច្បាប់
                <span class="sub-text">ស្នើសុំច្បាប់</span>
            </button>
            <button class="btn btn-mission" onclick="handleMission()">
                <span class="icon">🚗</span>
                បេសកម្ម
                <span class="sub-text">ស្នើសុំបេសកម្ម</span>
            </button>
        </div>

        <div class="table-container">
            <div class="table-title">
                📋 ប្រវត្តិការងារបុគ្គលិក
                <span class="badge-count">{{ work_history|length }}</span>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>ល.រ</th>
                        <th>ឈ្មោះ</th>
                        <th>កាលបរិច្ឆេទ</th>
                        <th>វគ្គ</th>
                        <th>ម៉ោងចូល</th>
                        <th>ម៉ោងចេញ</th>
                        <th>ម៉ោង/ថ្ងៃ</th>
                        <th>ស្ថានភាព</th>
                        <th>ព័ត៌មាន</th>
                        {% if session.role == 'admin' %}
                        <th>សកម្មភាព</th>
                        {% endif %}
                    </tr>
                </thead>
                <tbody>
    {% if work_history %}
        {% for item in work_history %}
        <tr>
            <td>{{ loop.index }}</td>
            <td>{{ item.full_name }}</td>
            <td>{{ item.date }}</td>
            <td>
                {% if item.type == 'attendance' %}
                    {% if item.shift == 'វគ្គ 1' %}
                    <span class="shift-badge shift-1">វគ្គ 1</span>
                    {% elif item.shift == 'វគ្គ 2' %}
                    <span class="shift-badge shift-2">វគ្គ 2</span>
                    {% elif item.shift == 'វគ្គ 3' %}
                    <span class="shift-badge shift-3">វគ្គ 3</span>
                    {% else %}
                    <span class="shift-badge shift-1">{{ item.shift }}</span>
                    {% endif %}
                {% elif item.type == 'leave' %}
                    <span class="shift-badge shift-leave">📋 ច្បាប់</span>
                {% elif item.type == 'mission' %}
                    <span class="shift-badge shift-mission">🚗 បេសកម្ម</span>
                {% endif %}
            </td>
            <td>{{ item.check_in_time or '' }}</td>
            <td>{{ item.check_out_time or '' }}</td>
            <td>{{ item.display_value or '' }}</td>
            <td>
                {% if item.type == 'attendance' %}
                    {% if item.status == 'បានបិទ' %}
                    <span class="status-badge present">បានបិទ</span>
                    {% elif item.status == 'កំពុងធ្វើការ' %}
                    <span class="status-badge working">កំពុងធ្វើការ</span>
                    {% else %}
                    <span class="status-badge outside">មិនទាន់ចូល</span>
                    {% endif %}
                {% else %}
                    <span class="status-badge approved">បានអនុម័ត</span>
                {% endif %}
            </td>
            <td>{{ item.display_info or '' }}</td>
            {% if session.role == 'admin' %}
            <td>
                {% if item.type == 'attendance' %}
                <a href="/edit_attendance/{{ item.id }}"
                   style="background:#1a73e8;color:white;padding:4px 12px;border-radius:6px;text-decoration:none;font-size:12px;display:inline-block;margin:2px;">
                   ✏️ កែ
                </a>
                <button onclick="deleteAttendance({{ item.id }})"
                        style="background:#dc3545;color:white;padding:4px 12px;border:none;border-radius:6px;cursor:pointer;font-size:12px;margin:2px;">
                   🗑️ លុប
                </button>
                {% else %}
                <span style="font-size:11px;color:#999;">មិនអាចកែ</span>
                {% endif %}
            </td>
            {% endif %}
        </tr>
        {% endfor %}
    {% else %}
        <tr>
            <td colspan="{% if session.role == 'admin' %}10{% else %}9{% endif %}"
                style="text-align:center;padding:30px;color:#aaa;">
                📭 មិនមានទិន្នន័យសម្រាប់ខែនេះ
            </td>
        </tr>
    {% endif %}
    </tbody>
            </table>
        </div>

        <div class="summary-grid">
            <div class="summary-card">
                <div class="card-title">📊 សង្ខេបបុគ្គលិក</div>
                <div class="summary-item" style="font-weight:600;color:#555;border-bottom:2px solid #e8ecf1;">
                    <span>ឈ្មោះ</span>
                    <span>ថ្ងៃធ្វើការ</span>
                </div>
                {% for emp in monthly_summary %}
                <div class="summary-item">
                    <span class="name">{{ emp.full_name }}</span>
                    <span class="value">{{ emp.days_worked_display }}</span>
                </div>
                {% endfor %}
                {% if not monthly_summary %}
                <div style="text-align:center;padding:15px;color:#aaa;">មិនមានទិន្នន័យ</div>
                {% endif %}
            </div>
            <div class="summary-card">
                <div class="card-title">📋 សង្ខេបវត្តមាន និងច្បាប់</div>
                <div class="summary-item" style="font-weight:600;color:#555;border-bottom:2px solid #e8ecf1;">
                    <span>ឈ្មោះ</span>
                    <span style="display:flex;gap:15px;">
                        <span>ម៉ោង</span>
                        <span>ច្បាប់</span>
                        <span>បេសកម្ម</span>
                    </span>
                </div>
                {% for emp in monthly_summary %}
                <div class="summary-item">
                    <span class="name">{{ emp.full_name }}</span>
                    <span style="display:flex;gap:50px;font-weight:600;">
                        <span class="value hours">{{ emp.total_hours_formatted }}</span>
                        <span class="value leave">{{ emp.total_leave_days }}</span>
                        <span class="value mission">{{ emp.total_mission_days }}</span>
                    </span>
                </div>
                {% endfor %}
                {% if not monthly_summary %}
                <div style="text-align:center;padding:15px;color:#aaa;">មិនមានទិន្នន័យ</div>
                {% endif %}
            </div>
        </div>

        <div class="footer">
            © 2026 ប្រព័ន្ធគ្រប់គ្រងបុគ្គលិក | រក្សាសិទ្ធិគ្រប់យ៉ាង
        </div>
    </div>

    <div id="checkInModal" class="modal">
        <div class="modal-content">
            <h3>✅ ចូលធ្វើការ</h3>
            <p>សូមជ្រើសរើសវគ្គដែលអ្នកចង់ចូល</p>
            <div id="checkInLocationInfo" class="location-info">
                ⏳ កំពុងចាប់យកទីតាំង...
            </div>
            <select id="shiftSelect">
                <option value="1">វគ្គ 1 (ព្រឹក 07:00-11:00)</option>
                <option value="2">វគ្គ 2 (រសៀល 13:00-17:00)</option>
                <option value="3">វគ្គ 3 (យប់ 18:00-08:00)</option>
            </select>
            <div class="btn-group">
                <button class="btn-success" onclick="submitCheckIn()">✅ ចូលធ្វើការ</button>
                <button class="btn-cancel" onclick="closeModal('checkInModal')">បោះបង់</button>
            </div>
        </div>
    </div>

    <div id="locationModal" class="modal">
        <div class="modal-content">
            <h3>📍 កំណត់ទីតាំងក្រុមហ៊ុន</h3>
            <p>សូមកំណត់ទីតាំងក្រុមហ៊ុនសម្រាប់ពិនិត្យមើលវត្តមានបុគ្គលិក (ចម្ងាយអនុញ្ញាត {{ allowed_distance }} ម៉ែត្រ)</p>
            <div id="locationInfo" class="location-info">
                {% if company_lat and company_lng %}
                📍 ទីតាំងបច្ចុប្បន្ន: {{ company_lat }}, {{ company_lng }}
                {% else %}
                ⚠️ មិនទាន់មានកំណត់ទីតាំង
                {% endif %}
            </div>
            <button class="btn-get-location" onclick="getCurrentLocation()">📡 ចាប់យកទីតាំងបច្ចុប្បន្ន</button>
            <input type="text" id="companyLat" placeholder="រយៈទទឹង (Latitude)" value="{{ company_lat or '' }}">
            <input type="text" id="companyLng" placeholder="រយៈបណ្តោយ (Longitude)" value="{{ company_lng or '' }}">
            <div class="btn-group">
                <button class="btn-save" onclick="saveCompanyLocation()">💾 រក្សាទុក</button>
                <button class="btn-cancel" onclick="closeLocationModal()">បោះបង់</button>
            </div>
        </div>
    </div>

    <!-- System Lock Modal - កែសម្រួលសម្រាប់បើកជាមួយម៉ោងកំណត់ -->
    <div id="systemLockModal" class="modal">
        <div class="modal-content">
            <h3>🔒 បិទ/បើកប្រព័ន្ធចូលធ្វើការ</h3>
            <p style="color:#dc3545;font-weight:600;">⚠️ ពេលបិទប្រព័ន្ធ បុគ្គលិកនឹងមិនអាចចុចចូល/ចេញធ្វើការបានទេ!</p>
            <div id="lockStatusInfo" class="location-info">
                <span id="lockStatusText">ប្រព័ន្ធកំពុងបើក</span>
            </div>
            <div style="margin-bottom:15px;">
                <label style="font-weight:600;color:#555;font-size:14px;">ម៉ោងបើកដោយស្វ័យប្រវត្តិ (HH:MM)</label>
                <input type="time" id="autoUnlockTime" value="06:00" step="60">
                <div style="font-size:12px;color:#888;margin-top:4px;">
                    🔓 ពេលដល់ម៉ោងនេះ ប្រព័ន្ធនឹងបើកដោយស្វ័យប្រវត្តិ
                </div>
            </div>
            <div class="btn-group">
                <button id="lockToggleBtnModal" class="btn-danger" onclick="toggleSystemLockFromModal(1)">🔒 បិទប្រព័ន្ធឥឡូវនេះ</button>
                <button class="btn-success" onclick="toggleSystemLockFromModal(0)" id="unlockBtnModal" style="display:none;">🔓 បើកប្រព័ន្ធឥឡូវនេះ</button>
                <button class="btn-cancel" onclick="closeModal('systemLockModal')">បោះបង់</button>
            </div>
            <div id="unlockScheduledInfo" style="display:none;margin-top:12px;padding:10px;background:#fff3cd;border-radius:8px;color:#856404;font-size:14px;">
                ⏰ ប្រព័ន្ធនឹងបើកដោយស្វ័យប្រវត្តិនៅម៉ោង <span id="scheduledUnlockTime">06:00</span>
            </div>
        </div>
    </div>

    <div id="requestsModal" class="modal">
        <div class="modal-content">
            <h3>📬 សំណើរង់ចាំការអនុម័ត</h3>
            <div id="requestsList">
                <p style="text-align:center;color:#888;padding:20px;">កំពុងផ្ទុក...</p>
            </div>
            <div class="btn-group">
                <button class="btn-cancel" onclick="closeRequestsModal()" style="width:100%;">បិទ</button>
            </div>
        </div>
    </div>

    <div id="cleanModal" class="modal">
        <div class="modal-content" style="max-width:500px;">
            <h3>🧹 សម្អាតទិន្នន័យ</h3>
            <p style="color:#dc3545;font-weight:600;margin-bottom:15px;">⚠️ ការសម្អាតទិន្នន័យមិនអាចស្តារឡើងវិញបានទេ!</p>
            <p style="margin-bottom:15px;color:#666;">សូមជ្រើសរើសប្រភេទទិន្នន័យដែលចង់សម្អាត៖</p>
            <div style="display:flex;flex-direction:column;gap:10px;">
                <button onclick="cleanData('all')" style="padding:12px;background:#dc3545;color:white;border:none;border-radius:8px;cursor:pointer;font-size:15px;font-family:'Khmer OS','Khmer OS Muol','Arial',sans-serif;">🗑️ សម្អាតទិន្នន័យទាំងអស់</button>
                <button onclick="cleanData('attendance')" style="padding:12px;background:#fbbc04;color:#333;border:none;border-radius:8px;cursor:pointer;font-size:15px;font-family:'Khmer OS','Khmer OS Muol','Arial',sans-serif;">🗑️ សម្អាតតែវត្តមាន</button>
                <button onclick="cleanData('leaves')" style="padding:12px;background:#fbbc04;color:#333;border:none;border-radius:8px;cursor:pointer;font-size:15px;font-family:'Khmer OS','Khmer OS Muol','Arial',sans-serif;">🗑️ សម្អាតតែសុំច្បាប់</button>
                <button onclick="cleanData('missions')" style="padding:12px;background:#fbbc04;color:#333;border:none;border-radius:8px;cursor:pointer;font-size:15px;font-family:'Khmer OS','Khmer OS Muol','Arial',sans-serif;">🗑️ សម្អាតតែបេសកម្ម</button>
                <button onclick="closeModal('cleanModal')" style="padding:12px;background:#e8ecf1;color:#333;border:none;border-radius:8px;cursor:pointer;font-size:15px;font-family:'Khmer OS','Khmer OS Muol','Arial',sans-serif;">❌ បោះបង់</button>
            </div>
        </div>
    </div>

    <script>
    var currentDataVersion = {{ data_version }};
    var dataCheckInterval = null;
    var userLocation = null;
    var notificationInterval = null;
    var isFirstCheck = true;
    var deferredPrompt = null;
    var systemLocked = false;
    var autoUnlockTime = '06:00';
    var userLocked = {{ user_lock.is_locked|default(0) }};

    window.addEventListener('beforeinstallprompt', function(e) {
        e.preventDefault();
        deferredPrompt = e;
        var installBtn = document.getElementById('installBtn');
        if (installBtn) {
            installBtn.style.display = 'inline-block';
        }
    });

    document.getElementById('installBtn').addEventListener('click', function() {
        if (deferredPrompt) {
            deferredPrompt.prompt();
            deferredPrompt.userChoice.then(function(choiceResult) {
                if (choiceResult.outcome === 'accepted') {
                    console.log('User accepted the install prompt');
                    document.getElementById('installBtn').style.display = 'none';
                } else {
                    console.log('User dismissed the install prompt');
                }
                deferredPrompt = null;
            });
        }
    });

    window.addEventListener('appinstalled', function() {
        console.log('App installed successfully!');
        document.getElementById('installBtn').style.display = 'none';
    });

    if (window.matchMedia('(display-mode: standalone)').matches) {
        document.getElementById('installBtn').style.display = 'none';
    }

    if ('serviceWorker' in navigator) {
        navigator.serviceWorker.register('/static/sw.js')
        .then(function(reg) {
            console.log('Service Worker registered successfully!');
        })
        .catch(function(err) {
            console.log('Service Worker registration failed:', err);
        });
    }

    // ===== USER LOCK FUNCTIONS =====
    function updateUserLockUI(locked, autoUnlockTime) {
        var banner = document.getElementById('userLockBanner');
        var text = document.getElementById('userLockText');
        userLocked = locked;

        if (locked) {
            banner.className = 'user-lock-banner show';
            var msg = '🔒 អ្នកត្រូវបានបិទការចូលធ្វើការដោយ Admin';
            if (autoUnlockTime) {
                msg += ' (បើកដោយស្វ័យប្រវត្តិនៅម៉ោង ' + autoUnlockTime + ')';
            }
            text.textContent = msg;
            document.getElementById('checkinBtn').classList.add('disabled');
            document.getElementById('checkoutBtn').classList.add('disabled');
        } else {
            banner.className = 'user-lock-banner';
            if (!systemLocked) {
                document.getElementById('checkinBtn').classList.remove('disabled');
                document.getElementById('checkoutBtn').classList.remove('disabled');
            }
        }
    }

    // ===== SYSTEM LOCK FUNCTIONS =====
    function updateLockUI(lockStatus) {
        var banner = document.getElementById('lockBanner');
        var lockText = document.getElementById('lockText');
        var lockTime = document.getElementById('lockTime');
        var toggleBtn = document.getElementById('lockToggleBtn');
        var lockToggleModal = document.getElementById('lockToggleBtnModal');
        var unlockModal = document.getElementById('unlockBtnModal');
        var unlockScheduledInfo = document.getElementById('unlockScheduledInfo');
        var scheduledUnlockTime = document.getElementById('scheduledUnlockTime');

        systemLocked = lockStatus.is_locked == 1;
        autoUnlockTime = lockStatus.auto_unlock_time || '06:00';

        // Update banner
        if (systemLocked) {
            banner.className = 'lock-banner show';
            lockText.textContent = '🔒 ប្រព័ន្ធកំពុងបិទការចូលធ្វើការ';
            lockTime.textContent = '⏰ បើកដោយស្វ័យប្រវត្តិនៅម៉ោង ' + autoUnlockTime;
            if (toggleBtn) {
                toggleBtn.textContent = '🔓 បើកប្រព័ន្ធ';
                toggleBtn.style.background = 'rgba(255,255,255,0.2)';
            }
            if (lockToggleModal) {
                lockToggleModal.style.display = 'none';
                unlockModal.style.display = 'block';
                unlockModal.textContent = '🔓 បើកប្រព័ន្ធជាមួយម៉ោងកំណត់';
            }
            if (unlockScheduledInfo) {
                unlockScheduledInfo.style.display = 'block';
                scheduledUnlockTime.textContent = autoUnlockTime;
            }
            // Disable checkin/checkout buttons
            document.getElementById('checkinBtn').classList.add('disabled');
            document.getElementById('checkoutBtn').classList.add('disabled');

            // Show lock message
            var statusMsg = document.getElementById('statusMessage');
            statusMsg.className = 'status-message locked';
            statusMsg.innerHTML = '🔒 ប្រព័ន្ធកំពុងបិទការចូលធ្វើការ! សូមរង់ចាំរហូតដល់ម៉ោង ' + autoUnlockTime + ' ឬទាក់ទង Admin!';
            statusMsg.style.display = 'block';
        } else {
            banner.className = 'lock-banner show unlocked';
            lockText.textContent = '🔓 ប្រព័ន្ធកំពុងបើក';
            lockTime.textContent = '✅ អនុញ្ញាតអោយចុចចូល/ចេញធ្វើការ';
            if (toggleBtn) {
                toggleBtn.textContent = '🔒 បិទប្រព័ន្ធ';
                toggleBtn.style.background = 'rgba(220,53,69,0.3)';
            }
            if (lockToggleModal) {
                lockToggleModal.style.display = 'block';
                lockToggleModal.textContent = '🔒 បិទប្រព័ន្ធឥឡូវនេះ';
                unlockModal.style.display = 'none';
            }
            if (unlockScheduledInfo) {
                unlockScheduledInfo.style.display = 'none';
            }
            // Enable buttons if user not locked
            if (!userLocked) {
                document.getElementById('checkinBtn').classList.remove('disabled');
                document.getElementById('checkoutBtn').classList.remove('disabled');
            }

            var statusMsg = document.getElementById('statusMessage');
            if (statusMsg.className === 'status-message locked') {
                statusMsg.className = 'status-message info';
                statusMsg.innerHTML = '🔵 ប្រព័ន្ធកំពុងបើក។ អ្នកអាចចុចចូល/ចេញធ្វើការបាន!';
                statusMsg.style.display = 'block';
            }
        }

        // Update lock status text in modal
        var statusText = document.getElementById('lockStatusText');
        if (statusText) {
            if (systemLocked) {
                statusText.textContent = '🔒 ប្រព័ន្ធកំពុងបិទ (បើកដោយស្វ័យប្រវត្តិនៅម៉ោង ' + autoUnlockTime + ')';
                statusText.style.color = '#dc3545';
            } else {
                statusText.textContent = '🔓 ប្រព័ន្ធកំពុងបើក';
                statusText.style.color = '#34a853';
            }
        }
    }

    function getSystemLockStatus() {
        fetch('/get_system_lock_status')
        .then(function(res) { return res.json(); })
        .then(function(data) {
            updateLockUI(data);
        })
        .catch(function(err) { console.log('Lock status error:', err); });
    }

    function openSystemLockModal() {
        getSystemLockStatus();
        var autoTime = document.getElementById('autoUnlockTime');
        if (autoTime) {
            autoTime.value = autoUnlockTime || '06:00';
        }
        openModal('systemLockModal');
    }

    // ===== TOGGLE SYSTEM LOCK - កែសម្រួល =====
    function toggleSystemLockFromModal(lockState) {
        // lockState: 1 = lock, 0 = unlock (schedule)
        var autoTime = document.getElementById('autoUnlockTime');
        var unlockTime = autoTime ? autoTime.value : '06:00';

        if (!unlockTime) {
            unlockTime = '06:00';
        }

        var message = '';
        if (lockState === 1) {
            message = 'តើអ្នកចង់បិទប្រព័ន្ធចូលធ្វើការមែនទេ?';
        } else {
            message = 'តើអ្នកចង់កំណត់ម៉ោងបើកប្រព័ន្ធដោយស្វ័យប្រវត្តិនៅម៉ោង ' + unlockTime + ' មែនទេ?\n\nប្រព័ន្ធនឹងបើកដោយស្វ័យប្រវត្តិនៅម៉ោង ' + unlockTime;
        }

        if (!confirm(message)) {
            return;
        }

        // For unlock (lockState = 0), we still send lock_state=1 to the server
        // because we want to keep the system locked until the scheduled time
        var actualLockState = (lockState === 1) ? 1 : 1; // Keep locked
        var isUnlockRequest = (lockState === 0);

        // If unlocking, we keep the system locked but set the auto_unlock_time
        // The system will auto-unlock when the time is reached
        var newState = isUnlockRequest ? 1 : lockState;

        fetch('/toggle_system_lock', {
            method: 'POST',
            headers: { 'Content-Type': 'application/json' },
            body: JSON.stringify({
                lock_state: newState,
                auto_unlock_time: unlockTime
            })
        })
        .then(function(res) { return res.json(); })
        .then(function(result) {
            alert(result.message);
            if (result.success) {
                if (isUnlockRequest) {
                    alert('⏰ ប្រព័ន្ធនឹងបើកដោយស្វ័យប្រវត្តិនៅម៉ោង ' + unlockTime);
                }
                getSystemLockStatus();
                if (document.getElementById('systemLockModal').classList.contains('show')) {
                    closeModal('systemLockModal');
                }
            }
        })
        .catch(function(err) {
            console.error('Error:', err);
            alert('មានបញ្ហា: ' + err.message);
        });
    }

    // ===== END SYSTEM LOCK FUNCTIONS =====

    function checkDataVersion() {
        fetch('/get_data_version')
        .then(function(res) { return res.json(); })
        .then(function(data) {
            if (data.version > currentDataVersion) {
                console.log('Data changed! Refreshing...');
                currentDataVersion = data.version;
                location.reload();
            }
        })
        .catch(function(err) { console.log('Data version check error:', err); });
    }

    function startDataVersionChecker() {
        checkDataVersion();
        dataCheckInterval = setInterval(checkDataVersion, 3000);
    }

    function stopDataVersionChecker() {
        if (dataCheckInterval) {
            clearInterval(dataCheckInterval);
            dataCheckInterval = null;
        }
    }

    function updateButtonStatus() {
        fetch('/get_checkin_status')
        .then(function(res) { return res.json(); })
        .then(function(data) {
            var checkinBtn = document.getElementById('checkinBtn');
            var checkoutBtn = document.getElementById('checkoutBtn');

            if (systemLocked || userLocked) {
                checkinBtn.classList.add('disabled');
                checkoutBtn.classList.add('disabled');
                return;
            }

            checkinBtn.disabled = false;
            checkoutBtn.disabled = false;

            if (data.has_checkin) {
                var shiftNames = {1: 'វគ្គ 1', 2: 'វគ្គ 2', 3: 'វគ្គ 3'};
                var shiftText = shiftNames[data.shift] || '';
                var statusMsg = document.getElementById('statusMessage');
                statusMsg.className = 'status-message warning';
                statusMsg.innerHTML =
                    '🟢 អ្នកកំពុងធ្វើការ (' + shiftText + ') ចាប់ពីម៉ោង ' +
                    (data.check_in_time ? data.check_in_time.split(' ')[1].slice(0,5) : '') +
                    '។ សូមចុច "ចេញពីធ្វើការ" មុនពេលចូលម្តងទៀត!';
                statusMsg.style.display = 'block';
            } else if (!systemLocked && !userLocked) {
                var statusMsg = document.getElementById('statusMessage');
                statusMsg.className = 'status-message info';
                statusMsg.innerHTML =
                    '🔵 អ្នកមិនទាន់ចូលធ្វើការនៅថ្ងៃនេះទេ។ សូមចុច "ចូលធ្វើការ" ដើម្បីចាប់ផ្តើម!';
                statusMsg.style.display = 'block';
            }
        })
        .catch(function(err) { console.log('Update status error:', err); });
    }

    function openModal(id) {
        document.getElementById(id).classList.add('show');
    }

    function closeModal(id) {
        var modal = document.getElementById(id);
        if (modal) {
            modal.classList.remove('show');
        }
    }

    function closeLocationModal() { closeModal('locationModal'); }
    function closeRequestsModal() { closeModal('requestsModal'); }
    function openLocationModal() { openModal('locationModal'); }
    function openRequestsModal() { openModal('requestsModal'); loadRequests(); }
    function openCleanModal() { openModal('cleanModal'); }

    function cleanData(type) {
        var typeNames = {
            'all': 'ទិន្នន័យទាំងអស់',
            'attendance': 'ទិន្នន័យវត្តមាន',
            'leaves': 'ទិន្នន័យសុំច្បាប់',
            'missions': 'ទិន្នន័យបេសកម្ម'
        };
        if (!confirm('តើអ្នកពិតជាចង់សម្អាត ' + typeNames[type] + ' មែនទេ?\n\nការសម្អាតនេះមិនអាចស្តារឡើងវិញបានទេ!')) return;
        if (!confirm('សូមបញ្ជាក់ម្តងទៀត: តើអ្នកប្រាកដជាចង់លុប ' + typeNames[type] + ' មែនទេ?')) return;

        fetch('/clean_data', {
            method: 'POST',
            headers: { 'Content-Type': 'application/json' },
            body: JSON.stringify({ type: type })
        })
        .then(function(response) { return response.json(); })
        .then(function(result) {
            alert(result.message);
            if (result.success) {
                closeModal('cleanModal');
                location.reload();
            }
        })
        .catch(function(err) {
            console.error('Error:', err);
            alert('មានបញ្ហា: ' + err);
        });
    }

    function loadRequests() {
        fetch('/get_pending_requests')
        .then(function(res) { return res.json(); })
        .then(function(data) {
            var html = '';
            if (data.leaves.length === 0 && data.missions.length === 0) {
                html = '<p style="text-align:center;color:#888;padding:20px;">📭 មិនមានសំណើរង់ចាំ</p>';
            }
            for (var i = 0; i < data.leaves.length; i++) {
                var item = data.leaves[i];
                var attachmentHtml = '';
                if (item.attachment) {
                    var ext = item.attachment.split('.').pop().toLowerCase();
                    if (['jpg', 'jpeg', 'png', 'gif', 'bmp', 'webp'].indexOf(ext) !== -1) {
                        attachmentHtml = '<div style="margin-top:5px;"><a href="' + item.attachment + '" target="_blank" style="color:#1a73e8;text-decoration:none;">🖼️ មើលរូបភាព</a></div>';
                    } else {
                        attachmentHtml = '<div style="margin-top:5px;"><a href="' + item.attachment + '" target="_blank" style="color:#1a73e8;text-decoration:none;">📎 ទាញយកឯកសារ</a></div>';
                    }
                }
                html += '<div class="request-item">' +
                    '<div class="request-user">📋 ' + item.full_name + ' (' + item.username + ')</div>' +
                    '<div class="request-detail">សុំច្បាប់: ' + item.days + ' ថ្ងៃ | ' + item.start_date + ' ដល់ ' + item.end_date + ' | មូលហេតុ: ' + (item.reason || 'មិនបានបញ្ជាក់') + '</div>' +
                    attachmentHtml +
                    '<div class="request-actions">' +
                        '<button class="approve-btn" onclick="approveRequest(\'leave\', ' + item.id + ')">✅ អនុម័ត</button>' +
                        '<button class="reject-btn" onclick="rejectRequest(\'leave\', ' + item.id + ')">❌ បដិសេធ</button>' +
                    '</div>' +
                '</div>';
            }
            for (var j = 0; j < data.missions.length; j++) {
                var item2 = data.missions[j];
                var attachmentHtml2 = '';
                if (item2.attachment) {
                    var ext2 = item2.attachment.split('.').pop().toLowerCase();
                    if (['jpg', 'jpeg', 'png', 'gif', 'bmp', 'webp'].indexOf(ext2) !== -1) {
                        attachmentHtml2 = '<div style="margin-top:5px;"><a href="' + item2.attachment + '" target="_blank" style="color:#1a73e8;text-decoration:none;">🖼️ មើលរូបភាព</a></div>';
                    } else {
                        attachmentHtml2 = '<div style="margin-top:5px;"><a href="' + item2.attachment + '" target="_blank" style="color:#1a73e8;text-decoration:none;">📎 ទាញយកឯកសារ</a></div>';
                    }
                }
                html += '<div class="request-item">' +
                    '<div class="request-user">🚗 ' + item2.full_name + ' (' + item2.username + ')</div>' +
                    '<div class="request-detail">បេសកម្ម: ' + item2.days + ' ថ្ងៃ | ' + item2.start_date + ' ដល់ ' + item2.end_date + ' | ទីតាំង: ' + (item2.destination || 'មិនបានបញ្ជាក់') + '</div>' +
                    attachmentHtml2 +
                    '<div class="request-actions">' +
                        '<button class="approve-btn" onclick="approveRequest(\'mission\', ' + item2.id + ')">✅ អនុម័ត</button>' +
                        '<button class="reject-btn" onclick="rejectRequest(\'mission\', ' + item2.id + ')">❌ បដិសេធ</button>' +
                    '</div>' +
                '</div>';
            }
            document.getElementById('requestsList').innerHTML = html;
        })
        .catch(function(err) {
            document.getElementById('requestsList').innerHTML = '<p style="text-align:center;color:red;">មានបញ្ហា: ' + err + '</p>';
        });
    }

    function approveRequest(type, id) {
        if (!confirm('តើអ្នកចង់អនុម័តសំណើនេះមែនទេ?')) return;
        fetch('/approve_request', {
            method: 'POST',
            headers: { 'Content-Type': 'application/json' },
            body: JSON.stringify({ type: type, id: id })
        })
        .then(function(res) { return res.json(); })
        .then(function(result) {
            alert(result.message);
            if (result.success) {
                loadRequests();
                window.location.reload();
            }
        })
        .catch(function(err) {
            console.error('Error:', err);
            alert('មានបញ្ហា: ' + err.message);
        });
    }

    function rejectRequest(type, id) {
        if (!confirm('តើអ្នកចង់បដិសេធសំណើនេះមែនទេ?')) return;
        fetch('/reject_request', {
            method: 'POST',
            headers: { 'Content-Type': 'application/json' },
            body: JSON.stringify({ type: type, id: id })
        })
        .then(function(res) { return res.json(); })
        .then(function(result) {
            alert(result.message);
            if (result.success) {
                loadRequests();
                window.location.reload();
            }
        })
        .catch(function(err) {
            console.error('Error:', err);
            alert('មានបញ្ហា: ' + err.message);
        });
    }

    function getCurrentLocation() {
        if (!navigator.geolocation) {
            alert('កម្មវិធីរុករករបស់អ្នកមិនគាំទ្រការចាប់ទីតាំងទេ!');
            return;
        }
        navigator.geolocation.getCurrentPosition(
            function(pos) {
                document.getElementById('companyLat').value = pos.coords.latitude;
                document.getElementById('companyLng').value = pos.coords.longitude;
                document.getElementById('locationInfo').innerHTML =
                    '📍 ទីតាំងបច្ចុប្បន្ន: ' + pos.coords.latitude + ', ' + pos.coords.longitude;
            },
            function(err) {
                alert('មិនអាចចាប់យកទីតាំងបាន! សូមបញ្ចូលដោយដៃ។');
                console.log(err);
            }
        );
    }

    function saveCompanyLocation() {
        var lat = document.getElementById('companyLat').value.trim();
        var lng = document.getElementById('companyLng').value.trim();
        if (!lat || !lng) {
            alert('សូមបញ្ចូលទីតាំងឲ្យបានពេញលេញ!');
            return;
        }
        fetch('/save_location', {
            method: 'POST',
            headers: { 'Content-Type': 'application/json' },
            body: JSON.stringify({ lat: parseFloat(lat), lng: parseFloat(lng) })
        })
        .then(function(res) { return res.json(); })
        .then(function(data) {
            if (data.success) {
                alert('✅ រក្សាទុកទីតាំងជោគជ័យ!');
                document.getElementById('locationInfo').innerHTML =
                    '📍 ទីតាំងបច្ចុប្បន្ន: ' + lat + ', ' + lng;
                closeLocationModal();
                location.reload();
            } else {
                alert('❌ ' + data.message);
            }
        })
        .catch(function(err) { alert('មានបញ្ហា: ' + err); });
    }

    function openCheckInModal() {
        if (systemLocked) {
            alert('⛔ ប្រព័ន្ធកំពុងបិទការចូលធ្វើការ! សូមរង់ចាំរហូតដល់ម៉ោង ' + autoUnlockTime + ' ឬទាក់ទង Admin!');
            return;
        }
        if (userLocked) {
            alert('⛔ អ្នកត្រូវបានបិទការចូលធ្វើការដោយ Admin! សូមទាក់ទង Admin!');
            return;
        }

        if (!navigator.geolocation) {
            alert('កម្មវិធីរុករករបស់អ្នកមិនគាំទ្រការចាប់ទីតាំងទេ!');
            return;
        }

        document.getElementById('checkInLocationInfo').innerHTML = '⏳ កំពុងចាប់យកទីតាំង...';

        navigator.geolocation.getCurrentPosition(
            function(pos) {
                userLocation = {
                    lat: pos.coords.latitude,
                    lng: pos.coords.longitude
                };
                document.getElementById('checkInLocationInfo').innerHTML =
                    '📍 ទីតាំងរបស់អ្នក: ' + pos.coords.latitude.toFixed(6) + ', ' + pos.coords.longitude.toFixed(6) +
                    '<br>📏 កំពុងពិនិត្យចម្ងាយ...';
                openModal('checkInModal');
            },
            function(err) {
                document.getElementById('checkInLocationInfo').innerHTML =
                    '⚠️ មិនអាចចាប់យកទីតាំង! សូមបើក GPS ហើយចុចសាកល្បងម្តងទៀត<br>' +
                    '<button onclick="openCheckInModal()" style="margin-top:10px;padding:8px 20px;background:#1a73e8;color:white;border:none;border-radius:8px;cursor:pointer;">🔄 សាកល្បងម្តងទៀត</button>';
                console.log(err);
            }
        );
    }

    function submitCheckIn() {
        if (!userLocation) {
            alert('សូមចាប់យកទីតាំងរបស់អ្នកមុន!');
            return;
        }

        var shift = document.getElementById('shiftSelect').value;

        fetch('/get_company_location')
        .then(function(res) { return res.json(); })
        .then(function(data) {
            if (!data.lat || !data.lng) {
                alert('សូមឲ្យ Admin កំណត់ទីតាំងក្រុមហ៊ុនជាមុនសិន!');
                return;
            }

            fetch('/check_in', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({
                    user_lat: userLocation.lat,
                    user_lng: userLocation.lng,
                    company_lat: data.lat,
                    company_lng: data.lng,
                    shift: parseInt(shift)
                })
            })
            .then(function(res) { return res.json(); })
            .then(function(result) {
                alert(result.message);
                if (result.success) {
                    closeModal('checkInModal');
                    window.location.reload();
                }
            })
            .catch(function(err) {
                console.error('Error:', err);
                alert('មានបញ្ហា: ' + err.message);
            });
        })
        .catch(function(err) {
            console.error('Error:', err);
            alert('មានបញ្ហា: ' + err.message);
        });
    }

    function handleCheckOut() {
        if (systemLocked) {
            alert('⛔ ប្រព័ន្ធកំពុងបិទការចូលធ្វើការ! សូមរង់ចាំរហូតដល់ម៉ោង ' + autoUnlockTime + ' ឬទាក់ទង Admin!');
            return;
        }
        if (userLocked) {
            alert('⛔ អ្នកត្រូវបានបិទការចូលធ្វើការដោយ Admin! សូមទាក់ទង Admin!');
            return;
        }

        if (!confirm('តើអ្នកចង់ចេញពីធ្វើការមែនទេ?')) return;

        if (!navigator.geolocation) {
            alert('កម្មវិធីរុករករបស់អ្នកមិនគាំទ្រការចាប់ទីតាំងទេ!');
            return;
        }

        var statusMsg = document.getElementById('statusMessage');
        if (statusMsg) {
            statusMsg.className = 'status-message info';
            statusMsg.innerHTML = '⏳ កំពុងចាប់យកទីតាំងសម្រាប់ចេញធ្វើការ...';
            statusMsg.style.display = 'block';
        }

        navigator.geolocation.getCurrentPosition(
            function(pos) {
                var location = {
                    lat: pos.coords.latitude,
                    lng: pos.coords.longitude
                };

                fetch('/get_company_location')
                .then(function(res) { return res.json(); })
                .then(function(data) {
                    if (!data.lat || !data.lng) {
                        alert('សូមឲ្យ Admin កំណត់ទីតាំងក្រុមហ៊ុនជាមុនសិន!');
                        return;
                    }

                    fetch('/check_out', {
                        method: 'POST',
                        headers: { 'Content-Type': 'application/json' },
                        body: JSON.stringify({
                            user_lat: location.lat,
                            user_lng: location.lng,
                            company_lat: data.lat,
                            company_lng: data.lng
                        })
                    })
                    .then(function(res) { return res.json(); })
                    .then(function(result) {
                        alert(result.message);
                        if (result.success) {
                            window.location.reload();
                        }
                    })
                    .catch(function(err) {
                        console.error('Error:', err);
                        alert('មានបញ្ហា: ' + err.message);
                    });
                })
                .catch(function(err) {
                    console.error('Error:', err);
                    alert('មានបញ្ហា: ' + err.message);
                });
            },
            function(err) {
                var statusMsg = document.getElementById('statusMessage');
                if (statusMsg) {
                    statusMsg.className = 'status-message danger';
                    statusMsg.innerHTML = '⚠️ មិនអាចចាប់យកទីតាំង! សូមបើក GPS ហើយព្យាយាមម្តងទៀត';
                    statusMsg.style.display = 'block';
                }
                console.log('Geolocation error:', err);
            }
        );
    }

    function handleLeave() {
        var html =
            '<div style="padding:10px;">' +
                '<div style="margin-bottom:10px;">' +
                    '<label>ចំនួនថ្ងៃ (អាចជា 0.5):</label>' +
                    '<input type="number" id="leaveDays" step="0.5" min="0.5" value="1" style="width:100%;padding:8px;border:2px solid #e8ecf1;border-radius:8px;">' +
                '</div>' +
                '<div style="margin-bottom:10px;">' +
                    '<label>ចាប់ពីថ្ងៃ:</label>' +
                    '<input type="date" id="leaveStartDate" style="width:100%;padding:8px;border:2px solid #e8ecf1;border-radius:8px;">' +
                '</div>' +
                '<div style="margin-bottom:10px;">' +
                    '<label>ដល់ថ្ងៃ:</label>' +
                    '<input type="date" id="leaveEndDate" style="width:100%;padding:8px;border:2px solid #e8ecf1;border-radius:8px;">' +
                '</div>' +
                '<div style="margin-bottom:10px;">' +
                    '<label>មូលហេតុ:</label>' +
                    '<textarea id="leaveReason" rows="3" style="width:100%;padding:8px;border:2px solid #e8ecf1;border-radius:8px;font-family:\'Khmer OS\',Arial;" placeholder="សូមបញ្ចូលមូលហេតុ..."></textarea>' +
                '</div>' +
                '<div style="margin-bottom:10px;">' +
                    '<label>ឯកសារភ្ជាប់ (រូបភាព ឬ PDF):</label>' +
                    '<input type="file" id="leaveAttachment" accept="image/*,.pdf,.doc,.docx" style="width:100%;padding:8px;border:2px solid #e8ecf1;border-radius:8px;font-family:\'Khmer OS\',Arial;">' +
                    '<div style="font-size:12px;color:#888;margin-top:4px;">📎 អាចភ្ជាប់រូបភាព, PDF, Word (អតិបរមា 5MB)</div>' +
                '</div>' +
                '<div id="attachmentPreview" style="display:none;margin-bottom:10px;padding:10px;background:#f8f9fa;border-radius:8px;text-align:center;">' +
                    '<img id="previewImage" src="" style="max-width:200px;max-height:200px;border-radius:8px;display:none;">' +
                    '<a id="previewLink" href="#" target="_blank" style="display:none;color:#1a73e8;text-decoration:none;">📄 មើលឯកសារ</a>' +
                    '<button onclick="removeAttachment()" style="display:block;margin-top:5px;padding:4px 12px;background:#dc3545;color:white;border:none;border-radius:6px;cursor:pointer;font-size:12px;">❌ លុប</button>' +
                '</div>' +
                '<div style="display:flex;gap:10px;margin-top:10px;">' +
                    '<button onclick="submitLeave()" style="flex:1;padding:10px;background:#34a853;color:white;border:none;border-radius:8px;cursor:pointer;font-size:16px;font-family:\'Khmer OS\',Arial;">✅ បញ្ជូន</button>' +
                    '<button onclick="closeModal(\'leaveModal\')" style="flex:1;padding:10px;background:#e8ecf1;color:#333;border:none;border-radius:8px;cursor:pointer;font-size:16px;font-family:\'Khmer OS\',Arial;">បោះបង់</button>' +
                '</div>' +
            '</div>';

        var modal = document.createElement('div');
        modal.id = 'leaveModal';
        modal.className = 'modal show';
        modal.innerHTML =
            '<div class="modal-content" style="max-width:500px;">' +
                '<h3>📋 សុំច្បាប់</h3>' +
                html +
            '</div>';
        document.body.appendChild(modal);

        var today = new Date().toISOString().split('T')[0];
        document.getElementById('leaveStartDate').value = today;
        document.getElementById('leaveEndDate').value = today;

        var fileInput = document.getElementById('leaveAttachment');
        fileInput.addEventListener('change', function(e) {
            var file = this.files[0];
            if (file) {
                var previewDiv = document.getElementById('attachmentPreview');
                var previewImage = document.getElementById('previewImage');
                var previewLink = document.getElementById('previewLink');

                previewDiv.style.display = 'block';
                previewImage.style.display = 'none';
                previewLink.style.display = 'none';

                if (file.type.startsWith('image/')) {
                    var reader = new FileReader();
                    reader.onload = function(e) {
                        previewImage.src = e.target.result;
                        previewImage.style.display = 'block';
                    };
                    reader.readAsDataURL(file);
                } else {
                    previewLink.textContent = '📄 ' + file.name;
                    previewLink.style.display = 'block';
                    previewLink.href = '#';
                    previewLink.onclick = function(e) {
                        e.preventDefault();
                        alert('ឯកសារ: ' + file.name + '\nទំហំ: ' + (file.size / 1024).toFixed(2) + ' KB');
                    };
                }
            }
        });
    }

    function removeAttachment() {
        var fileInput = document.getElementById('leaveAttachment');
        fileInput.value = '';
        document.getElementById('attachmentPreview').style.display = 'none';
        document.getElementById('previewImage').style.display = 'none';
        document.getElementById('previewLink').style.display = 'none';
    }

    function submitLeave() {
        var days = document.getElementById('leaveDays').value;
        var start_date = document.getElementById('leaveStartDate').value;
        var end_date = document.getElementById('leaveEndDate').value;
        var reason = document.getElementById('leaveReason').value.trim();
        var fileInput = document.getElementById('leaveAttachment');
        var file = fileInput.files[0];

        if (!start_date || !end_date) {
            alert('សូមជ្រើសរើសថ្ងៃ!');
            return;
        }
        if (!reason) {
            alert('សូមបញ្ចូលមូលហេតុ!');
            return;
        }
        if (parseFloat(days) <= 0) {
            alert('សូមបញ្ចូលចំនួនថ្ងៃឲ្យបានត្រឹមត្រូវ!');
            return;
        }

        if (file && file.size > 5 * 1024 * 1024) {
            alert('ឯកសារធំពេក! សូមជ្រើសរើសឯកសារដែលមានទំហំតិចជាង 5MB');
            return;
        }

        var formData = new FormData();
        formData.append('start_date', start_date);
        formData.append('end_date', end_date);
        formData.append('days', parseFloat(days));
        formData.append('reason', reason);
        if (file) {
            formData.append('attachment', file);
        }

        fetch('/request_leave', {
            method: 'POST',
            body: formData
        })
        .then(function(response) { return response.json(); })
        .then(function(result) {
            alert(result.message);
            if (result.success) {
                closeModal('leaveModal');
                window.location.reload();
            }
        })
        .catch(function(err) {
            console.error("Error:", err);
            alert('មានបញ្ហា: ' + err.message);
        });
    }

    function handleMission() {
        var html =
            '<div style="padding:10px;">' +
                '<div style="margin-bottom:10px;">' +
                    '<label>ចំនួនថ្ងៃ (អាចជា 0.5):</label>' +
                    '<input type="number" id="missionDays" step="0.5" min="0.5" value="1" style="width:100%;padding:8px;border:2px solid #e8ecf1;border-radius:8px;">' +
                '</div>' +
                '<div style="margin-bottom:10px;">' +
                    '<label>ចាប់ពីថ្ងៃ:</label>' +
                    '<input type="date" id="missionStartDate" style="width:100%;padding:8px;border:2px solid #e8ecf1;border-radius:8px;">' +
                '</div>' +
                '<div style="margin-bottom:10px;">' +
                    '<label>ដល់ថ្ងៃ:</label>' +
                    '<input type="date" id="missionEndDate" style="width:100%;padding:8px;border:2px solid #e8ecf1;border-radius:8px;">' +
                '</div>' +
                '<div style="margin-bottom:10px;">' +
                    '<label>ទីតាំងបេសកម្ម:</label>' +
                    '<input type="text" id="missionDestination" style="width:100%;padding:8px;border:2px solid #e8ecf1;border-radius:8px;font-family:\'Khmer OS\',Arial;" placeholder="សូមបញ្ចូលទីតាំង...">' +
                '</div>' +
                '<div style="margin-bottom:10px;">' +
                    '<label>គោលបំណង:</label>' +
                    '<textarea id="missionPurpose" rows="3" style="width:100%;padding:8px;border:2px solid #e8ecf1;border-radius:8px;font-family:\'Khmer OS\',Arial;" placeholder="សូមបញ្ចូលគោលបំណង..."></textarea>' +
                '</div>' +
                '<div style="margin-bottom:10px;">' +
                    '<label>ឯកសារភ្ជាប់ (រូបភាព ឬ PDF):</label>' +
                    '<input type="file" id="missionAttachment" accept="image/*,.pdf,.doc,.docx" style="width:100%;padding:8px;border:2px solid #e8ecf1;border-radius:8px;font-family:\'Khmer OS\',Arial;">' +
                    '<div style="font-size:12px;color:#888;margin-top:4px;">📎 អាចភ្ជាប់រូបភាព, PDF, Word (អតិបរមា 5MB)</div>' +
                '</div>' +
                '<div id="missionAttachmentPreview" style="display:none;margin-bottom:10px;padding:10px;background:#f8f9fa;border-radius:8px;text-align:center;">' +
                    '<img id="missionPreviewImage" src="" style="max-width:200px;max-height:200px;border-radius:8px;display:none;">' +
                    '<a id="missionPreviewLink" href="#" target="_blank" style="display:none;color:#1a73e8;text-decoration:none;">📄 មើលឯកសារ</a>' +
                    '<button onclick="removeMissionAttachment()" style="display:block;margin-top:5px;padding:4px 12px;background:#dc3545;color:white;border:none;border-radius:6px;cursor:pointer;font-size:12px;">❌ លុប</button>' +
                '</div>' +
                '<div style="display:flex;gap:10px;margin-top:10px;">' +
                    '<button onclick="submitMission()" style="flex:1;padding:10px;background:#1a73e8;color:white;border:none;border-radius:8px;cursor:pointer;font-size:16px;font-family:\'Khmer OS\',Arial;">✅ បញ្ជូន</button>' +
                    '<button onclick="closeModal(\'missionModal\')" style="flex:1;padding:10px;background:#e8ecf1;color:#333;border:none;border-radius:8px;cursor:pointer;font-size:16px;font-family:\'Khmer OS\',Arial;">បោះបង់</button>' +
                '</div>' +
            '</div>';

        var modal = document.createElement('div');
        modal.id = 'missionModal';
        modal.className = 'modal show';
        modal.innerHTML =
            '<div class="modal-content" style="max-width:500px;">' +
                '<h3>🚗 សុំបេសកម្ម</h3>' +
                html +
            '</div>';
        document.body.appendChild(modal);

        var today = new Date().toISOString().split('T')[0];
        document.getElementById('missionStartDate').value = today;
        document.getElementById('missionEndDate').value = today;

        var fileInput = document.getElementById('missionAttachment');
        fileInput.addEventListener('change', function(e) {
            var file = this.files[0];
            if (file) {
                var previewDiv = document.getElementById('missionAttachmentPreview');
                var previewImage = document.getElementById('missionPreviewImage');
                var previewLink = document.getElementById('missionPreviewLink');

                previewDiv.style.display = 'block';
                previewImage.style.display = 'none';
                previewLink.style.display = 'none';

                if (file.type.startsWith('image/')) {
                    var reader = new FileReader();
                    reader.onload = function(e) {
                        previewImage.src = e.target.result;
                        previewImage.style.display = 'block';
                    };
                    reader.readAsDataURL(file);
                } else {
                    previewLink.textContent = '📄 ' + file.name;
                    previewLink.style.display = 'block';
                    previewLink.href = '#';
                    previewLink.onclick = function(e) {
                        e.preventDefault();
                        alert('ឯកសារ: ' + file.name + '\nទំហំ: ' + (file.size / 1024).toFixed(2) + ' KB');
                    };
                }
            }
        });
    }

    function removeMissionAttachment() {
        var fileInput = document.getElementById('missionAttachment');
        fileInput.value = '';
        document.getElementById('missionAttachmentPreview').style.display = 'none';
        document.getElementById('missionPreviewImage').style.display = 'none';
        document.getElementById('missionPreviewLink').style.display = 'none';
    }

    function submitMission() {
        var days = document.getElementById('missionDays').value;
        var start_date = document.getElementById('missionStartDate').value;
        var end_date = document.getElementById('missionEndDate').value;
        var destination = document.getElementById('missionDestination').value.trim();
        var purpose = document.getElementById('missionPurpose').value.trim();
        var fileInput = document.getElementById('missionAttachment');
        var file = fileInput.files[0];

        if (!start_date || !end_date) {
            alert('សូមជ្រើសរើសថ្ងៃ!');
            return;
        }
        if (!destination) {
            alert('សូមបញ្ចូលទីតាំងបេសកម្ម!');
            return;
        }
        if (parseFloat(days) <= 0) {
            alert('សូមបញ្ចូលចំនួនថ្ងៃឲ្យបានត្រឹមត្រូវ!');
            return;
        }

        if (file && file.size > 5 * 1024 * 1024) {
            alert('ឯកសារធំពេក! សូមជ្រើសរើសឯកសារដែលមានទំហំតិចជាង 5MB');
            return;
        }

        var formData = new FormData();
        formData.append('start_date', start_date);
        formData.append('end_date', end_date);
        formData.append('days', parseFloat(days));
        formData.append('destination', destination);
        formData.append('purpose', purpose || 'បំពេញបេសកម្ម');
        if (file) {
            formData.append('attachment', file);
        }

        fetch('/request_mission', {
            method: 'POST',
            body: formData
        })
        .then(function(response) { return response.json(); })
        .then(function(result) {
            alert(result.message);
            if (result.success) {
                closeModal('missionModal');
                window.location.reload();
            }
        })
        .catch(function(err) {
            console.error("Error:", err);
            alert('មានបញ្ហា: ' + err.message);
        });
    }

    function checkNewRequests() {
        fetch('/check_new_requests')
        .then(function(res) { return res.json(); })
        .then(function(data) {
            var requestBtn = document.querySelector('.header-right .nav-btn[onclick="openRequestsModal()"]');
            if (requestBtn) {
                var oldBadge = requestBtn.querySelector('span');
                if (oldBadge) {
                    oldBadge.remove();
                }
                if (data.count > 0) {
                    var badge = document.createElement('span');
                    badge.style.cssText = 'background:red;color:white;border-radius:50%;padding:0 6px;font-size:11px;margin-left:3px;';
                    badge.textContent = data.count;
                    requestBtn.appendChild(badge);
                }
            }

            if (!isFirstCheck && data.new && data.message) {
                showNotification(data.message);
            }
            isFirstCheck = false;
        })
        .catch(function(err) { console.log('Notification error:', err); });
    }

    function deleteAttendance(id) {
        console.log('Delete button clicked for ID:', id);

        if (!confirm('តើអ្នកចង់លុបទិន្នន័យនេះមែនទេ?')) {
            return;
        }

        fetch('/delete_attendance/' + id, {
            method: 'POST',
            headers: {
                'Content-Type': 'application/json'
            }
        })
        .then(function(response) {
            console.log('Response status:', response.status);
            return response.json();
        })
        .then(function(result) {
            console.log('Result:', result);
            alert(result.message);
            if (result.success) {
                window.location.reload();
            }
        })
        .catch(function(err) {
            console.error('Error:', err);
            alert('មានបញ្ហា: ' + err.message);
        });
    }

    var style = document.createElement('style');
    style.textContent =
        '@keyframes slideIn {' +
            'from { transform: translateX(100%); opacity: 0; }' +
            'to { transform: translateX(0); opacity: 1; }' +
        '}' +
        '@keyframes slideOut {' +
            'from { transform: translateX(0); opacity: 1; }' +
            'to { transform: translateX(100%); opacity: 0; }' +
        '}';
    document.head.appendChild(style);

    // Initialize
    getSystemLockStatus();
    updateButtonStatus();
    startDataVersionChecker();

    {% if session.role == 'admin' %}
    startNotificationChecker();
    {% endif %}

    setInterval(updateButtonStatus, 10000);
    setInterval(getSystemLockStatus, 60000);

    window.addEventListener('beforeunload', function() {
        stopDataVersionChecker();
        stopNotificationChecker();
    });
    </script>
</body>
</html>
'''

# ============================================================
# HTML TEMPLATES (unchanged)
# ============================================================

REGISTER_HTML = '''<!DOCTYPE html>
<html>
<head>
    <title>ចុះឈ្មោះអ្នកប្រើប្រាស់</title>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif; background: #f0f2f5; padding: 15px; min-height: 100vh; }
        .container { max-width: 600px; margin: 0 auto; }
        .header {
            background: linear-gradient(135deg, #1a73e8, #0d47a1);
            color: white;
            padding: 15px 20px;
            border-radius: 12px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            box-shadow: 0 4px 15px rgba(26, 115, 232, 0.3);
            margin-bottom: 20px;
            flex-wrap: wrap;
            gap: 10px;
        }
        .header-left h2 { font-size: 18px; font-weight: 600; }
        .header-right { display: flex; align-items: center; gap: 10px; }
        .header-right .back-link {
            color: white;
            text-decoration: none;
            background: rgba(255,255,255,0.2);
            padding: 6px 16px;
            border-radius: 20px;
            font-size: 13px;
            transition: all 0.3s;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
        }
        .header-right .back-link:hover { background: rgba(255,255,255,0.35); }
        .header-right .user-name {
            color: white;
            padding: 6px 14px;
            border-radius: 20px;
            font-size: 13px;
            background: rgba(255,255,255,0.15);
        }
        .form-container {
            background: white;
            padding: 30px;
            border-radius: 12px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.08);
        }
        .form-container h3 { color: #333; font-size: 20px; margin-bottom: 5px; }
        .form-container .sub-title { color: #888; font-size: 14px; margin-bottom: 20px; }
        .form-group { margin-bottom: 18px; }
        .form-group label { display: block; font-weight: 600; color: #555; font-size: 14px; margin-bottom: 5px; }
        .form-group label .required { color: #dc3545; }
        .form-group input, .form-group select {
            width: 100%;
            padding: 12px 14px;
            border: 2px solid #e8ecf1;
            border-radius: 10px;
            font-size: 15px;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
            transition: all 0.3s;
            background: #fafafa;
        }
        .form-group input:focus, .form-group select:focus {
            outline: none;
            border-color: #1a73e8;
            background: white;
            box-shadow: 0 0 0 3px rgba(26, 115, 232, 0.1);
        }
        .form-group .hint { font-size: 12px; color: #aaa; margin-top: 4px; }
        .form-row { display: grid; grid-template-columns: 1fr 1fr; gap: 15px; }
        .btn-submit {
            width: 100%;
            padding: 14px;
            background: linear-gradient(135deg, #34a853, #1e7e34);
            color: white;
            border: none;
            border-radius: 10px;
            font-size: 18px;
            font-weight: 600;
            cursor: pointer;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
            transition: all 0.3s;
            margin-top: 5px;
        }
        .btn-submit:hover { transform: translateY(-2px); box-shadow: 0 6px 20px rgba(52, 168, 83, 0.3); }
        .btn-submit:active { transform: scale(0.98); }
        .message {
            padding: 12px 16px;
            border-radius: 10px;
            margin-bottom: 20px;
            font-size: 14px;
        }
        .message.success { background: #d4edda; color: #155724; border-left: 4px solid #28a745; }
        .message.error { background: #f8d7da; color: #721c24; border-left: 4px solid #dc3545; }
        .footer { text-align: center; padding: 20px 0 5px; color: #aaa; font-size: 12px; }
        @media (max-width: 600px) {
            .form-row { grid-template-columns: 1fr; gap: 0; }
            .header { flex-direction: column; text-align: center; }
            .header-left h2 { font-size: 16px; }
            .form-container { padding: 20px; }
            .form-group input, .form-group select { font-size: 14px; padding: 10px 12px; }
        }
        @media (max-width: 400px) {
            .form-container { padding: 15px; }
            .form-group input, .form-group select { font-size: 13px; padding: 8px 10px; }
            .btn-submit { font-size: 16px; padding: 12px; }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <div class="header-left">
                <h2>📝 ចុះឈ្មោះអ្នកប្រើប្រាស់</h2>
            </div>
            <div class="header-right">
                <span class="user-name">👤 {{ session.username }}</span>
                <a href="/dashboard" class="back-link">← ត្រលប់</a>
            </div>
        </div>
        {% if message %}
        <div class="message {{ message_type }}">{{ message }}</div>
        {% endif %}
        <div class="form-container">
            <h3>📋 បំពេញព័ត៌មាន</h3>
            <div class="sub-title">សូមបំពេញព័ត៌មានឲ្យបានត្រឹមត្រូវ</div>
            <form method="POST">
                <div class="form-group">
                    <label>ឈ្មោះអ្នកប្រើ <span class="required">*</span></label>
                    <input type="text" name="username" placeholder="ឧ: sok_sovan" required>
                    <div class="hint">ត្រូវមានយ៉ាងតិច 3 តួ</div>
                </div>
                <div class="form-row">
                    <div class="form-group">
                        <label>ពាក្យសម្ងាត់ <span class="required">*</span></label>
                        <input type="password" name="password" placeholder="បញ្ចូលពាក្យសម្ងាត់" required>
                        <div class="hint">ត្រូវមានយ៉ាងតិច 4 តួ</div>
                    </div>
                    <div class="form-group">
                        <label>បញ្ជាក់ពាក្យសម្ងាត់ <span class="required">*</span></label>
                        <input type="password" name="confirm_password" placeholder="បញ្ចូលម្តងទៀត" required>
                    </div>
                </div>
                <div class="form-group">
                    <label>ឈ្មោះពេញ <span class="required">*</span></label>
                    <input type="text" name="full_name" placeholder="ឧ: សុខ សុវណ្ណ" required>
                </div>
                <div class="form-row">
                    <div class="form-group">
                        <label>អ៊ីមែល</label>
                        <input type="email" name="email" placeholder="sok.sovan@email.com">
                    </div>
                    <div class="form-group">
                        <label>លេខទូរស័ព្ទ</label>
                        <input type="text" name="phone" placeholder="012 345 678">
                    </div>
                </div>
                <div class="form-group">
                    <label>តួនាទី</label>
                    <select name="role">
                        <option value="user">អ្នកប្រើប្រាស់</option>
                        <option value="admin">អ្នកគ្រប់គ្រង</option>
                        <option value="manager">អ្នកគ្រប់គ្រងផ្នែក</option>
                    </select>
                </div>
                <button type="submit" class="btn-submit">✅ ចុះឈ្មោះ</button>
            </form>
        </div>
        <div class="footer">© 2026 ប្រព័ន្ធគ្រប់គ្រងបុគ្គលិក | រក្សាសិទ្ធិគ្រប់យ៉ាង</div>
    </div>
</body>
</html>'''

CHANGE_PASSWORD_HTML = '''<!DOCTYPE html>
<html>
<head>
    <title>ប្តូរពាក្យសម្ងាត់</title>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: "Khmer OS", Arial, sans-serif; background: #f0f2f5; padding: 15px; min-height: 100vh; display: flex; justify-content: center; align-items: center; }
        .container { max-width: 500px; width: 100%; margin: 0 auto; }
        .header {
            background: linear-gradient(135deg, #1a73e8, #0d47a1);
            color: white;
            padding: 15px 20px;
            border-radius: 12px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 20px;
            flex-wrap: wrap;
            gap: 10px;
        }
        .header-left h2 { font-size: 18px; font-weight: 600; }
        .header-right { display: flex; align-items: center; gap: 10px; }
        .header-right .back-link {
            color: white;
            text-decoration: none;
            background: rgba(255,255,255,0.2);
            padding: 6px 16px;
            border-radius: 20px;
            font-size: 13px;
            transition: all 0.3s;
        }
        .header-right .back-link:hover { background: rgba(255,255,255,0.35); }
        .header-right .user-name {
            color: white;
            padding: 6px 14px;
            border-radius: 20px;
            font-size: 13px;
            background: rgba(255,255,255,0.15);
        }
        .form-container {
            background: white;
            padding: 30px;
            border-radius: 12px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.08);
        }
        .form-container h3 { color: #333; font-size: 20px; margin-bottom: 5px; }
        .form-container .sub-title { color: #888; font-size: 14px; margin-bottom: 20px; }
        .form-group { margin-bottom: 18px; }
        .form-group label { display: block; font-weight: 600; color: #555; font-size: 14px; margin-bottom: 5px; }
        .form-group label .required { color: #dc3545; }
        .form-group input {
            width: 100%;
            padding: 12px 14px;
            border: 2px solid #e8ecf1;
            border-radius: 10px;
            font-size: 15px;
            font-family: "Khmer OS", Arial, sans-serif;
            transition: all 0.3s;
            background: #fafafa;
        }
        .form-group input:focus {
            outline: none;
            border-color: #1a73e8;
            background: white;
            box-shadow: 0 0 0 3px rgba(26, 115, 232, 0.1);
        }
        .form-group .hint { font-size: 12px; color: #aaa; margin-top: 4px; }
        .btn-submit {
            width: 100%;
            padding: 14px;
            background: linear-gradient(135deg, #1a73e8, #0d47a1);
            color: white;
            border: none;
            border-radius: 10px;
            font-size: 18px;
            font-weight: 600;
            cursor: pointer;
            font-family: "Khmer OS", Arial, sans-serif;
            transition: all 0.3s;
            margin-top: 5px;
        }
        .btn-submit:hover { transform: translateY(-2px); box-shadow: 0 6px 20px rgba(26, 115, 232, 0.3); }
        .btn-submit:active { transform: scale(0.98); }
        .message {
            padding: 12px 16px;
            border-radius: 10px;
            margin-bottom: 20px;
            font-size: 14px;
        }
        .message.success { background: #d4edda; color: #155724; border-left: 4px solid #28a745; }
        .message.error { background: #f8d7da; color: #721c24; border-left: 4px solid #dc3545; }
        .message.warning { background: #fff3cd; color: #856404; border-left: 4px solid #ffc107; }
        .footer { text-align: center; padding: 20px 0 5px; color: #aaa; font-size: 12px; }
        @media (max-width: 600px) {
            .header { flex-direction: column; text-align: center; }
            .form-container { padding: 20px; }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <div class="header-left">
                <h2>🔑 ប្តូរពាក្យសម្ងាត់</h2>
            </div>
            <div class="header-right">
                <span class="user-name">👤 {{ session.username }}</span>
                <a href="/dashboard" class="back-link">← ត្រលប់</a>
            </div>
        </div>
        {% if message %}
        <div class="message {{ message_type }}">{{ message }}</div>
        {% endif %}
        <div class="form-container">
            <h3>🔐 ប្តូរពាក្យសម្ងាត់</h3>
            <div class="sub-title">សូមបញ្ចូលពាក្យសម្ងាត់បច្ចុប្បន្ន និងពាក្យសម្ងាត់ថ្មី</div>
            <form method="POST">
                <div class="form-group">
                    <label>ពាក្យសម្ងាត់បច្ចុប្បន្ន <span class="required">*</span></label>
                    <input type="password" name="current_password" placeholder="បញ្ចូលពាក្យសម្ងាត់បច្ចុប្បន្ន" required>
                </div>
                <div class="form-group">
                    <label>ពាក្យសម្ងាត់ថ្មី <span class="required">*</span></label>
                    <input type="password" name="new_password" placeholder="បញ្ចូលពាក្យសម្ងាត់ថ្មី" required>
                    <div class="hint">ត្រូវមានយ៉ាងតិច 4 តួ</div>
                </div>
                <div class="form-group">
                    <label>បញ្ជាក់ពាក្យសម្ងាត់ថ្មី <span class="required">*</span></label>
                    <input type="password" name="confirm_password" placeholder="បញ្ចូលម្តងទៀត" required>
                </div>
                <button type="submit" class="btn-submit">✅ ប្តូរពាក្យសម្ងាត់</button>
            </form>
        </div>
        <div class="footer">
            © 2026 ប្រព័ន្ធគ្រប់គ្រងបុគ្គលិក | រក្សាសិទ្ធិគ្រប់យ៉ាង
        </div>
    </div>
</body>
</html>'''

USER_MANAGEMENT_HTML = '''<!DOCTYPE html>
<html>
<head>
    <title>គ្រប់គ្រងអ្នកប្រើប្រាស់</title>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif; background: #f0f2f5; padding: 15px; min-height: 100vh; }
        .container { max-width: 1200px; margin: 0 auto; }
        .header {
            background: linear-gradient(135deg, #1a73e8, #0d47a1);
            color: white;
            padding: 12px 20px;
            border-radius: 12px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            box-shadow: 0 4px 15px rgba(26, 115, 232, 0.3);
            margin-bottom: 20px;
            flex-wrap: wrap;
            gap: 10px;
        }
        .header-left h2 { font-size: 18px; font-weight: 600; }
        .header-right { display: flex; align-items: center; gap: 10px; }
        .header-right .back-link {
            color: white;
            text-decoration: none;
            background: rgba(255,255,255,0.2);
            padding: 6px 16px;
            border-radius: 20px;
            font-size: 13px;
            transition: all 0.3s;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
        }
        .header-right .back-link:hover { background: rgba(255,255,255,0.35); }
        .header-right .user-name {
            color: white;
            padding: 6px 14px;
            border-radius: 20px;
            font-size: 13px;
            background: rgba(255,255,255,0.15);
        }
        .message {
            padding: 12px 16px;
            border-radius: 10px;
            margin-bottom: 15px;
            font-size: 14px;
        }
        .message.success { background: #d4edda; color: #155724; border-left: 4px solid #28a745; }
        .message.error { background: #f8d7da; color: #721c24; border-left: 4px solid #dc3545; }
        .message.warning { background: #fff3cd; color: #856404; border-left: 4px solid #ffc107; }
        .table-container {
            background: white;
            border-radius: 12px;
            padding: 20px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.06);
            overflow-x: auto;
        }
        .table-container .table-title {
            font-size: 18px;
            font-weight: 600;
            color: #333;
            margin-bottom: 15px;
            display: flex;
            align-items: center;
            gap: 10px;
        }
        .table-container .table-title .badge-count {
            background: #1a73e8;
            color: white;
            padding: 2px 12px;
            border-radius: 20px;
            font-size: 14px;
        }
        table { width: 100%; border-collapse: collapse; }
        table thead th {
            background: #f8f9fa;
            color: #555;
            padding: 12px 15px;
            text-align: left;
            font-size: 13px;
            font-weight: 600;
            border-bottom: 2px solid #e8ecf1;
        }
        table tbody td {
            padding: 12px 15px;
            border-bottom: 1px solid #f0f2f5;
            color: #333;
            font-size: 14px;
        }
        table tbody tr:hover { background: #f8f9fa; }
        .btn-action {
            padding: 5px 12px;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            font-size: 12px;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
            margin: 2px;
        }
        .btn-edit { background: #1a73e8; color: white; }
        .btn-edit:hover { background: #1557b0; }
        .btn-delete { background: #dc3545; color: white; }
        .btn-delete:hover { background: #b02a37; }
        .btn-add {
            background: #34a853;
            color: white;
            padding: 8px 20px;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            font-size: 14px;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
            margin-bottom: 15px;
        }
        .btn-add:hover { background: #1e7e34; }
        .btn-setting {
            background: #ff9800;
            color: white;
            padding: 5px 12px;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            font-size: 12px;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
            margin: 2px;
        }
        .btn-setting:hover { background: #e68900; }
        .btn-setting.active {
            background: #4caf50;
        }
        .role-badge {
            display: inline-block;
            padding: 3px 12px;
            border-radius: 12px;
            font-size: 12px;
            font-weight: 600;
        }
        .role-admin { background: #ffebee; color: #c62828; }
        .role-manager { background: #fff3e0; color: #e65100; }
        .role-user { background: #e3f2fd; color: #0d47a1; }
        .deadline-badge {
            display: inline-block;
            padding: 2px 10px;
            border-radius: 12px;
            font-size: 11px;
            font-weight: 600;
            background: #ff9800;
            color: white;
        }
        .deadline-badge.active {
            background: #4caf50;
        }
        .deadline-badge.inactive {
            background: #9e9e9e;
        }
        .user-lock-badge {
            display: inline-block;
            padding: 2px 10px;
            border-radius: 12px;
            font-size: 11px;
            font-weight: 600;
            background: #dc3545;
            color: white;
        }
        .user-lock-badge.unlocked {
            background: #4caf50;
        }
        .modal {
            display: none;
            position: fixed;
            top: 0; left: 0; width: 100%; height: 100%;
            background: rgba(0,0,0,0.5);
            justify-content: center;
            align-items: center;
            z-index: 999;
        }
        .modal.show { display: flex; }
        .modal-content {
            background: white;
            padding: 30px;
            border-radius: 16px;
            max-width: 500px;
            width: 90%;
            box-shadow: 0 10px 40px rgba(0,0,0,0.3);
            max-height: 90vh;
            overflow-y: auto;
        }
        .modal-content h3 { color: #333; margin-bottom: 15px; }
        .modal-content label { display: block; font-weight: 600; color: #555; font-size: 14px; margin-bottom: 5px; }
        .modal-content input, .modal-content select {
            width: 100%;
            padding: 10px 14px;
            border: 2px solid #e8ecf1;
            border-radius: 10px;
            font-size: 14px;
            margin-bottom: 12px;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
        }
        .modal-content input:focus, .modal-content select:focus {
            outline: none;
            border-color: #1a73e8;
        }
        .modal-content .btn-group {
            display: flex;
            gap: 10px;
            margin-top: 10px;
        }
        .modal-content .btn-group button {
            flex: 1;
            padding: 12px;
            border: none;
            border-radius: 10px;
            cursor: pointer;
            font-size: 15px;
            font-family: 'Khmer OS', 'Khmer OS Muol', 'Arial', sans-serif;
            font-weight: 500;
        }
        .btn-save { background: #1a73e8; color: white; }
        .btn-save:hover { background: #1557b0; }
        .btn-cancel { background: #e8ecf1; color: #333; }
        .btn-cancel:hover { background: #d5d8dd; }
        .toggle-label {
            display: flex;
            align-items: center;
            gap: 10px;
            cursor: pointer;
            padding: 10px 0;
        }
        .toggle-label input[type="checkbox"] {
            width: 50px;
            height: 26px;
            appearance: none;
            background: #ccc;
            border-radius: 13px;
            position: relative;
            cursor: pointer;
            transition: 0.3s;
            margin: 0;
        }
        .toggle-label input[type="checkbox"]:checked {
            background: #4caf50;
        }
        .toggle-label input[type="checkbox"]::before {
            content: '';
            position: absolute;
            width: 22px;
            height: 22px;
            background: white;
            border-radius: 50%;
            top: 2px;
            left: 2px;
            transition: 0.3s;
        }
        .toggle-label input[type="checkbox"]:checked::before {
            left: 26px;
        }
        .footer { text-align: center; padding: 20px 0 5px; color: #aaa; font-size: 12px; }
        @media (max-width: 768px) {
            .header { flex-direction: column; text-align: center; }
            .header-right { justify-content: center; }
            table thead th, table tbody td { font-size: 12px; padding: 8px 10px; }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <div class="header-left">
                <h2>👤 គ្រប់គ្រងអ្នកប្រើប្រាស់</h2>
            </div>
            <div class="header-right">
                <span class="user-name">👤 {{ session.username }}</span>
                <a href="/dashboard" class="back-link">← ត្រលប់</a>
            </div>
        </div>
        {% if message %}
        <div class="message {{ message_type }}">{{ message }}</div>
        {% endif %}
        <div class="table-container">
            <div class="table-title">
                📋 បញ្ជីអ្នកប្រើប្រាស់
                <span class="badge-count">{{ users|length }}</span>
                <button class="btn-add" onclick="openAddUserModal()">➕ បន្ថែមអ្នកប្រើ</button>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>ល.រ</th>
                        <th>ឈ្មោះអ្នកប្រើ</th>
                        <th>ឈ្មោះពេញ</th>
                        <th>អ៊ីមែល</th>
                        <th>ទូរស័ព្ទ</th>
                        <th>តួនាទី</th>
                        <th>ម៉ោងកំណត់</th>
                        <th>ស្ថានភាព</th>
                        <th>សកម្មភាព</th>
                    </tr>
                </thead>
                <tbody>
                    {% for user in users %}
                    {% set setting = settings.get(user.id) %}
                    {% set lock = user_locks.get(user.id) %}
                    <tr>
                        <td>{{ loop.index }}</td>
                        <td>{{ user.username }}</td>
                        <td>{{ user.full_name }}</td>
                        <td>{{ user.email or '-' }}</td>
                        <td>{{ user.phone or '-' }}</td>
                        <td>
                            <span class="role-badge
                                {% if user.role == 'admin' %}role-admin
                                {% elif user.role == 'manager' %}role-manager
                                {% else %}role-user{% endif %}">
                                {{ user.role }}
                            </span>
                        </td>
                        <td>
                            {% if setting and setting.is_active == 1 %}
                                <span class="deadline-badge active">⏰ {{ setting.check_in_deadline or 'មិនកំណត់' }}</span>
                            {% elif setting %}
                                <span class="deadline-badge inactive">⏰ {{ setting.check_in_deadline or 'មិនកំណត់' }}</span>
                            {% else %}
                                <span class="deadline-badge inactive">មិនកំណត់</span>
                            {% endif %}
                        </td>
                        <td>
                            {% if lock and lock.is_locked == 1 %}
                                <span class="user-lock-badge">🔒 បិទ</span>
                                {% if lock.auto_unlock_time %}
                                <br><span style="font-size:10px;color:#888;">បើកនៅ {{ lock.auto_unlock_time }}</span>
                                {% endif %}
                            {% else %}
                                <span class="user-lock-badge unlocked">🔓 បើក</span>
                            {% endif %}
                        </td>
                        <td>
                            <button class="btn-action btn-edit" onclick="openEditUserModal({{ user.id }})">✏️ កែ</button>
                            <button class="btn-setting" onclick="openAttendanceSettingModal({{ user.id }}, '{{ user.username }}')">⚙️ ម៉ោង</button>
                            <button class="btn-action" onclick="toggleUserLock({{ user.id }}, '{{ user.username }}', {% if lock and lock.is_locked == 1 %}true{% else %}false{% endif %})"
                                    style="background:{% if lock and lock.is_locked == 1 %}#dc3545;color:white{% else %}#4caf50;color:white{% endif %};padding:5px 12px;border:none;border-radius:6px;cursor:pointer;font-size:12px;font-family:'Khmer OS','Khmer OS Muol','Arial',sans-serif;margin:2px;">
                                {% if lock and lock.is_locked == 1 %}🔓 បើក{% else %}🔒 បិទ{% endif %}
                            </button>
                            <button class="btn-action" onclick="resetPassword({{ user.id }}, '{{ user.username }}')" style="background:#fbbc04;color:#333;padding:5px 12px;border:none;border-radius:6px;cursor:pointer;font-size:12px;font-family:'Khmer OS','Khmer OS Muol','Arial',sans-serif;margin:2px;">🔑 ពាក្យសម្ងាត់</button>
                            {% if user.id != session.user_id %}
                            <button class="btn-action btn-delete" onclick="deleteUser({{ user.id }}, '{{ user.username }}')">🗑️ លុប</button>
                            {% else %}
                            <span style="font-size:11px;color:#999;">(អ្នកចូល)</span>
                            {% endif %}
                        </td>
                    </tr>
                    {% endfor %}
                    {% if not users %}
                    <tr>
                        <td colspan="9" style="text-align:center;padding:30px;color:#aaa;">📭 មិនមានអ្នកប្រើ</td>
                    </tr>
                    {% endif %}
                </tbody>
            </table>
        </div>
        <div class="footer">© 2026 ប្រព័ន្ធគ្រប់គ្រងបុគ្គលិក | រក្សាសិទ្ធិគ្រប់យ៉ាង</div>
    </div>

    <!-- Edit User Modal -->
    <div id="editUserModal" class="modal">
        <div class="modal-content">
            <h3>✏️ កែប្រែអ្នកប្រើ</h3>
            <form id="editUserForm">
                <input type="hidden" id="editUserId">
                <label>ឈ្មោះអ្នកប្រើ</label>
                <input type="text" id="editUsername" required>
                <label>ឈ្មោះពេញ</label>
                <input type="text" id="editFullName" required>
                <label>អ៊ីមែល</label>
                <input type="email" id="editEmail">
                <label>ទូរស័ព្ទ</label>
                <input type="text" id="editPhone">
                <label>តួនាទី</label>
                <select id="editRole">
                    <option value="user">អ្នកប្រើប្រាស់</option>
                    <option value="admin">អ្នកគ្រប់គ្រង</option>
                    <option value="manager">អ្នកគ្រប់គ្រងផ្នែក</option>
                </select>
                <div class="btn-group">
                    <button type="button" class="btn-save" onclick="submitEditUser()">💾 រក្សាទុក</button>
                    <button type="button" class="btn-cancel" onclick="closeModal('editUserModal')">បោះបង់</button>
                </div>
            </form>
        </div>
    </div>

    <!-- Add User Modal -->
    <div id="addUserModal" class="modal">
        <div class="modal-content">
            <h3>➕ បន្ថែមអ្នកប្រើថ្មី</h3>
            <form id="addUserForm">
                <label>ឈ្មោះអ្នកប្រើ <span style="color:red;">*</span></label>
                <input type="text" id="addUsername" required>
                <label>ពាក្យសម្ងាត់ <span style="color:red;">*</span></label>
                <input type="password" id="addPassword" required>
                <label>ឈ្មោះពេញ <span style="color:red;">*</span></label>
                <input type="text" id="addFullName" required>
                <label>អ៊ីមែល</label>
                <input type="email" id="addEmail">
                <label>ទូរស័ព្ទ</label>
                <input type="text" id="addPhone">
                <label>តួនាទី</label>
                <select id="addRole">
                    <option value="user">អ្នកប្រើប្រាស់</option>
                    <option value="admin">អ្នកគ្រប់គ្រង</option>
                    <option value="manager">អ្នកគ្រប់គ្រងផ្នែក</option>
                </select>
                <div class="btn-group">
                    <button type="button" class="btn-save" onclick="submitAddUser()">✅ បន្ថែម</button>
                    <button type="button" class="btn-cancel" onclick="closeModal('addUserModal')">បោះបង់</button>
                </div>
            </form>
        </div>
    </div>

    <!-- Attendance Setting Modal -->
    <div id="attendanceSettingModal" class="modal">
        <div class="modal-content">
            <h3>⚙️ កំណត់ម៉ោងចូលធ្វើការ</h3>
            <p style="color:#666;font-size:14px;margin-bottom:15px;">
                កំណត់ម៉ោងចុងក្រោយដែលអ្នកប្រើអាចចុចចូលធ្វើការបាន។
                ប្រសិនបើអ្នកប្រើចុចចូលធ្វើការលើសពីម៉ោងកំណត់ ប្រព័ន្ធនឹងបដិសេធ។
            </p>
            <form id="attendanceSettingForm">
                <input type="hidden" id="settingUserId">
                <label>ឈ្មោះអ្នកប្រើ</label>
                <input type="text" id="settingUsername" disabled style="background:#f5f5f5;">

                <div class="toggle-label">
                    <span>បើក/បិទ ការកំណត់</span>
                    <input type="checkbox" id="settingIsActive">
                </div>

                <label>ម៉ោងកំណត់ (HH:MM)</label>
                <input type="time" id="settingDeadline" step="60">
                <div style="font-size:12px;color:#888;margin-top:-8px;margin-bottom:10px;">
                    ឧទាហរណ៍: 08:00 មានន័យថាអ្នកប្រើត្រូវចូលធ្វើការមុនម៉ោង 08:00
                </div>

                <div class="btn-group">
                    <button type="button" class="btn-save" onclick="submitAttendanceSetting()">💾 រក្សាទុក</button>
                    <button type="button" class="btn-cancel" onclick="closeModal('attendanceSettingModal')">បោះបង់</button>
                </div>
            </form>
        </div>
    </div>

    <script>
        function openEditUserModal(userId) {
            fetch('/get_user/' + userId)
            .then(res => res.json())
            .then(data => {
                document.getElementById('editUserId').value = data.id;
                document.getElementById('editUsername').value = data.username;
                document.getElementById('editFullName').value = data.full_name;
                document.getElementById('editEmail').value = data.email || '';
                document.getElementById('editPhone').value = data.phone || '';
                document.getElementById('editRole').value = data.role || 'user';
                openModal('editUserModal');
            })
            .catch(err => alert('មានបញ្ហា: ' + err));
        }

        function submitEditUser() {
            var id = document.getElementById('editUserId').value;
            var data = {
                username: document.getElementById('editUsername').value.trim(),
                full_name: document.getElementById('editFullName').value.trim(),
                email: document.getElementById('editEmail').value.trim(),
                phone: document.getElementById('editPhone').value.trim(),
                role: document.getElementById('editRole').value
            };
            if (!data.username || !data.full_name) {
                alert('សូមបំពេញឈ្មោះអ្នកប្រើ និងឈ្មោះពេញ!');
                return;
            }
            fetch('/update_user/' + id, {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify(data)
            })
            .then(res => res.json())
            .then(result => {
                alert(result.message);
                if (result.success) {
                    closeModal('editUserModal');
                    location.reload();
                }
            })
            .catch(err => alert('មានបញ្ហា: ' + err));
        }

        function openAddUserModal() {
            document.getElementById('addUsername').value = '';
            document.getElementById('addPassword').value = '';
            document.getElementById('addFullName').value = '';
            document.getElementById('addEmail').value = '';
            document.getElementById('addPhone').value = '';
            document.getElementById('addRole').value = 'user';
            openModal('addUserModal');
        }

        function submitAddUser() {
            var data = {
                username: document.getElementById('addUsername').value.trim(),
                password: document.getElementById('addPassword').value.trim(),
                full_name: document.getElementById('addFullName').value.trim(),
                email: document.getElementById('addEmail').value.trim(),
                phone: document.getElementById('addPhone').value.trim(),
                role: document.getElementById('addRole').value
            };
            if (!data.username || !data.password || !data.full_name) {
                alert('សូមបំពេញព័ត៌មានឲ្យបានពេញលេញ!');
                return;
            }
            if (data.password.length < 4) {
                alert('ពាក្យសម្ងាត់ត្រូវមានយ៉ាងតិច 4 តួ!');
                return;
            }
            fetch('/add_user', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify(data)
            })
            .then(res => res.json())
            .then(result => {
                alert(result.message);
                if (result.success) {
                    closeModal('addUserModal');
                    location.reload();
                }
            })
            .catch(err => alert('មានបញ្ហា: ' + err));
        }

        function deleteUser(userId, username) {
            if (!confirm('តើអ្នកចង់លុបអ្នកប្រើ "' + username + '" មែនទេ? (ទិន្នន័យទាំងអស់របស់គាត់នឹងត្រូវលុប!)')) return;
            fetch('/delete_user/' + userId, { method: 'POST' })
            .then(res => res.json())
            .then(result => {
                alert(result.message);
                if (result.success) location.reload();
            })
            .catch(err => alert('មានបញ្ហា: ' + err));
        }

        function resetPassword(userId, username) {
            var newPassword = prompt('សូមបញ្ចូលពាក្យសម្ងាត់ថ្មីសម្រាប់ "' + username + '" (យ៉ាងតិច 4 តួ):');
            if (newPassword === null) return;
            if (!newPassword || newPassword.length < 4) {
                alert('ពាក្យសម្ងាត់ត្រូវមានយ៉ាងតិច 4 តួ!');
                return;
            }
            fetch('/admin_reset_password/' + userId, {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ new_password: newPassword })
            })
            .then(response => response.json())
            .then(result => {
                alert(result.message);
                if (result.success) location.reload();
            })
            .catch(err => {
                console.error('Error:', err);
                alert('មានបញ្ហា: ' + err);
            });
        }

        // ===== USER LOCK FUNCTIONS =====
        function toggleUserLock(userId, username, isLocked) {
            var action = isLocked ? 'បើក' : 'បិទ';
            var confirmMsg = 'តើអ្នកចង់' + action + 'ការចូលធ្វើការរបស់ "' + username + '" មែនទេ?';

            if (!confirm(confirmMsg)) return;

            var autoUnlockTime = null;
            if (!isLocked) {
                autoUnlockTime = prompt('សូមបញ្ចូលពេលវេលាបើកដោយស្វ័យប្រវត្តិ (HH:MM) ឬទុកចោលសម្រាប់មិនកំណត់:', '');
                if (autoUnlockTime !== null && autoUnlockTime !== '') {
                    var timeRegex = /^([0-1][0-9]|2[0-3]):[0-5][0-9]$/;
                    if (!timeRegex.test(autoUnlockTime)) {
                        alert('ទ្រង់ទ្រាយម៉ោងមិនត្រឹមត្រូវ! សូមប្រើ HH:MM (ឧទាហរណ៍: 08:00)');
                        return;
                    }
                }
                if (autoUnlockTime === '') {
                    autoUnlockTime = null;
                }
            }

            var newState = isLocked ? 0 : 1;

            fetch('/toggle_user_lock', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({
                    user_id: userId,
                    lock_state: newState,
                    auto_unlock_time: autoUnlockTime
                })
            })
            .then(res => res.json())
            .then(result => {
                alert(result.message);
                if (result.success) location.reload();
            })
            .catch(err => {
                console.error('Error:', err);
                alert('មានបញ្ហា: ' + err);
            });
        }

        // ===== ATTENDANCE SETTING FUNCTIONS =====
        function openAttendanceSettingModal(userId, username) {
            document.getElementById('settingUserId').value = userId;
            document.getElementById('settingUsername').value = username;

            fetch('/get_attendance_setting/' + userId)
            .then(res => res.json())
            .then(data => {
                if (data.check_in_deadline) {
                    document.getElementById('settingDeadline').value = data.check_in_deadline;
                } else {
                    document.getElementById('settingDeadline').value = '';
                }
                document.getElementById('settingIsActive').checked = data.is_active == 1;
                openModal('attendanceSettingModal');
            })
            .catch(err => alert('មានបញ្ហា: ' + err));
        }

        function submitAttendanceSetting() {
            var userId = document.getElementById('settingUserId').value;
            var deadline = document.getElementById('settingDeadline').value;
            var isActive = document.getElementById('settingIsActive').checked ? 1 : 0;

            if (isActive && !deadline) {
                alert('សូមបញ្ចូលម៉ោងកំណត់!');
                return;
            }

            fetch('/save_attendance_setting', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({
                    user_id: parseInt(userId),
                    check_in_deadline: deadline,
                    is_active: isActive
                })
            })
            .then(res => res.json())
            .then(result => {
                alert(result.message);
                if (result.success) {
                    closeModal('attendanceSettingModal');
                    location.reload();
                }
            })
            .catch(err => alert('មានបញ្ហា: ' + err));
        }

        function openModal(id) { document.getElementById(id).classList.add('show'); }
        function closeModal(id) { document.getElementById(id).classList.remove('show'); }

        document.querySelectorAll('.modal').forEach(function(modal) {
            modal.addEventListener('click', function(e) {
                if (e.target === this) this.classList.remove('show');
            });
        });
    </script>
</body>
</html>'''

REPORT_HTML = '''<!DOCTYPE html>
<html>
<head>
    <title>របាយការណ៍</title>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: "Khmer OS", Arial, sans-serif; background: #f0f2f5; padding: 15px; }
        .container { max-width: 1200px; margin: 0 auto; }
        .header {
            background: linear-gradient(135deg, #1a73e8, #0d47a1);
            color: white;
            padding: 12px 20px;
            border-radius: 12px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 20px;
            flex-wrap: wrap;
            gap: 10px;
        }
        .header-left h2 { font-size: 18px; font-weight: 600; }
        .header-right { display: flex; align-items: center; gap: 10px; flex-wrap: wrap; }
        .header-right .back-link {
            color: white;
            text-decoration: none;
            background: rgba(255,255,255,0.2);
            padding: 6px 16px;
            border-radius: 20px;
            font-size: 13px;
            transition: all 0.3s;
        }
        .header-right .back-link:hover { background: rgba(255,255,255,0.35); }
        .header-right .user-name {
            color: white;
            padding: 6px 14px;
            border-radius: 20px;
            font-size: 13px;
            background: rgba(255,255,255,0.15);
        }
        .filter-box {
            background: white;
            padding: 20px;
            border-radius: 12px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.06);
            margin-bottom: 20px;
            display: flex;
            gap: 15px;
            flex-wrap: wrap;
            align-items: end;
        }
        .filter-box .form-group { flex: 1; min-width: 150px; }
        .filter-box label { display: block; font-weight: 600; color: #555; font-size: 14px; margin-bottom: 5px; }
        .filter-box input, .filter-box select {
            width: 100%;
            padding: 10px 14px;
            border: 2px solid #e8ecf1;
            border-radius: 10px;
            font-size: 14px;
            font-family: "Khmer OS", Arial, sans-serif;
        }
        .filter-box input:focus, .filter-box select:focus { outline: none; border-color: #1a73e8; }
        .filter-box .btn-filter {
            padding: 10px 30px;
            background: #1a73e8;
            color: white;
            border: none;
            border-radius: 10px;
            cursor: pointer;
            font-size: 16px;
            font-family: "Khmer OS", Arial, sans-serif;
            font-weight: 500;
            min-width: 120px;
            transition: all 0.3s;
        }
        .filter-box .btn-filter:hover { background: #1557b0; }
        .filter-box .btn-filter.monthly { background: #34a853; }
        .filter-box .btn-filter.monthly:hover { background: #1e7e34; }
        .filter-box .btn-excel {
            padding: 10px 30px;
            background: #fbbc04;
            color: #333;
            border: none;
            border-radius: 10px;
            cursor: pointer;
            font-size: 16px;
            font-family: "Khmer OS", Arial, sans-serif;
            font-weight: 500;
            min-width: 120px;
            transition: all 0.3s;
        }
        .filter-box .btn-excel:hover { background: #e5a800; }
        .tab-buttons {
            display: flex;
            gap: 10px;
            margin-bottom: 20px;
            flex-wrap: wrap;
        }
        .tab-button {
            padding: 12px 25px;
            border: none;
            border-radius: 10px;
            cursor: pointer;
            font-size: 16px;
            font-family: "Khmer OS", Arial, sans-serif;
            font-weight: 500;
            background: #e8ecf1;
            color: #555;
            transition: all 0.3s;
        }
        .tab-button.active {
            background: #1a73e8;
            color: white;
        }
        .tab-button:hover:not(.active) {
            background: #d5d8dd;
        }
        .tab-content { display: none; }
        .tab-content.active { display: block; }
        .table-container {
            background: white;
            border-radius: 12px;
            padding: 20px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.06);
            overflow-x: auto;
            margin-bottom: 20px;
        }
        .table-container .table-title {
            font-size: 18px;
            font-weight: 600;
            color: #333;
            margin-bottom: 15px;
            display: flex;
            align-items: center;
            gap: 10px;
        }
        .table-container .table-title .badge-count {
            background: #1a73e8;
            color: white;
            padding: 2px 12px;
            border-radius: 20px;
            font-size: 14px;
        }
        table { width: 100%; border-collapse: collapse; }
        table thead th {
            background: #f8f9fa;
            color: #555;
            padding: 12px 15px;
            text-align: left;
            font-size: 13px;
            font-weight: 600;
            border-bottom: 2px solid #e8ecf1;
        }
        table tbody td {
            padding: 12px 15px;
            border-bottom: 1px solid #f0f2f5;
            color: #333;
            font-size: 14px;
        }
        table tbody tr:hover { background: #f8f9fa; }
        .rank-badge {
            display: inline-block;
            border-radius: 50%;
            width: 28px;
            height: 28px;
            text-align: center;
            line-height: 28px;
            font-size: 14px;
            font-weight: 600;
        }
        .rank-1 { background: #ffd700; color: #333; }
        .rank-2 { background: #c0c0c0; color: #333; }
        .rank-3 { background: #cd7f32; color: white; }
        .rank-other { background: #e8ecf1; color: #555; }
        .shift-badge {
            display: inline-block;
            padding: 2px 10px;
            border-radius: 12px;
            font-size: 11px;
            font-weight: 600;
        }
        .shift-1 { background: #e3f2fd; color: #0d47a1; }
        .shift-2 { background: #fff3e0; color: #e65100; }
        .shift-3 { background: #f3e5f5; color: #4a148c; }
        .shift-leave { background: #fff3cd; color: #856404; }
        .shift-mission { background: #d1ecf1; color: #0c5460; }
        .status-badge {
            display: inline-block;
            padding: 4px 14px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: 600;
        }
        .status-badge.present { background: #e8f5e9; color: #2e7d32; }
        .status-badge.outside { background: #ffebee; color: #c62828; }
        .status-badge.working { background: #fff3cd; color: #856404; }
        .status-badge.leave { background: #fff3cd; color: #856404; }
        .status-badge.mission { background: #d1ecf1; color: #0c5460; }
        .status-badge.approved { background: #d4edda; color: #155724; }
        .status-badge.pending { background: #fff3cd; color: #856404; }
        .status-badge.rejected { background: #f8d7da; color: #721c24; }
        .stats-cards {
            display: grid;
            grid-template-columns: repeat(4, 1fr);
            gap: 15px;
            margin-bottom: 20px;
        }
        .stat-card {
            background: white;
            padding: 18px 15px;
            border-radius: 12px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.06);
            text-align: center;
        }
        .stat-card .stat-icon { font-size: 28px; display: block; margin-bottom: 4px; }
        .stat-card .number { font-size: 30px; font-weight: 700; color: #1a73e8; }
        .stat-card .number.green { color: #34a853; }
        .stat-card .number.orange { color: #fbbc04; }
        .stat-card .number.purple { color: #7c3aed; }
        .stat-card .label { font-size: 13px; color: #888; margin-top: 2px; }
        .footer { text-align: center; padding: 20px 0 5px; color: #aaa; font-size: 12px; }
        .info-text { font-size: 12px; color: #666; margin-top: 2px; }
        @media (max-width: 768px) {
            .header { flex-direction: column; text-align: center; }
            .header-right { justify-content: center; }
            .filter-box { flex-direction: column; }
            .filter-box .form-group { min-width: 100%; }
            table thead th, table tbody td { font-size: 12px; padding: 8px 10px; }
            .tab-buttons { flex-wrap: wrap; }
            .tab-button { flex: 1; min-width: 100px; text-align: center; padding: 10px 15px; font-size: 14px; }
            .stats-cards { grid-template-columns: repeat(2, 1fr); }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <div class="header-left"><h2>📊 របាយការណ៍</h2></div>
            <div class="header-right">
                <span class="user-name">👤 {{ session.username }}</span>
                <a href="/dashboard" class="back-link">← ត្រលប់</a>
            </div>
        </div>

        <div class="filter-box">
            <div class="form-group">
                <label>ចាប់ពីថ្ងៃ</label>
                <input type="date" id="startDate" value="{{ start_date }}">
            </div>
            <div class="form-group">
                <label>ដល់ថ្ងៃ</label>
                <input type="date" id="endDate" value="{{ end_date }}">
            </div>
            <div class="form-group">
                <label>ជ្រើសរើសខែ</label>
                <select id="monthSelect">
                    <option value="">-- ជ្រើសរើសខែ --</option>
                    <option value="1">មករា</option>
                    <option value="2">កុម្ភៈ</option>
                    <option value="3">មីនា</option>
                    <option value="4">មេសា</option>
                    <option value="5">ឧសភា</option>
                    <option value="6">មិថុនា</option>
                    <option value="7">កក្កដា</option>
                    <option value="8">សីហា</option>
                    <option value="9">កញ្ញា</option>
                    <option value="10">តុលា</option>
                    <option value="11">វិច្ឆិកា</option>
                    <option value="12">ធ្នូ</option>
                </select>
            </div>
            <div class="form-group">
                <label>ឆ្នាំ</label>
                <input type="number" id="yearSelect" value="{{ current_year }}" min="2020" max="2030">
            </div>
            <button class="btn-filter" onclick="filterReport()">🔍 ស្វែងរក</button>
            <button class="btn-filter monthly" onclick="loadMonthly()">📅 ប្រចាំខែ</button>
            <button class="btn-excel" onclick="exportExcel()">📥 Excel</button>
        </div>

        <div class="tab-buttons">
            <button class="tab-button active" onclick="switchTab('daily')">📋 ប្រវត្តិការងារ</button>
            <button class="tab-button" onclick="switchTab('summary')">📊 សង្ខេបប្រចាំខែ</button>
        </div>

        <div id="tab-daily" class="tab-content active">
            <div class="table-container">
                <div class="table-title">
                    📋 ប្រវត្តិការងារ
                    <span class="badge-count">{{ daily|length }}</span>
                </div>
                <table>
                    <thead>
                        <tr>
                            <th>ល.រ</th>
                            <th>ឈ្មោះបុគ្គលិក</th>
                            <th>កាលបរិច្ឆេទ</th>
                            <th>ប្រភេទ</th>
                            <th>វគ្គ</th>
                            <th>ម៉ោងចូល</th>
                            <th>ម៉ោងចេញ</th>
                            <th>ចំនួន</th>
                            <th>ស្ថានភាព/ព័ត៌មាន</th>
                        </tr>
                    </thead>
                    <tbody>
    {% for item in daily %}
    <tr>
        <td>{{ loop.index }}</td>
        <td>{{ item.full_name }}</td>
        <td>{{ item.date }}</td>
        <td>
            {% if item.type == 'attendance' %}
                <span class="status-badge present">វត្តមាន</span>
            {% elif item.type == 'leave' %}
                <span class="status-badge leave">ច្បាប់</span>
            {% elif item.type == 'mission' %}
                <span class="status-badge mission">បេសកម្ម</span>
            {% endif %}
        </td>
        <td>
            {% if item.type == 'attendance' %}
                <span class="shift-badge shift-{{ item.shift }}">វគ្គ {{ item.shift }}</span>
            {% elif item.type == 'leave' %}
                <span class="shift-badge shift-leave">📋 ច្បាប់</span>
            {% elif item.type == 'mission' %}
                <span class="shift-badge shift-mission">🚗 បេសកម្ម</span>
            {% endif %}
        </td>
        <td>
            {% if item.type == 'attendance' %}
                {% if item.check_in %}
                    {% if ' ' in item.check_in %}
                        {{ item.check_in.split(' ')[1][:5] }}
                    {% else %}
                        {{ item.check_in[:5] }}
                    {% endif %}
                {% else %}
                    -
                {% endif %}
            {% else %}
                {{ item.start_date or '' }}
            {% endif %}
        </td>
        <td>
            {% if item.type == 'attendance' %}
                {% if item.check_out %}
                    {% if ' ' in item.check_out %}
                        {{ item.check_out.split(' ')[1][:5] }}
                    {% else %}
                        {{ item.check_out[:5] }}
                    {% endif %}
                {% else %}
                    -
                {% endif %}
            {% else %}
                {{ item.end_date or '' }}
            {% endif %}
        </td>
        <td>
            {% if item.type == 'attendance' %}
                {% if item.total_hours %}
                    {% set hours = item.total_hours|int %}
                    {% set minutes = ((item.total_hours - hours) * 60)|int %}
                    {{ '%02d' % hours }}:{{ '%02d' % minutes }}
                {% else %}
                    -
                {% endif %}
            {% elif item.type == 'leave' %}
                {{ item.days }} ថ្ងៃ
            {% elif item.type == 'mission' %}
                {{ item.days }} ថ្ងៃ
            {% endif %}
        </td>
        <td>
            {% if item.type == 'attendance' %}
                {% if item.status == 'បានបិទ' %}
                    <span class="status-badge present">បានបិទ</span>
                {% elif item.status == 'កំពុងធ្វើការ' %}
                    <span class="status-badge working">កំពុងធ្វើការ</span>
                {% else %}
                    <span class="status-badge outside">មិនទាន់ចូល</span>
                {% endif %}
            {% elif item.type == 'leave' %}
                <span class="status-badge approved">✅ Approved</span>
                <div class="info-text">មូលហេតុ: {{ item.reason or 'មិនបានបញ្ជាក់' }}</div>
            {% elif item.type == 'mission' %}
                <span class="status-badge approved">✅ Approved</span>
                <div class="info-text">ទីតាំង: {{ item.destination or 'មិនបានបញ្ជាក់' }}</div>
            {% endif %}
        </td>
    </tr>
    {% endfor %}
    {% if not daily %}
    <tr>
        <td colspan="9" style="text-align:center;padding:30px;color:#aaa;">📭 មិនមានទិន្នន័យ</td>
    </tr>
    {% endif %}
</tbody>
                </table>
            </div>
        </div>

        <div id="tab-summary" class="tab-content">
            <div class="table-container">
                <div class="table-title">
                    📊 របាយការណ៍សង្ខេបប្រចាំខែ
                    <span class="badge-count">{{ summary|length }}</span>
                </div>
                <table>
                    <thead>
                        <tr>
                            <th>ល.រ</th>
                            <th>ឈ្មោះបុគ្គលិក</th>
                            <th>ថ្ងៃធ្វើការ</th>
                            <th>ម៉ោងធ្វើការសរុប</th>
                            <th>វគ្គយប់</th>
                            <th>ម៉ោងយប់</th>
                            <th>ចំនួនថ្ងៃសុំច្បាប់</th>
                            <th>ចំនួនថ្ងៃបេសកម្ម</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for item in summary %}
                        <tr>
                            <td>
                                {% if loop.index == 1 %}<span class="rank-badge rank-1">1</span>
                                {% elif loop.index == 2 %}<span class="rank-badge rank-2">2</span>
                                {% elif loop.index == 3 %}<span class="rank-badge rank-3">3</span>
                                {% else %}<span class="rank-badge rank-other">{{ loop.index }}</span>{% endif %}
                            </td>
                            <td>{{ item.full_name }}</td>
                            <td>{{ item.days_worked or 0 }}</td>
                            <td>
                                {% set hours = (item.total_hours or 0)|int %}
                                {% set minutes = (((item.total_hours or 0) - hours) * 60)|int %}
                                {{ '%02d' % hours }}:{{ '%02d' % minutes }}
                            </td>
                            <td>{{ item.night_shifts or 0 }}</td>
                            <td>
                                {% set nhours = (item.night_hours or 0)|int %}
                                {% set nminutes = (((item.night_hours or 0) - nhours) * 60)|int %}
                                {{ '%02d' % nhours }}:{{ '%02d' % nminutes }}
                            </td>
                            <td>{{ item.total_leave_days or 0 }}</td>
                            <td>{{ item.total_mission_days or 0 }}</td>
                        </tr>
                        {% endfor %}
                        {% if not summary %}
                        <tr>
                            <td colspan="8" style="text-align:center;padding:30px;color:#aaa;">📭 មិនមានទិន្នន័យ</td>
                        </tr>
                        {% endif %}
                    </tbody>
                </table>
            </div>

            <div class="stats-cards">
                <div class="stat-card">
                    <span class="stat-icon">👥</span>
                    <div class="number">{{ summary|length }}</div>
                    <div class="label">បុគ្គលិកសរុប</div>
                </div>
                <div class="stat-card">
                    <span class="stat-icon">📅</span>
                    <div class="number green">{{ summary|sum(attribute='days_worked') }}</div>
                    <div class="label">ថ្ងៃធ្វើការសរុប</div>
                </div>
                <div class="stat-card">
                    <span class="stat-icon">⏱️</span>
                    <div class="number orange">{{ summary|sum(attribute='total_hours')|round(1) }}</div>
                    <div class="label">ម៉ោងសរុប</div>
                </div>
                <div class="stat-card">
                    <span class="stat-icon">🌙</span>
                    <div class="number purple">{{ summary|sum(attribute='night_shifts') }}</div>
                    <div class="label">វគ្គយប់សរុប</div>
                </div>
            </div>
        </div>

        <div class="footer">© 2026 ប្រព័ន្ធគ្រប់គ្រងបុគ្គលិក | រក្សាសិទ្ធិគ្រប់យ៉ាង</div>
    </div>

    <script>
        function switchTab(tab) {
            var tabs = document.querySelectorAll('.tab-content');
            var buttons = document.querySelectorAll('.tab-button');

            for (var i = 0; i < tabs.length; i++) {
                tabs[i].classList.remove('active');
            }
            for (var i = 0; i < buttons.length; i++) {
                buttons[i].classList.remove('active');
            }

            document.getElementById('tab-' + tab).classList.add('active');

            for (var i = 0; i < buttons.length; i++) {
                if (buttons[i].getAttribute('onclick') === "switchTab('" + tab + "')") {
                    buttons[i].classList.add('active');
                }
            }
        }

        function filterReport() {
            var startDate = document.getElementById('startDate').value;
            var endDate = document.getElementById('endDate').value;

            if (!startDate || !endDate) {
                alert('សូមជ្រើសរើសថ្ងៃ!');
                return;
            }

            window.location.href = '/report?start=' + encodeURIComponent(startDate) + '&end=' + encodeURIComponent(endDate);
        }

        function loadMonthly() {
            var month = document.getElementById('monthSelect').value;
            var year = document.getElementById('yearSelect').value;

            if (!month) {
                alert('សូមជ្រើសរើសខែ!');
                return;
            }

            if (!year) {
                year = new Date().getFullYear();
            }

            window.location.href = '/report?month=' + encodeURIComponent(month) + '&year=' + encodeURIComponent(year);
        }

        function exportExcel() {
            var startDate = document.getElementById('startDate').value;
            var endDate = document.getElementById('endDate').value;
            var month = document.getElementById('monthSelect').value;
            var year = document.getElementById('yearSelect').value;

            var url = '/export_excel?';

            if (month && year) {
                url += 'month=' + encodeURIComponent(month) + '&year=' + encodeURIComponent(year);
            } else if (startDate && endDate) {
                url += 'start=' + encodeURIComponent(startDate) + '&end=' + encodeURIComponent(endDate);
            } else {
                alert('សូមជ្រើសរើសថ្ងៃខែមុនពេលទាញ Excel!');
                return;
            }

            window.open(url, '_blank');
        }

        function setDefaultDates() {
            var startDate = document.getElementById('startDate');
            var endDate = document.getElementById('endDate');

            if (startDate && !startDate.value) {
                var today = new Date();
                var firstDay = new Date(today.getFullYear(), today.getMonth(), 1);
                startDate.value = firstDay.toISOString().split('T')[0];
            }
            if (endDate && !endDate.value) {
                var today = new Date();
                endDate.value = today.toISOString().split('T')[0];
            }
        }

        document.addEventListener('DOMContentLoaded', function() {
            setDefaultDates();
        });

        window.onload = function() {
            setDefaultDates();
        };
    </script>
</body>
</html>'''

# ============================================================
# MAIN
# ============================================================

import os
import sys

# ហៅ init_db នៅពេល app ចាប់ផ្តើម
with app.app_context():
    try:
        init_db()
        print("✅ Database initialized successfully!")
    except Exception as e:
        print(f"❌ Database initialization error: {e}")
        sys.exit(1)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
